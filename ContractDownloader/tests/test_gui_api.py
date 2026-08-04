import os
from pathlib import Path
import pytest
import pandas as pd
from unittest.mock import MagicMock, AsyncMock, patch
from src.db.session_db import SessionDB
from src.gui.api import AppAPI

@pytest.fixture
def temp_db(tmp_path):
    db_path = tmp_path / "test_gui.db"
    db = SessionDB(str(db_path))
    yield db
    db.close()

@pytest.fixture
def dummy_excel(tmp_path):
    file_path = tmp_path / "po_input.xlsx"
    df = pd.DataFrame({
        'PO Number': ['PO-100', 'PO-101', 'PO-102'],
        'legal entity': ['COMP-A', 'COMP-B', 'COMP-A'],
        'Other Col': ['x', 'y', 'z']
    })
    df.to_excel(file_path, index=False)
    return str(file_path)

def test_import_file_excel(temp_db, dummy_excel):
    api = AppAPI(temp_db, "/tmp/downloads")
    result = api.import_file(dummy_excel)
    
    assert result['success'] is True
    assert result['session_id'] is not None
    assert result['total_pos'] == 3

    session = temp_db.get_session(result['session_id'])
    assert session['execution_type'] == 'PROD'
    
    # Verify POs are stored in DB
    stats = temp_db.get_company_stats(result['session_id'], "COMP-A")
    assert stats['total'] == 2

def test_get_session_history(temp_db):
    api = AppAPI(temp_db, "/tmp/downloads")
    
    # Create two sessions directly
    s1 = temp_db.create_session("file1.xlsx")
    s2 = temp_db.create_session("file2.xlsx")
    
    history = api.get_session_history()
    assert len(history) == 2
    assert history[0]['input_file'] == "file2.xlsx" # Should be reverse chronological
    assert history[1]['input_file'] == "file1.xlsx"
    assert history[0]['execution_type'] == "PROD"

def test_confirm_and_retry_company(temp_db):
    api = AppAPI(temp_db, "/tmp/downloads")
    session_id = temp_db.create_session("file.xlsx")
    
    from src.db.session_db import PODownload
    temp_db.add_po(PODownload(session_id, "PO-1", "COMP-A", "SKIPPED_VERIFICATION_REQUIRED"))
    temp_db.add_po(PODownload(session_id, "PO-2", "COMP-A", "SUCCESS"))
    
    result = api.confirm_and_retry_company(session_id, "COMP-A")
    assert result['success'] is True
    
    # PO-1 should be PENDING again
    po1 = temp_db.get_po(session_id, "PO-1")
    assert po1['status'] == "PENDING"
    
    # PO-2 should remain SUCCESS
    po2 = temp_db.get_po(session_id, "PO-2")
    assert po2['status'] == "SUCCESS"


def test_authentication_exposes_progress_states_without_blocking(temp_db, monkeypatch):
    import asyncio
    import time
    from src.main import TurboAPI

    async def fake_authenticate(*, load_from_file=True, fresh=False, status_callback=None):
        status_callback("starting", "Opening Edge and loading Coupa…")
        await asyncio.sleep(0.08)
        status_callback("user_action_required", "Complete the Coupa sign-in in the Edge window.")
        await asyncio.sleep(0.08)
        status_callback("validating", "Sign-in detected; validating the Coupa session…")
        return {"_coupa_session": "session-value"}

    monkeypatch.setattr("src.main.get_coupa_cookies", fake_authenticate)
    api = TurboAPI(temp_db, "/tmp/downloads")

    started = api.authenticate()
    assert started == {"success": True, "started": True}

    states = []
    deadline = time.monotonic() + 2
    while time.monotonic() < deadline:
        status = api.get_authentication_status()
        states.append(status["state"])
        if status["state"] == "success":
            break
        time.sleep(0.01)

    assert states[-1] == "success"
    assert api._cookies == {"_coupa_session": "session-value"}


def test_relative_download_path_is_resolved_under_user_home(temp_db, monkeypatch, tmp_path):
    monkeypatch.setattr("src.gui.api.Path.home", lambda: tmp_path)
    api = AppAPI(temp_db, "Downloads/CoupaAttachments")

    assert api.default_download_dir == str(tmp_path / "Downloads" / "CoupaAttachments")


def test_edge_driver_version_components_can_be_compared():
    from src.gui.api import AppAPI

    assert AppAPI._version_components("Microsoft Edge 140.0.3485.54") == (140, 0, 3485, 54)
    assert AppAPI._version_components("MSEdgeDriver 140.0.3485.54")[:3] == (140, 0, 3485)


def test_window_uses_88_percent_width_and_is_centered():
    from src.main import calculate_window_geometry

    geometry = calculate_window_geometry(1512, 982)

    assert geometry == {"width": 1331, "height": 840, "x": 90, "y": 71}


def test_window_minimum_width_gives_title_and_controls_room():
    from src.main import calculate_window_geometry

    small = calculate_window_geometry(900, 700)
    assert small["width"] >= 1080
    assert small["height"] >= 700


def test_macos_coupa_url_uses_configured_default_handler(temp_db, monkeypatch):
    from src.main import TurboAPI

    monkeypatch.setattr("src.main.sys.platform", "darwin")
    api = TurboAPI(temp_db, "/tmp/downloads")
    url = "https://unilever.coupahost.com/order_headers/17105916"
    with patch("src.main.subprocess.run") as run:
        result = api.open_external_url(url)

    assert result["success"] is True
    run.assert_called_once_with(
        ["/usr/bin/open", url],
        check=True,
        timeout=10,
    )


def test_open_coupa_po_strips_prefix_and_uses_default_browser(temp_db):
    from src.main import TurboAPI

    api = TurboAPI(temp_db, "/tmp/downloads")
    with patch.object(api, "open_external_url", return_value={"success": True}) as open_browser:
        result = api.open_coupa_po("PO17138914")

    assert result["success"] is True
    open_browser.assert_called_once_with("https://unilever.coupahost.com/order_headers/17138914")


def test_font_scale_setting_is_validated_and_persisted(temp_db, monkeypatch, tmp_path):
    monkeypatch.setattr("src.gui.api.Path.home", lambda: tmp_path)
    api = AppAPI(temp_db, "Downloads/CoupaAttachments")

    result = api.set_app_settings({
        "download_root": "Downloads/CoupaAttachments",
        "font_scale": 1.2,
    })

    assert result["success"] is True
    assert result["settings"]["font_scale"] == 1.2
    assert AppAPI(temp_db, "Downloads/CoupaAttachments").get_app_settings()["font_scale"] == 1.2

    clamped = api.set_app_settings({
        "download_root": "Downloads/CoupaAttachments",
        "font_scale": 9,
    })
    assert clamped["settings"]["font_scale"] == 1.3


def test_reset_new_run_deletes_only_app_generated_template(temp_db, monkeypatch, tmp_path):
    monkeypatch.setattr("src.gui.api.Path.home", lambda: tmp_path)
    template_dir = tmp_path / "Documents" / "Contract Downloader" / "templates"
    template_dir.mkdir(parents=True)
    generated = template_dir / "input_template_20260729-213140.xlsx"
    generated.write_bytes(b"template")
    external = tmp_path / "input.xlsx"
    external.write_bytes(b"original")
    api = AppAPI(temp_db, "Downloads/CoupaAttachments")

    deleted = api.reset_new_run(str(generated))
    preserved = api.reset_new_run(str(external))

    assert deleted == {"success": True, "deleted": True, "path": str(generated)}
    assert not generated.exists()
    assert preserved == {"success": True, "deleted": False, "preserved": True}
    assert external.exists()


def test_python_portable_disables_startup_update_checks_by_default(temp_db, monkeypatch, tmp_path):
    monkeypatch.setenv("COUPA_PYTHON_PORTABLE", "1")
    monkeypatch.setattr("src.gui.api.Path.home", lambda: tmp_path)

    settings = AppAPI(temp_db, "Downloads/CoupaAttachments").get_app_settings()

    assert settings["python_portable"] is True
    assert settings["auto_updates"] is False


def test_import_file_csv_hierarchy_output_subdir(temp_db, tmp_path):
    csv_path = tmp_path / "hierarchy.csv"
    csv_path.write_text(
        "PO_NUMBER;SUPPLIER;<|>;Year;Quarter;Management Unit L1 Name\n"
        "PO9001;Manpowergroup Inc.;;2026;Q12026;Yellow Wood\n",
        encoding="utf-8",
    )

    api = AppAPI(temp_db, "/tmp/downloads")
    result = api.import_file(str(csv_path))

    assert result["success"] is True
    session_id = result["session_id"]

    cur = temp_db.conn.cursor()
    row = cur.execute(
        "SELECT output_subdir FROM po_downloads WHERE session_id = ? AND po_number = ?",
        (session_id, "PO9001"),
    ).fetchone()
    assert row is not None
    assert row["output_subdir"] == "Manpowergroup_Inc/2026/Q12026/Yellow_Wood"


# ── Column mapping and grouped validation ────────────────────────────────


def test_validate_input_file_groups_errors_by_category(temp_db, tmp_path):
    csv_path = tmp_path / "dirty.csv"
    csv_path.write_text(
        "PO_NUMBER;SUPPLIER\n"
        "PO1;CompA\n"
        "PO1;CompA\n"
        "PO@2;CompB\n"
        ";CompC\n",
        encoding="utf-8",
    )
    api = AppAPI(temp_db, "/tmp/downloads")
    result = api.validate_input_file(str(csv_path))

    groups = {group["id"]: group for group in result["groups"]}
    assert groups["duplicate_pos"]["fix_action"] == "remove_duplicate_pos"
    assert groups["invalid_chars"]["fix_action"] == "clean_invalid_chars"
    assert "partial_rows" in groups


def test_validate_input_file_detects_blank_rows_in_xlsx(temp_db, tmp_path):
    xlsx_path = tmp_path / "blank_rows.xlsx"
    frame = pd.DataFrame({
        "PO_NUMBER": ["PO1", None, "PO2"],
        "SUPPLIER": ["CompA", None, "CompB"],
    })
    frame.to_excel(xlsx_path, index=False)
    api = AppAPI(temp_db, "/tmp/downloads")
    result = api.validate_input_file(str(xlsx_path))

    groups = {group["id"]: group for group in result["groups"]}
    assert groups["blank_rows"]["fix_action"] == "remove_blank_rows"
    assert result["valid"] is False


def test_validate_input_file_detects_blank_csv_lines_and_canonical_duplicates(temp_db, tmp_path):
    csv_path = tmp_path / "blank_lines.csv"
    csv_path.write_text("PO_NUMBER;SUPPLIER\nPO1;CompA\n\n PO1 ;CompA\n", encoding="utf-8")
    result = AppAPI(temp_db, "/tmp/downloads").validate_input_file(str(csv_path))

    groups = {group["id"]: group for group in result["groups"]}
    assert result["total_rows"] == 3
    assert groups["blank_rows"]["rows"] == [3]
    assert groups["duplicate_pos"]["rows"] == ["PO1"]


def test_validate_input_file_blocks_multiple_pos_in_one_cell(temp_db, tmp_path):
    csv_path = tmp_path / "multiple_pos.csv"
    csv_path.write_text("PO_NUMBER;SUPPLIER\nPO123 - PO456;CompA\n", encoding="utf-8")
    result = AppAPI(temp_db, "/tmp/downloads").validate_input_file(str(csv_path))
    groups = {group["id"]: group for group in result["groups"]}
    assert result["valid"] is False
    assert groups["multiple_pos_in_cell"]["rows"] == [2]
    assert groups["multiple_pos_in_cell"]["fix_action"] == "split_multiple_pos"


def test_validate_input_file_normalizes_required_whitespace(temp_db, tmp_path):
    csv_path = tmp_path / "whitespace.csv"
    csv_path.write_text("PO_NUMBER;SUPPLIER\n  PO1 ; CompA \n", encoding="utf-8")
    api = AppAPI(temp_db, "/tmp/downloads")

    result = api.validate_input_file(str(csv_path))
    groups = {group["id"]: group for group in result["groups"]}
    assert result["valid"] is True
    assert groups["required_value_whitespace"]["fix_action"] == "normalize_required_values"

    repaired = api.repair_input_file(str(csv_path), ["normalize_required_values"])
    assert repaired["success"] is True
    assert "  PO1 " not in csv_path.read_text(encoding="utf-8-sig")


def test_validate_input_file_blocks_supplier_conflicts_and_placeholders(temp_db, tmp_path):
    conflict = tmp_path / "conflict.csv"
    conflict.write_text("PO_NUMBER;SUPPLIER\nPO1;CompA\nPO1;CompB\n", encoding="utf-8")
    placeholder = tmp_path / "placeholder.csv"
    placeholder.write_text("PO_NUMBER;SUPPLIER\nUNK;CompA\n", encoding="utf-8")
    api = AppAPI(temp_db, "/tmp/downloads")

    conflict_result = api.validate_input_file(str(conflict))
    placeholder_result = api.validate_input_file(str(placeholder))

    assert "po_supplier_conflict" in {group["id"] for group in conflict_result["groups"]}
    assert not any(group.get("fixable") for group in conflict_result["groups"] if group["id"] == "po_supplier_conflict")
    assert "placeholder_pos" in {group["id"] for group in placeholder_result["groups"]}
    repair = api.repair_input_file(str(conflict), ["remove_duplicate_pos"])
    assert repair["success"] is False
    imported = api.import_file(str(conflict))
    assert imported["success"] is False


def test_validate_input_file_detects_excel_error_cells(temp_db, tmp_path):
    csv_path = tmp_path / "formula_error.csv"
    csv_path.write_text("PO_NUMBER;SUPPLIER;Year\nPO1;CompA;#REF!\n", encoding="utf-8")
    result = AppAPI(temp_db, "/tmp/downloads").validate_input_file(str(csv_path))
    assert "excel_cell_errors" in {group["id"] for group in result["groups"]}


def test_validate_input_file_suggests_mapping_from_column_values(temp_db, tmp_path):
    csv_path = tmp_path / "value_mapped.csv"
    csv_path.write_text("Document;Vendor Name\nPO100;CompA\nPO101;CompB\n", encoding="utf-8")
    result = AppAPI(temp_db, "/tmp/downloads").validate_input_file(str(csv_path))
    assert result["valid"] is True
    assert result["mapping"] == {"po": "Document", "supplier": "Vendor Name"}


def test_map_input_columns_rejects_same_required_column(temp_db, tmp_path):
    csv_path = tmp_path / "mapped.csv"
    csv_path.write_text("Document;Vendor\nPO1;CompA\n", encoding="utf-8")
    result = AppAPI(temp_db, "/tmp/downloads").map_input_columns(
        str(csv_path), {"po": "Document", "supplier": "Document"}
    )
    assert result["success"] is False
    assert result["mapping_conflict"] is True


def test_validate_download_directory_checks_writability(temp_db, tmp_path):
    api = AppAPI(temp_db, "/tmp/downloads")
    result = api.validate_download_directory(str(tmp_path / "new-run"))
    assert result["success"] is True
    assert result["path"] == str(tmp_path / "new-run")


def test_repair_preview_and_split_multiple_pos(temp_db, tmp_path):
    csv_path = tmp_path / "split.csv"
    csv_path.write_text("PO_NUMBER;SUPPLIER;Year\nPO123/PO456;CompA;2026\n", encoding="utf-8")
    api = AppAPI(temp_db, "/tmp/downloads")

    preview = api.preview_repair_input_file(str(csv_path), "split_multiple_pos")
    assert preview["success"] is True
    assert [change["new"] for change in preview["changes"]] == ["PO123", "PO456"]
    repaired = api.repair_input_file(str(csv_path), ["split_multiple_pos"], preview["expected_fingerprint"])
    assert repaired["success"] is True
    assert pd.read_csv(csv_path, sep=";") ["PO_NUMBER"].tolist() == ["PO123", "PO456"]


def test_repair_input_file_cleans_invalid_chars(temp_db, tmp_path):
    csv_path = tmp_path / "dirty.csv"
    csv_path.write_text(
        "PO_NUMBER;SUPPLIER\n"
        "PO@2;CompB\n"
        "PO3;CompA\n",
        encoding="utf-8",
    )
    api = AppAPI(temp_db, "/tmp/downloads")
    result = api.repair_input_file(str(csv_path), ["clean_invalid_chars"])

    assert result["success"] is True
    assert result["cleaned_invalid_chars"] == 1
    content = csv_path.read_text(encoding="utf-8-sig")
    assert "PO2" in content
    assert "PO@2" not in content


def test_repair_input_file_uses_persisted_column_mapping(temp_db, tmp_path):
    csv_path = tmp_path / "mapped_dirty.csv"
    csv_path.write_text(
        "Document;Vendor Name\n"
        "PO@2;CompB\n"
        "PO3;CompA\n",
        encoding="utf-8",
    )
    api = AppAPI(temp_db, "/tmp/downloads")
    mapped = api.map_input_columns(str(csv_path), {"po": "Document", "supplier": "Vendor Name"})
    assert mapped["success"] is True

    result = api.repair_input_file(str(csv_path), ["clean_invalid_chars"])

    assert result["success"] is True
    assert result["cleaned_invalid_chars"] == 1
    assert "PO2" in csv_path.read_text(encoding="utf-8-sig")


def test_map_input_columns_enables_nonstandard_file(temp_db, tmp_path):
    csv_path = tmp_path / "master_data.csv"
    csv_path.write_text(
        "Document;Vendor Name\n"
        "PO1234;Acme Corp\n"
        "PO5678;Globex\n",
        encoding="utf-8",
    )
    api = AppAPI(temp_db, "/tmp/downloads")
    info = api.get_input_columns(str(csv_path))
    assert info["success"] is True
    assert info["detected"]["po"] is None

    mapped = api.map_input_columns(str(csv_path), {"po": "Document", "supplier": "Vendor Name"})
    assert mapped["success"] is True
    assert mapped["mapping"]["po"] == "Document"

    validation = api.validate_input_file(str(csv_path))
    assert validation["valid"] is True
    assert validation["valid_po_count"] == 2

    imported = api.import_file(str(csv_path))
    assert imported["success"] is True
    assert imported["total_pos"] == 2


def test_import_file_deduplicates_canonical_po_values(temp_db, tmp_path):
    csv_path = tmp_path / "canonical.csv"
    csv_path.write_text("PO_NUMBER;SUPPLIER\nPO1;CompA\n PO1 ;CompA\n", encoding="utf-8")
    result = AppAPI(temp_db, "/tmp/downloads").import_file(str(csv_path))
    assert result["success"] is True
    assert result["total_pos"] == 1


def test_import_file_supplier_is_always_first_level(temp_db, tmp_path):
    csv_path = tmp_path / "hierarchy.csv"
    csv_path.write_text(
        "PO_NUMBER;SUPPLIER;<|>;Year\n"
        "PO9001;Manpowergroup Inc.;;2026\n",
        encoding="utf-8",
    )
    api = AppAPI(temp_db, "/tmp/downloads")
    result = api.import_file(str(csv_path))
    assert result["success"] is True
    row = temp_db.conn.execute(
        "SELECT output_subdir FROM po_downloads WHERE session_id = ? AND po_number = 'PO9001'",
        (result["session_id"],),
    ).fetchone()
    assert row["output_subdir"] == "Manpowergroup_Inc/2026"


def test_validate_input_file_reports_empty_hierarchy_columns(temp_db, tmp_path):
    csv_path = tmp_path / "empty_col.csv"
    csv_path.write_text(
        "PO_NUMBER;SUPPLIER;<|>;Year;Country\n"
        "PO1;CompA;;2026;\n",
        encoding="utf-8",
    )
    api = AppAPI(temp_db, "/tmp/downloads")
    result = api.validate_input_file(str(csv_path))
    assert "Country" in result["empty_hierarchy_columns"]
    assert "Year" not in result["empty_hierarchy_columns"]


def test_open_filtered_input_view_annotates_original_xlsx_with_backup(temp_db, monkeypatch, tmp_path):
    from openpyxl import load_workbook

    monkeypatch.setattr("src.gui.api.Path.home", lambda: tmp_path)
    xlsx_path = tmp_path / "input.xlsx"
    pd.DataFrame({
        "PO_NUMBER": ["PO00000001", "PO00000001", "PO00000002"],
        "SUPPLIER": ["CompA", "CompA", "CompB"],
    }).to_excel(xlsx_path, index=False)
    original = xlsx_path.read_bytes()
    api = AppAPI(temp_db, "/tmp/downloads")
    api._open_spreadsheet_path = lambda path: None

    result = api.open_filtered_input_view(str(xlsx_path))

    assert result["success"] is True
    assert result["path"] == str(xlsx_path)
    assert result["backup_path"]
    assert Path(result["backup_path"]).exists()
    assert xlsx_path.read_bytes() != original
    sheet = load_workbook(xlsx_path).active
    assert sheet.max_column == 2
    assert sheet.row_dimensions[2].hidden is not True
    assert sheet.row_dimensions[4].hidden is True
    assert sheet["A2"].comment is not None
    assert "Duplicate PO" in sheet["A2"].comment.text
    assert sheet.auto_filter.ref.endswith(f"{sheet.max_row}")


def test_open_filtered_input_view_converts_csv_to_annotated_xlsx_and_preserves_source(temp_db, monkeypatch, tmp_path):
    from openpyxl import load_workbook

    monkeypatch.setattr("src.gui.api.Path.home", lambda: tmp_path)
    csv_path = tmp_path / "input.csv"
    original = "PO_NUMBER;SUPPLIER\nPO@00000001;CompA\nPO00000002;CompB\n"
    csv_path.write_text(original, encoding="utf-8")
    api = AppAPI(temp_db, "/tmp/downloads")
    api._open_spreadsheet_path = lambda path: None

    result = api.open_filtered_input_view(str(csv_path))

    working_path = tmp_path / "input.xlsx"
    assert result["success"] is True
    assert result["path"] == str(working_path)
    assert result["original_path"] == str(csv_path)
    assert result["converted_from_csv"] == str(csv_path)
    assert csv_path.read_text(encoding="utf-8") == original
    assert result["backup_path"]
    assert Path(result["backup_path"]).exists()
    sheet = load_workbook(working_path).active
    assert sheet.row_dimensions[2].hidden is not True
    assert sheet.row_dimensions[3].hidden is True
    assert sheet["A2"].comment is not None
    assert sheet.auto_filter.filterColumn[0].colId == 0
    assert result["filtered_rows"] == 1


def test_validate_input_file_explains_folder_safety_rows(temp_db, tmp_path):
    csv_path = tmp_path / "folder-values.csv"
    csv_path.write_text(
        "PO_NUMBER;SUPPLIER;Year\n"
        "PO00000001;ACME/BR;2026\n"
        "PO00000002;ACME_BR;2026\n",
        encoding="utf-8",
    )
    result = AppAPI(temp_db, "/tmp/downloads").validate_input_file(str(csv_path))

    group = next(group for group in result["groups"] if group["id"] == "folder_value_safety")
    assert group["rows"] == [2, 3]
    assert {item["row"] for item in group["row_details"]} == {2, 3}
    assert group["severity"] == "warning"
    assert group["title"] == "Folder names will be sanitized"
    assert "download can continue" in group["message"]
    assert any("path separator" in item["reason"] for item in group["row_details"])
    assert any("replaced with '_'" in item["reason"] for item in group["row_details"])
    assert any("collapse" in item["reason"] for item in group["row_details"])


# ── Run description (audit context) ──────────────────────────────────────


def test_session_description_persisted(temp_db):
    db = SessionDB(temp_db.db_path)
    session_id = db.create_session("file.xlsx", description="Análise para a Prianca — POs 2026")
    session = db.get_session(session_id)
    assert session["description"] == "Análise para a Prianca — POs 2026"

    db.update_session_description(session_id, "Requerido por auditoria Q3")
    assert db.get_session(session_id)["description"] == "Requerido por auditoria Q3"


# ── Working copy for archived inputs (TurboAPI GUI bridge) ───────────────


def test_archived_input_gets_working_copy_instead_of_in_place_edit(temp_db, monkeypatch, tmp_path):
    """Reusing a previous run's archived input must copy it, never edit it."""
    from src.main import TurboAPI
    import sqlite3
    from pathlib import Path

    monkeypatch.setattr("src.gui.api.Path.home", lambda: tmp_path)
    archive = tmp_path / "run_1" / "input_source_1.csv"
    archive.parent.mkdir(parents=True)
    original_content = "PO_NUMBER;SUPPLIER\nPO1;CompA\n"
    archive.write_text(original_content, encoding="utf-8")

    supervisor_db = tmp_path / ".contract_downloader" / "cli_sessions.db"
    supervisor_db.parent.mkdir(parents=True)
    conn = sqlite3.connect(str(supervisor_db))
    conn.execute("CREATE TABLE sessions (id INTEGER PRIMARY KEY AUTOINCREMENT, input_file TEXT, input_file_path TEXT, status TEXT)")
    conn.execute(
        "INSERT INTO sessions (input_file, input_file_path, status) VALUES ('input_source_1.csv', ?, 'SUCCESS')",
        (str(archive),),
    )
    conn.commit()
    conn.close()

    api = TurboAPI(temp_db, "/tmp/downloads")
    api.cli_backend.db_path = supervisor_db

    working, copied = api._working_copy_for(str(archive))

    assert copied is True
    assert working != str(archive)
    assert Path(working).exists()
    assert archive.read_text(encoding="utf-8") == original_content
    assert Path(working).read_text(encoding="utf-8") == original_content


def test_regular_input_is_not_copied(temp_db, monkeypatch, tmp_path):
    from src.main import TurboAPI

    monkeypatch.setattr("src.gui.api.Path.home", lambda: tmp_path)
    source = tmp_path / "my_input.csv"
    source.write_text("PO_NUMBER;SUPPLIER\nPO1;CompA\n", encoding="utf-8")
    api = TurboAPI(temp_db, "/tmp/downloads")

    working, copied = api._working_copy_for(str(source))

    assert copied is False
    assert working == str(source)
