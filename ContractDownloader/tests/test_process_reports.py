import pytest
from openpyxl import load_workbook

from process_all_pos import export_original_like_excel_report, read_input_dataframe
from src.db.session_db import PODownload, SessionDB


@pytest.fixture
def temp_db(tmp_path):
    db = SessionDB(str(tmp_path / "sessions.db"))
    yield db
    db.close()


def test_read_input_dataframe_supports_excel_and_cp1252_csv(tmp_path):
    excel_path = tmp_path / "input.xlsx"
    import pandas as pd
    pd.DataFrame({"PO_NUMBER": ["PO-1"], "SUPPLIER": ["Companhia São Paulo"]}).to_excel(excel_path, index=False)
    assert list(read_input_dataframe(str(excel_path))["PO_NUMBER"]) == ["PO-1"]

    csv_path = tmp_path / "input.csv"
    csv_path.write_bytes("PO_NUMBER;SUPPLIER\nPO-2;Companhia São Paulo\n".encode("cp1252"))
    frame = read_input_dataframe(str(csv_path))
    assert frame.iloc[0]["SUPPLIER"] == "Companhia São Paulo"


def test_report_preserves_input_columns_and_updates_retry_result(temp_db, tmp_path):
    input_path = tmp_path / "input.csv"
    input_path.write_text(
        "PO_NUMBER;SUPPLIER;<|>;Year;Management Unit\n"
        "PO-1;Supplier A;;2026;Foods\n"
        "PO-2;Supplier B;;2026;Finance\n",
        encoding="utf-8",
    )
    session_id = temp_db.create_session(input_path.name)
    temp_db.add_po(PODownload(session_id, "PO-1", "Supplier A", status="PENDING"))
    temp_db.add_po(PODownload(session_id, "PO-2", "Supplier B", status="PENDING"))
    temp_db.update_po_status(
        session_id,
        "PO-1",
        "ERROR",
        attachment_count=0,
        error_message="First attempt failed",
    )
    temp_db.update_po_status(session_id, "PO-2", "SUCCESS", attachment_count=2)

    report_path = tmp_path / "run" / f"report_session_{session_id}.xlsx"
    export_original_like_excel_report(temp_db, session_id, str(report_path), str(input_path))
    headers = list(next(load_workbook(report_path, read_only=True).active.iter_rows(values_only=True)))
    assert headers[:5] == ["PO_NUMBER", "SUPPLIER", "<|>", "Year", "Management Unit"]
    assert "STATUS" in headers
    assert "LAST_PROCESSED" in headers
    assert "ERROR_MESSAGE" in headers

    first_timestamp = temp_db.get_po(session_id, "PO-1")["updated_at"]
    temp_db.update_po_status(
        session_id,
        "PO-1",
        "ERROR",
        attachment_count=0,
        error_message="Retry failed again",
    )
    second_timestamp = temp_db.get_po(session_id, "PO-1")["updated_at"]
    assert second_timestamp >= first_timestamp

    export_original_like_excel_report(temp_db, session_id, str(report_path), str(input_path))
    rows = list(load_workbook(report_path, read_only=True).active.iter_rows(values_only=True))
    po_index = headers.index("PO_NUMBER")
    status_index = headers.index("STATUS")
    timestamp_index = headers.index("LAST_PROCESSED")
    error_index = headers.index("ERROR_MESSAGE")
    retry_row = next(row for row in rows[1:] if row[po_index] == "PO-1")
    assert retry_row[status_index] == "ERROR"
    assert retry_row[error_index] == "Retry failed again"
    assert retry_row[timestamp_index] == second_timestamp
