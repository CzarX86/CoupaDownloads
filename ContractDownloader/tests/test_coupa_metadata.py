from openpyxl import load_workbook
import pandas as pd

from src.db.coupa_metadata import CoupaMetadataRepository
from src.engine.coupa_metadata import CoupaLineMetadata, CoupaMetadataExtractor, CoupaPOMetadata
from src.reports.coupa_excel import enrich_excel_report
from src.db.session_db import SessionDB


HTML = """
<html>
  <div id="topForm">
    <div class="form_element" role="group">
      <span class="group_label">Ship To User</span> Shehara Gamage
    </div>
    <div class="form_element" role="group">
      <span class="group_label">Payment Term</span><span class="data">P090</span>
    </div>
  </div>
  <section id="shipTo">
    <div class="form_element wrapper_vertical">
      <label class="group_label">Address</label>
      <span class="address">London
        <div class="form_element wrapper_vertical">
          <label class="group_label">Company Code</label> 5069
        </div>
      </span>
    </div>
  </section>
  <table>
    <tr id="order_line_row_123">
      <td>
        <div class="line_num">1</div>
        <div class="form_element orderLinePrice">
          <span class="group_label">Price</span><span title="EUR">42,303.00</span>
        </div>
        <div class="account-wrapper">
          <div class="form_element account-content">
            <span class="group_label">Account</span>
            <span class="accounts_show">
              <ul><li><span class="bold">5069-Unilever IT</span></li></ul>
            </span>
            <ul class="billing_tooltip">
              <li>CommCode: Technology Consulting (UN01540302)</li>
              <li>CC FuncA: R5600-General Overheads</li>
              <li>CC FncID: RFNC50190-Technology</li>
            </ul>
          </div>
        </div>
      </td>
    </tr>
  </table>
</html>
"""


def test_coupa_metadata_extractor_reads_header_line_and_account_tooltip():
    metadata = CoupaMetadataExtractor.extract(HTML, "PO12345678", "https://coupa/order_headers/12345678")

    assert metadata.metadata_status == "EXTRACTED"
    assert metadata.company_code == "5069"
    assert metadata.ship_to_user == "Shehara Gamage"
    assert metadata.payment_term == "P090"
    assert metadata.source_url.endswith("12345678")
    assert metadata.lines == (
        CoupaLineMetadata(
            line_number=1,
            crg_code="5600",
            crg_raw="R5600-General Overheads",
            currency_code="EUR",
        ),
    )


def test_coupa_metadata_repository_round_trip(tmp_path):
    db = SessionDB(str(tmp_path / "sessions.db"))
    try:
        repository = CoupaMetadataRepository(db)
        repository.save(
            4,
            CoupaPOMetadata(
                po_number="PO1",
                company_code="5069",
                ship_to_user="User",
                payment_term="P090",
                lines=(CoupaLineMetadata(1, "5600", "R5600-General Overheads", "EUR"),),
            ),
        )

        assert repository.get_po_metadata(4, "PO1")["company_code"] == "5069"
        assert repository.get_line_metadata(4, "PO1")[0]["currency_code"] == "EUR"
    finally:
        db.close()


def test_enrich_excel_report_writes_summary_and_lines_sheet(tmp_path):
    report_path = tmp_path / "report.xlsx"
    pd.DataFrame({"PO_NUMBER": ["PO1"], "SUPPLIER": ["Input supplier"]}).to_excel(report_path, index=False)

    enrich_excel_report(
        report_path,
        [{
            "po_number": "PO1",
            "company_code": "5069",
            "ship_to_user": "User",
            "payment_term": "P090",
            "metadata_status": "EXTRACTED",
            "scraped_at": "2026-08-03T23:00:00+00:00",
        }],
        [{
            "po_number": "PO1",
            "line_number": 1,
            "crg_code": "5600",
            "crg_raw": "R5600-General Overheads",
            "currency_code": "EUR",
        }],
    )

    workbook = load_workbook(report_path, read_only=True)
    summary_headers = list(next(workbook[workbook.sheetnames[0]].iter_rows(values_only=True)))
    summary_row = list(next(workbook[workbook.sheetnames[0]].iter_rows(min_row=2, values_only=True)))
    assert "COUPA_COMPANY_CODE" in summary_headers
    assert summary_row[summary_headers.index("COUPA_COMPANY_CODE")] == "5069"
    assert summary_row[summary_headers.index("COUPA_CRG_CODES")] == "5600"
    assert summary_row[summary_headers.index("COUPA_CURRENCY_CODES")] == "EUR"

    lines = workbook["COUPA_LINES"]
    assert list(lines.iter_rows(values_only=True)) == [
        ("PO_NUMBER", "LINE_NUMBER", "COUPA_CRG_CODE", "COUPA_CRG_RAW", "COUPA_CURRENCY_CODE"),
        ("PO1", 1, "5600", "R5600-General Overheads", "EUR"),
    ]
