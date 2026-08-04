import json
import os
import sys
import time
import asyncio
import platform
import plistlib
import shutil
import subprocess
import threading
from datetime import datetime, timezone
from pathlib import Path, PurePosixPath
from typing import Dict, Any, List, Optional

import pandas as pd

from src.db.session_db import SessionDB, PODownload
from src.engine.crawler import CoupaCrawler
from src.engine.tls import system_ssl_context
from src.auth import AuthService, AuthState
from src.auth.browser import BrowserKind


class AppAPI:
    DEFAULT_SETTINGS = {
        "concurrency": 4,
        "retry_attempts": 1,
        "auto_updates": True,
        "retention": "all",
        "msg_processing": "convert_extract",
        "deduplicate_files": True,
        "language": "en",
        "font_scale": 1.1,
        "auth_browser": "auto",
    }

    def __init__(self, db: SessionDB, default_download_dir: str):
        self.db = db
        self._settings_path = Path.home() / ".contract_downloader" / "gui_settings.json"
        self.default_download_dir = self._load_download_root(default_download_dir)
        self._runtime_lock = threading.Lock()
        self.auth_service = AuthService()
        self._cookies: Optional[Dict[str, str]] = None
        self._runtime: Dict[int, Dict[str, Any]] = {}
        self._file_observations: Dict[str, tuple[tuple[int, int], float]] = {}

    def _set_session_status(self, session_id: int, status: str):
        cursor = self.db.conn.cursor()
        cursor.execute(
            "UPDATE sessions SET status = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (status, session_id),
        )
        self.db.conn.commit()

    @staticmethod
    def _is_python_portable() -> bool:
        return os.environ.get("COUPA_PYTHON_PORTABLE") == "1"

    def _read_settings(self) -> Dict[str, Any]:
        settings = dict(self.DEFAULT_SETTINGS)
        if self._is_python_portable():
            settings["auto_updates"] = False
        try:
            stored = json.loads(self._settings_path.read_text(encoding="utf-8"))
            if isinstance(stored, dict):
                settings.update(stored)
        except (OSError, ValueError):
            pass
        settings["concurrency"] = max(1, min(8, int(settings.get("concurrency", 4))))
        settings["retry_attempts"] = max(1, min(3, int(settings.get("retry_attempts", 1))))
        settings["auto_updates"] = bool(settings.get("auto_updates", True))
        if settings.get("retention") not in {"all", "10", "30", "90"}:
            settings["retention"] = "all"
        if settings.get("msg_processing") not in {"disabled", "convert", "convert_extract"}:
            settings["msg_processing"] = "convert_extract"
        settings["deduplicate_files"] = bool(settings.get("deduplicate_files", True))
        settings["auth_browser"] = str(settings.get("auth_browser", "auto")).strip().lower()
        if settings["auth_browser"] not in {"auto", "edge", "chrome"}:
            settings["auth_browser"] = "auto"
        settings["language"] = "pt-BR" if settings.get("language") == "pt-BR" else "en"
        try:
            settings["font_scale"] = max(1.0, min(1.3, float(settings.get("font_scale", 1.1))))
        except (TypeError, ValueError):
            settings["font_scale"] = 1.1
        return settings

    @staticmethod
    def _absolute_user_path(value: str) -> str:
        path = Path(str(value or "")).expanduser()
        if not path.is_absolute():
            path = Path.home() / path
        return str(path.resolve(strict=False))

    def _load_download_root(self, fallback: str) -> str:
        try:
            settings = self._read_settings()
            value = str(settings.get("download_root", "")).strip()
            if value:
                return self._absolute_user_path(value)
        except (OSError, ValueError):
            pass
        return self._absolute_user_path(fallback)

    def _persist_download_root(self, value: str) -> str:
        root = self._absolute_user_path(value)
        try:
            self._settings_path.parent.mkdir(parents=True, exist_ok=True)
            settings = {}
            if self._settings_path.exists():
                settings = json.loads(self._settings_path.read_text(encoding="utf-8"))
            settings["download_root"] = root
            self._settings_path.write_text(json.dumps(settings, indent=2), encoding="utf-8")
        except (OSError, ValueError):
            pass
        self.default_download_dir = root
        return root

    # ------------------------------------------------------------------
    # Column mapping persistence (per input file)
    # ------------------------------------------------------------------
    def _column_mappings_path(self):
        return self._settings_path.parent / "column_mappings.json"

    def _load_column_mappings(self) -> Dict[str, Dict[str, str]]:
        try:
            data = json.loads(self._column_mappings_path().read_text(encoding="utf-8"))
            return data if isinstance(data, dict) else {}
        except (OSError, ValueError):
            return {}

    def _save_column_mappings(self, mappings: Dict[str, Dict[str, str]]) -> None:
        try:
            self._column_mappings_path().parent.mkdir(parents=True, exist_ok=True)
            self._column_mappings_path().write_text(json.dumps(mappings, indent=2), encoding="utf-8")
        except OSError:
            pass

    def _mapping_for(self, filepath: str) -> Dict[str, str]:
        key = str(Path(filepath).expanduser().resolve())
        stored = self._load_column_mappings().get(key, {})
        return {k: v for k, v in stored.items() if v}

    @staticmethod
    def _input_encoding(path: Path) -> str:
        if path.suffix.lower() != ".csv":
            return "utf-8-sig"
        sample = path.read_bytes()[:4096]
        try:
            sample.decode("utf-8-sig")
            return "utf-8-sig"
        except UnicodeDecodeError:
            return "cp1252"

    @classmethod
    def _read_input_dataframe(cls, path: Path, nrows: Optional[int] = None) -> pd.DataFrame:
        from src.engine.input_schema import detect_csv_separator

        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xls", ".xlsm"}:
            return pd.read_excel(path, sheet_name=0, dtype=str, nrows=nrows)
        encoding = cls._input_encoding(path)
        raw = path.read_bytes()[:8192].decode(encoding, errors="replace")
        separator = detect_csv_separator(raw)
        return pd.read_csv(
            path,
            sep=separator,
            dtype=str,
            encoding=encoding,
            nrows=nrows,
            skip_blank_lines=False,
        )

    def get_input_columns(self, filepath: str) -> Dict[str, Any]:
        """Return the input headers plus auto-detected PO/Supplier columns."""
        from src.engine.input_schema import columns_of_dataframe, detect_required_columns, resolve_data_mapping

        path = Path(filepath)
        if not path.exists():
            return {"success": False, "error": f"File not found: {filepath}"}
        try:
            frame = self._read_input_dataframe(path, nrows=5)
        except Exception as exc:
            return {"success": False, "error": f"Could not read the input file: {exc}"}
        columns = columns_of_dataframe(frame)
        suggested_mapping, suggestions = resolve_data_mapping(frame, columns, self._mapping_for(str(path)))
        return {
            "success": True,
            "columns": columns,
            "detected": detect_required_columns(columns),
            "suggestions": suggestions,
            "suggested_mapping": suggested_mapping,
            "mapping": self._mapping_for(str(path)),
        }

    def map_input_columns(self, filepath: str, mapping: Dict[str, Any]) -> Dict[str, Any]:
        """Store an explicit PO/Supplier column mapping and re-validate."""
        from src.engine.input_schema import REQUIRED_FIELDS, normalize_column_name, resolve_mapping

        path = Path(filepath)
        if not path.exists():
            return {"success": False, "error": f"File not found: {filepath}"}
        columns = self.get_input_columns(str(path)).get("columns", [])
        requested = {str(key).lower(): str(value or "").strip() for key, value in (mapping or {}).items()}
        resolved = resolve_mapping(columns, requested)
        missing = [field for field in REQUIRED_FIELDS if not resolved.get(field)]
        if missing:
            return {
                "success": False,
                "error": f"Mapping is incomplete. Required fields: {', '.join(REQUIRED_FIELDS)}.",
                "missing": missing,
            }
        if normalize_column_name(resolved["po"]) == normalize_column_name(resolved["supplier"]):
            return {
                "success": False,
                "error": "PO Number and Supplier must be mapped to different columns.",
                "mapping_conflict": True,
            }
        stored = self._load_column_mappings()
        key = str(path.expanduser().resolve())
        stored[key] = {"po": resolved["po"], "supplier": resolved["supplier"]}
        self._save_column_mappings(stored)
        validation = self.validate_input_file(str(path))
        return {"success": True, "mapping": stored[key], **validation}

    def get_app_settings(self) -> Dict[str, Any]:
        settings = self._read_settings()
        settings["download_root"] = self.default_download_dir
        settings["python_portable"] = self._is_python_portable()
        settings["auth_browsers"] = self.auth_service.browser_options(settings.get("auth_browser"))
        return settings

    def set_app_settings(self, values: Dict[str, Any]) -> Dict[str, Any]:
        current = self._read_settings()
        try:
            current["concurrency"] = max(1, min(8, int(values.get("concurrency", current["concurrency"]))))
            current["retry_attempts"] = max(1, min(3, int(values.get("retry_attempts", current["retry_attempts"]))))
        except (TypeError, ValueError):
            return {"success": False, "error": "Concurrency and retry values must be numeric."}
        current["auto_updates"] = bool(values.get("auto_updates", current["auto_updates"]))
        retention = str(values.get("retention", current["retention"]))
        if retention not in {"all", "10", "30", "90"}:
            return {"success": False, "error": "Invalid history retention option."}
        current["retention"] = retention
        msg_processing = str(values.get("msg_processing", current["msg_processing"]))
        if msg_processing not in {"disabled", "convert", "convert_extract"}:
            return {"success": False, "error": "Invalid MSG processing option."}
        current["msg_processing"] = msg_processing
        current["deduplicate_files"] = bool(values.get("deduplicate_files", current["deduplicate_files"]))
        auth_browser = str(values.get("auth_browser", current["auth_browser"])).strip().lower()
        if auth_browser not in {"auto", "edge", "chrome"}:
            return {"success": False, "error": "Invalid Coupa sign-in browser option."}
        current["auth_browser"] = auth_browser
        current["language"] = "pt-BR" if values.get("language") == "pt-BR" else "en"
        try:
            current["font_scale"] = max(1.0, min(1.3, float(values.get("font_scale", current["font_scale"]))))
        except (TypeError, ValueError):
            return {"success": False, "error": "Text size must be numeric."}
        root = str(values.get("download_root", self.default_download_dir)).strip()
        if not root:
            return {"success": False, "error": "Download folder cannot be empty."}
        current["download_root"] = self._absolute_user_path(root)
        try:
            self._settings_path.parent.mkdir(parents=True, exist_ok=True)
            self._settings_path.write_text(json.dumps(current, indent=2), encoding="utf-8")
        except OSError as exc:
            return {"success": False, "error": f"Could not save settings: {exc}"}
        self.default_download_dir = current["download_root"]
        return {"success": True, "settings": current}

    def _new_run_directory(self, root: str):
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        return str(Path(root).expanduser() / f"run_{timestamp}")

    def get_default_download_directory(self) -> str:
        return self._new_run_directory(self.default_download_dir)

    def validate_download_directory(self, value: str) -> Dict[str, Any]:
        """Check that a destination exists or can be created before step 5."""
        raw = str(value or "").strip()
        if not raw:
            return {"success": False, "error": "Choose a download folder first."}
        path = Path(self._absolute_user_path(raw))
        try:
            if path.exists():
                if not path.is_dir():
                    return {"success": False, "error": "The download destination is not a folder."}
                if not os.access(path, os.W_OK):
                    return {"success": False, "error": "The download destination is not writable."}
            else:
                parent = path.parent
                while not parent.exists() and parent != parent.parent:
                    parent = parent.parent
                if not parent.exists() or not os.access(parent, os.W_OK):
                    return {"success": False, "error": "The download destination cannot be created or is not writable."}
            return {"success": True, "path": str(path)}
        except OSError as exc:
            return {"success": False, "error": f"Could not access the download destination: {exc}"}

    def set_default_download_directory(self, value: str) -> Dict[str, Any]:
        if not str(value or "").strip():
            return {"success": False, "error": "Download folder cannot be empty."}
        root = self._persist_download_root(value)
        return {"success": True, "path": root}

    @staticmethod
    def _native_window():
        """Return the active pywebview window when running inside the desktop app."""
        try:
            import webview
            windows = getattr(webview, "windows", [])
            return windows[0] if windows else None
        except Exception:
            return None

    @staticmethod
    def _open_path(path: Path) -> None:
        """Open a file with the platform default application."""
        if sys.platform.startswith("win"):
            os.startfile(str(path))
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(path)])
        else:
            subprocess.Popen(["xdg-open", str(path)])

    def _open_spreadsheet_path(self, path: Path) -> None:
        """Prefer Microsoft Excel when available, then use the OS default."""
        if sys.platform == "darwin":
            try:
                subprocess.run(
                    ["open", "-a", "Microsoft Excel", str(path)],
                    check=True,
                    capture_output=True,
                    timeout=10,
                )
                return
            except (OSError, subprocess.SubprocessError):
                pass
        self._open_path(path)

    def open_input_path(self, filepath: str) -> Dict[str, Any]:
        """Open a user-selected input file in the platform default editor."""
        path = Path(str(filepath or "")).expanduser().resolve()
        if not path.is_file():
            return {"success": False, "error": "The selected input file could not be found."}
        try:
            self._open_path(path)
            return {"success": True, "path": str(path)}
        except (OSError, ValueError) as exc:
            return {"success": False, "error": f"Could not open the input file: {exc}"}

    def open_filtered_input_view(self, filepath: str) -> Dict[str, Any]:
        """Open the selected input with validation annotations applied in place.

        XLSX/XLSM files receive a safety backup, cell highlights, comments and
        hidden clean rows. CSV files are converted to a sibling XLSX working
        copy; the original CSV is never overwritten. Legacy XLS files remain
        read-only because they cannot be safely annotated with openpyxl.
        """
        path = Path(str(filepath or "")).expanduser().resolve()
        if not path.is_file():
            return {"success": False, "error": "The selected input file could not be found."}

        validation = self.validate_input_file(str(path))
        if validation.get("errors") and not validation.get("groups"):
            return {
                "success": False,
                "error": "The validation view could not be built because the file could not be mapped to rows.",
                "validation": validation,
            }

        try:
            import re
            from openpyxl import load_workbook
            from openpyxl.comments import Comment
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
            from src.engine.input_schema import detect_csv_separator, normalize_po_value, resolve_mapping

            source_path = path
            extension = path.suffix.lower()
            timestamp = time.strftime("%Y%m%d-%H%M%S")
            backup_path: Optional[Path] = None
            converted_from_csv = False

            if extension == ".csv":
                # Keep the received CSV as the source of record. The XLSX is a
                # working copy that can carry comments, colors and real Excel
                # filter criteria without silently destroying the CSV.
                encoding = self._input_encoding(path)
                raw = path.read_bytes()[:8192].decode(encoding, errors="replace")
                separator = detect_csv_separator(raw)
                frame = pd.read_csv(
                    path,
                    sep=separator,
                    dtype=str,
                    encoding=encoding,
                    skip_blank_lines=False,
                )
                working_path = path.with_suffix(".xlsx")
                if working_path.exists():
                    working_path = path.with_name(f"{path.stem}.validation-{timestamp}.xlsx")
                frame.to_excel(working_path, index=False)
                path = working_path
                extension = ".xlsx"
                converted_from_csv = True

            # Build issue indexes using the same one-based spreadsheet row
            # numbers returned by validate_input_file.
            row_issues: Dict[int, list[tuple[str, str]]] = {}
            po_issues: Dict[str, list[tuple[str, str]]] = {}
            global_issues: list[tuple[str, str]] = []
            groups = validation.get("groups") or []

            def add_issue(target: Dict[int, list[tuple[str, str]]], row: int, severity: str, label: str) -> None:
                if row < 2:
                    return
                entries = target.setdefault(row, [])
                item = (severity, label)
                if item not in entries:
                    entries.append(item)

            def add_po_issue(value: Any, severity: str, label: str) -> None:
                key = normalize_po_value(value)
                if not key or key in {"nan", "none"}:
                    return
                entries = po_issues.setdefault(key, [])
                item = (severity, label)
                if item not in entries:
                    entries.append(item)

            row_group_ids = {
                "blank_rows", "partial_rows", "excel_cell_errors",
                "required_value_whitespace", "placeholder_supplier",
                "multiple_pos_in_cell", "ambiguous_po_value", "folder_value_safety",
            }
            mapped_group_ids: set[str] = set()
            for group in groups:
                group_id = str(group.get("id") or "")
                severity = "warning" if group.get("severity") == "warning" else "error"
                label = str(group.get("title") or group_id or "Validation issue")
                row_details: Dict[int, list[Dict[str, Any]]] = {}
                for item in group.get("row_details") or []:
                    if item.get("row") is not None:
                        row_details.setdefault(int(item["row"]), []).append(item)
                values = group.get("rows") or []
                if values:
                    mapped_group_ids.add(group_id)
                for value in values:
                    if group_id in row_group_ids:
                        try:
                            row_number = int(value)
                            details = row_details.get(row_number, [])
                            row_label = label
                            detail_parts = [part for detail in details for part in detail.get("parts", [])]
                            if detail_parts:
                                row_label = f"{label}: {', '.join(detail_parts)}"
                            add_issue(row_issues, row_number, severity, row_label)
                        except (TypeError, ValueError):
                            pass
                    else:
                        add_po_issue(value, severity, label)

            empty_hierarchy = validation.get("empty_hierarchy_columns") or []
            if empty_hierarchy:
                global_issues.extend(
                    ("warning", f"Empty hierarchy column: {column}")
                    for column in empty_hierarchy
                )

            # Missing columns and file-level errors cannot be tied to one PO;
            # mark all data rows so the in-place validation view remains actionable.
            for group in groups:
                if str(group.get("id") or "") not in mapped_group_ids and not group.get("rows"):
                    severity = "warning" if group.get("severity") == "warning" else "error"
                    if severity != "error":
                        continue
                    label = str(group.get("title") or group.get("id") or "Validation issue")
                    global_issues.append((severity, label))

            # Only truly file-level failures become global. Row and PO groups
            # already carry an explicit target and must never be inferred from
            # English message fragments.
            if not groups:
                global_issues.extend(("error", str(message)) for message in validation.get("errors") or [])
                global_issues.extend(("warning", str(message)) for message in validation.get("warnings") or [])
            elif any("No valid PO entries" in str(message) for message in validation.get("errors") or []):
                global_issues.append(("error", "No valid PO entries found after cleaning."))

            def row_metadata(row_number: int, po_value: Any = "") -> tuple[str, str]:
                entries = list(row_issues.get(row_number, []))
                po_key = normalize_po_value(po_value)
                for item in po_issues.get(po_key, []):
                    if item not in entries:
                        entries.append(item)
                for item in global_issues:
                    if item not in entries:
                        entries.append(item)
                status = "ERROR" if any(severity == "error" for severity, _ in entries) else "WARNING" if entries else "OK"
                labels = "; ".join(label for _, label in entries)
                return status, labels

            def apply_filter(workbook, sheet, data_rows: list[tuple[int, Any]], headers: list[str], po_index: Optional[int]) -> dict[str, int]:
                error_fill = PatternFill("solid", fgColor="F4CCCC")
                warning_fill = PatternFill("solid", fgColor="FFF2CC")
                marker = "[Contract Downloader validation]"
                counts = {"ERROR": 0, "WARNING": 0, "OK": 0}
                issue_values: list[str] = []
                for row_number in range(2, sheet.max_row + 1):
                    sheet.row_dimensions[row_number].hidden = False
                    for cell in sheet[row_number]:
                        if cell.comment and cell.comment.text.startswith(marker):
                            cell.comment = None
                            cell.fill = PatternFill(fill_type=None)

                for row_number, po_value in data_rows:
                    status, issues = row_metadata(row_number, po_value)
                    counts[status] += 1
                    sheet.row_dimensions[row_number].hidden = status == "OK"
                    if status == "OK":
                        continue
                    if po_index is not None:
                        cell_value = sheet.cell(row_number, po_index + 1).value
                        if cell_value not in (None, ""):
                            issue_values.append(str(cell_value))
                    fill = error_fill if status == "ERROR" else warning_fill
                    for cell in sheet[row_number]:
                        cell.fill = fill
                    target_index = (po_index + 1) if po_index is not None else 1
                    target = sheet.cell(row_number, target_index)
                    target.comment = Comment(f"{marker}\n{status}: {issues or 'Validation issue'}", "Contract Downloader")

                sheet.auto_filter.ref = f"A1:{get_column_letter(max(sheet.max_column, 1))}{max(sheet.max_row, 1)}"
                sheet.auto_filter.filterColumn = []
                unique_issue_values = list(dict.fromkeys(issue_values))
                if po_index is not None and unique_issue_values:
                    sheet.auto_filter.add_filter_column(po_index, unique_issue_values)
                sheet.sheet_properties.filterMode = bool(unique_issue_values)
                if sheet.freeze_panes is None:
                    sheet.freeze_panes = "A2"
                workbook.active = workbook.index(sheet)
                return counts

            if extension in {".xlsx", ".xlsm"}:
                backup_path = path.with_name(f"{path.stem}.validation-backup-{timestamp}{path.suffix}")
                shutil.copy2(path, backup_path)
                workbook = load_workbook(path, keep_vba=extension == ".xlsm")
                sheet = workbook.worksheets[0]
                headers = [str(sheet.cell(1, column).value or "") for column in range(1, sheet.max_column + 1)]
                mapping = validation.get("mapping") or resolve_mapping(headers, self._mapping_for(str(path)))
                po_column = mapping.get("po")
                po_index = next((index for index, value in enumerate(headers) if value == po_column), None)
                data_rows = []
                for row_number in range(2, sheet.max_row + 1):
                    po_value = sheet.cell(row_number, po_index + 1).value if po_index is not None else ""
                    data_rows.append((row_number, po_value))
                counts = apply_filter(workbook, sheet, data_rows, headers, po_index)
                workbook.save(path)
            else:
                # Legacy XLS cannot persist these annotations safely.
                frame = pd.read_excel(path, sheet_name=0, dtype=str)
                data_rows = [(row_number, frame.iloc[row_number - 2].get((validation.get("mapping") or {}).get("po"), "")) for row_number in range(2, len(frame) + 2)]
                counts = {"ERROR": 0, "WARNING": 0, "OK": 0}
                for row_number, po_value in data_rows:
                    status, _ = row_metadata(row_number, po_value)
                    counts[status] += 1

            try:
                self._open_spreadsheet_path(path)
            except (OSError, ValueError) as exc:
                return {"success": False, "error": f"Input was prepared but could not be opened: {exc}", "path": str(path)}
            conversion_message = (
                f" CSV preserved at {source_path}; working XLSX created at {path}."
                if converted_from_csv else ""
            )
            return {
                "success": True,
                "path": str(path),
                "original_path": str(source_path),
                "converted_from_csv": str(source_path) if converted_from_csv else None,
                "backup_path": str(backup_path) if backup_path else None,
                "total_rows": sum(counts.values()),
                "filtered_rows": counts["ERROR"] + counts["WARNING"],
                "status_counts": counts,
                "message": (f"Input opened with validation rows visible.{conversion_message} Backup: {backup_path}" if backup_path else f"Input opened.{conversion_message} Legacy XLS cannot persist Excel filters or cell highlighting."),
            }
        except Exception as exc:
            return {"success": False, "error": f"Could not create the filtered validation view: {exc}"}

    def select_file(self) -> Dict[str, Any]:
        """Open a native input-file dialog and return the selected path."""
        window = self._native_window()
        if window is None:
            return {"success": False, "error": "Native file dialog is unavailable."}
        try:
            import webview
            paths = window.create_file_dialog(
                webview.OPEN_DIALOG,
                allow_multiple=False,
                file_types=("Input files (*.csv;*.xlsx;*.xls;*.xlsm)",),
            )
            path = paths[0] if paths else ""
            return {"success": bool(path), "path": path}
        except Exception as exc:
            return {"success": False, "error": str(exc)}

    def select_directory(self) -> str:
        """Open a native destination-folder dialog."""
        Path(self.default_download_dir).mkdir(parents=True, exist_ok=True)
        window = self._native_window()
        if window is None:
            return self.get_default_download_directory()
        try:
            import webview
            paths = window.create_file_dialog(webview.FOLDER_DIALOG)
            if paths:
                root = self._persist_download_root(str(paths[0]))
                return self._new_run_directory(root)
            return self.get_default_download_directory()
        except Exception:
            return self.default_download_dir

    def _append_log(self, session_id: int, log_type: str, message: str):
        with self._runtime_lock:
            runtime = self._runtime.get(session_id)
            if not runtime:
                return
            runtime["latest_logs"].append({"type": log_type, "message": message})
            runtime["latest_logs"] = runtime["latest_logs"][-30:]

    async def _run_session_async(self, session_id: int, download_dir: str):
        cursor = self.db.conn.cursor()
        cursor.execute(
            "SELECT po_number, company_code FROM po_downloads WHERE session_id = ? AND status = 'PENDING' ORDER BY id ASC",
            (session_id,),
        )
        rows = cursor.fetchall()

        runtime = self._runtime.get(session_id, {})
        concurrency = max(1, int(runtime.get("concurrency", 11)))

        # Use the same authentication policy as the GUI and canonical CLI.
        # This fallback path never invents a second cache or browser flow.
        auth_result = await self.auth_service.ensure_session(interactive=False)
        cookies = dict(auth_result.cookies) or None
        if cookies and auth_result.state in {AuthState.VALID, AuthState.UNAVAILABLE}:
            self._cookies = dict(cookies)
        else:
            cookies = None
        if not cookies:
            raise RuntimeError(auth_result.message or "Coupa authentication is required before downloading.")

        crawler = CoupaCrawler(
            self.db,
            session_id,
            download_dir,
            concurrency=concurrency,
            cookies=cookies,
        )
        self._set_session_status(session_id, "RUNNING")
        self._append_log(
            session_id,
            "System",
            f"Starting session {session_id} with {len(rows)} POs (workers={concurrency})",
        )

        semaphore = asyncio.Semaphore(concurrency)
        stopped = False

        async def process_one(po_number: str, company_code: str):
            async with semaphore:
                _check_runtime_stop(session_id, self._runtime, self._runtime_lock)
                await _check_runtime_pause(session_id, self._runtime, self._runtime_lock)

                self._append_log(session_id, "Info", f"Processing PO {po_number} ({company_code})")
                result = await crawler.process_po(po_number, company_code)
                po_row = self.db.get_po(session_id, po_number)
                po_status = (po_row or {}).get("status", "ERROR")
                failed = po_status in {"ERROR", "SKIPPED_VERIFICATION_REQUIRED"}

                with self._runtime_lock:
                    current = self._runtime.get(session_id)
                    if current:
                        current["processed"] += 1
                        if failed:
                            current["errors"] += 1

                if failed:
                    self._append_log(
                        session_id,
                        "Error",
                        f"PO {po_number} failed: {(po_row or {}).get('error_message') or po_status}",
                    )
                else:
                    self._append_log(session_id, "Success", f"PO {po_number} completed")
                return result

        try:
            tasks = [process_one(row["po_number"], row["company_code"]) for row in rows]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            for result in results:
                if isinstance(result, SessionStoppedError):
                    stopped = True
                    continue
                if isinstance(result, Exception):
                    self._append_log(session_id, "Error", f"Task error: {result}")
                    with self._runtime_lock:
                        current = self._runtime.get(session_id)
                        if current:
                            current["errors"] += 1

            with self._runtime_lock:
                current = self._runtime.get(session_id)
                processed = int(current.get("processed", 0)) if current else 0
                errors = int(current.get("errors", 0)) if current else 0
                pending = max(len(rows) - processed, 0)
                if stopped or (current and current.get("stop_requested")):
                    final_status = "STOPPED"
                elif errors == 0 and pending == 0:
                    final_status = "SUCCESS"
                elif processed > errors:
                    final_status = "PARTIAL"
                else:
                    final_status = "FAILED"
                if current:
                    current["status"] = final_status

            self._set_session_status(session_id, final_status)
            message = f"Session {session_id} finished with status: {final_status}"
            self._append_log(session_id, "System", message)

        finally:
            await crawler.close()

    def _run_session_thread(self, session_id: int, download_dir: str):
        try:
            asyncio.run(self._run_session_async(session_id, download_dir))
        except Exception as e:
            with self._runtime_lock:
                runtime = self._runtime.get(session_id)
                if runtime:
                    runtime["status"] = "FAILED"
                    runtime["errors"] += 1
            self._set_session_status(session_id, "FAILED")
            self._append_log(session_id, "Error", f"Fatal session error: {e}")

    def inspect_input_file(self, filepath: str) -> Dict[str, Any]:
        """Return a cross-platform best-effort state for an editable input file."""
        path = Path(filepath)
        if not path.exists():
            return {"exists": False, "ready": False, "open_detected": False, "stable": False, "error": "File not found."}

        try:
            stat = path.stat()
            fingerprint = (int(stat.st_mtime_ns), int(stat.st_size))
            now = time.monotonic()
            previous = self._file_observations.get(str(path))
            stable = bool(previous and previous[0] == fingerprint and now - previous[1] >= 0.8)
            if not previous or previous[0] != fingerprint:
                self._file_observations[str(path)] = (fingerprint, now)

            lock_names = [
                f"~${path.name}",
                f".~lock.{path.name}#",
                f".{path.name}.lock",
            ]
            lock_files = [str(path.parent / name) for name in lock_names if (path.parent / name).exists()]
            readable = os.access(path, os.R_OK)
            open_detected = bool(lock_files)
            return {
                "exists": True,
                "readable": readable,
                "stable": stable,
                "open_detected": open_detected,
                "lock_files": lock_files,
                "mtime_ns": fingerprint[0],
                "size": fingerprint[1],
                "ready": readable and stable and not open_detected,
                "error": None,
            }
        except OSError as exc:
            return {"exists": True, "ready": False, "open_detected": True, "stable": False, "error": str(exc)}

    def start_download(self, session_id: int, download_dir: str, concurrency: int = 11) -> Dict[str, Any]:
        try:
            session_id = int(session_id)
            download_dir = (download_dir or self.default_download_dir).strip() or self.default_download_dir
            os.makedirs(download_dir, exist_ok=True)
            self.default_download_dir = download_dir

            cursor = self.db.conn.cursor()
            cursor.execute(
                "SELECT COUNT(*) as total FROM po_downloads WHERE session_id = ? AND status = 'PENDING'",
                (session_id,),
            )
            total = int(cursor.fetchone()["total"])

            if total == 0:
                return {"success": False, "error": "No pending POs found in this session."}

            with self._runtime_lock:
                existing = self._runtime.get(session_id)
                if existing and existing.get("status") in {"RUNNING", "PAUSED"}:
                    return {"success": False, "error": "Session already running."}

                self._runtime[session_id] = {
                    "status": "RUNNING",
                    "started_at": time.time(),
                    "processed": 0,
                    "total": total,
                    "errors": 0,
                    "paused": False,
                    "stop_requested": False,
                    "concurrency": concurrency,
                    "latest_logs": [],
                }

            self._set_session_status(session_id, "RUNNING")

            worker = threading.Thread(
                target=self._run_session_thread,
                args=(session_id, download_dir),
                daemon=True,
            )
            with self._runtime_lock:
                self._runtime[session_id]["thread"] = worker

            worker.start()
            return {"success": True, "session_id": session_id, "total": total}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def get_active_session_status(self, session_id: int) -> Dict[str, Any]:
        session_id = int(session_id)
        with self._runtime_lock:
            runtime = self._runtime.get(session_id)

        if runtime:
            elapsed = max(time.time() - runtime.get("started_at", time.time()), 0.1)
            processed = int(runtime.get("processed", 0))
            total = int(runtime.get("total", 0))
            errors = int(runtime.get("errors", 0))
            speed = (processed / elapsed) * 60.0 if processed > 0 else 0.0
            remaining = max(total - processed, 0)
            eta_seconds = int((remaining / speed) * 60) if speed > 0 else 0
            eta = time.strftime("%M:%S", time.gmtime(eta_seconds)) if speed > 0 else "--:--"

            with self._runtime_lock:
                latest_logs = list(runtime.get("latest_logs", []))
                runtime["latest_logs"] = []
                status = runtime.get("status", "PENDING")

            return {
                "status": status,
                "total": total,
                "processed": processed,
                "speed": speed,
                "eta": eta,
                "errors": errors,
                "latest_logs": latest_logs,
            }

        cursor = self.db.conn.cursor()
        cursor.execute("SELECT status FROM sessions WHERE id = ?", (session_id,))
        session_row = cursor.fetchone()
        session_status = session_row["status"] if session_row else "UNKNOWN"

        cursor.execute(
            """
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN status != 'PENDING' THEN 1 ELSE 0 END) as processed,
                SUM(CASE WHEN status = 'ERROR' THEN 1 ELSE 0 END) as errors
            FROM po_downloads
            WHERE session_id = ?
            """,
            (session_id,),
        )
        stats = cursor.fetchone()
        total = int(stats["total"] or 0)
        processed = int(stats["processed"] or 0)
        errors = int(stats["errors"] or 0)

        return {
            "status": session_status,
            "total": total,
            "processed": processed,
            "speed": 0.0,
            "eta": "--:--",
            "errors": errors,
            "latest_logs": [],
        }

    def pause_download(self, session_id: int) -> Dict[str, Any]:
        try:
            session_id = int(session_id)
            with self._runtime_lock:
                runtime = self._runtime.get(session_id)
                if not runtime:
                    return {"success": False, "error": "Session not active."}
                runtime["paused"] = True
                runtime["status"] = "PAUSED"
            self._set_session_status(session_id, "PAUSED")
            self._append_log(session_id, "System", "Session paused")
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def resume_download(self, session_id: int) -> Dict[str, Any]:
        try:
            session_id = int(session_id)
            with self._runtime_lock:
                runtime = self._runtime.get(session_id)
                if not runtime:
                    return {"success": False, "error": "Session not active."}
                runtime["paused"] = False
                runtime["status"] = "RUNNING"
            self._set_session_status(session_id, "RUNNING")
            self._append_log(session_id, "System", "Session resumed")
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def stop_download(self, session_id: int) -> Dict[str, Any]:
        try:
            session_id = int(session_id)
            with self._runtime_lock:
                runtime = self._runtime.get(session_id)
                if not runtime:
                    return {"success": False, "error": "Session not active."}
                runtime["stop_requested"] = True
            self._append_log(session_id, "System", "Stop requested by user")
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": str(e)}

    @staticmethod
    def _safe_user_path(path: str) -> str:
        """Redact the user's home directory before a report is shared."""
        value = str(path or "")
        home = str(Path.home())
        return value.replace(home, "~")

    @staticmethod
    def _package_version(distribution: str, import_name: str | None = None) -> str:
        try:
            from importlib.metadata import version
            return version(distribution)
        except Exception:
            try:
                import importlib
                module = importlib.import_module(import_name or distribution.replace("-", "_"))
                return str(getattr(module, "__version__", "bundled"))
            except Exception:
                return "not installed"

    @staticmethod
    def _application_version() -> str:
        candidates = []
        if getattr(sys, "_MEIPASS", None):
            candidates.append(Path(sys._MEIPASS) / ".version")
        candidates.extend([
            Path(__file__).resolve().parents[2] / ".version",
            Path.cwd() / ".version",
        ])
        for path in candidates:
            try:
                value = path.read_text(encoding="utf-8").strip()
                if value:
                    return value.lstrip("v")
            except OSError:
                continue
        return "unknown"

    @staticmethod
    def _find_edge() -> Optional[str]:
        candidates = [
            shutil.which("msedge"),
            shutil.which("microsoft-edge"),
            "/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge",
            r"C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedge.exe",
            r"C:\\Program Files\\Microsoft\\Edge\\Application\\msedge.exe",
        ]
        return next((candidate for candidate in candidates if candidate and Path(candidate).exists()), None)

    @staticmethod
    def _find_chrome() -> Optional[str]:
        candidates = [
            shutil.which("google-chrome"),
            shutil.which("chrome"),
            "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
            r"C:\\Users\\%USERNAME%\\AppData\\Local\\Google\\Chrome\\Application\\chrome.exe",
            r"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe",
            r"C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe",
        ]
        expanded = [os.path.expandvars(candidate) for candidate in candidates if candidate]
        return next((candidate for candidate in expanded if Path(candidate).exists()), None)

    @staticmethod
    def _driver_path(browser: str) -> Optional[str]:
        """Resolve a Selenium-managed driver without starting a browser."""
        try:
            from selenium.webdriver.common.selenium_manager import SeleniumManager

            result = SeleniumManager().binary_paths(["--browser", browser])
            path = result.get("driver_path") if isinstance(result, dict) else None
            return str(path) if path and Path(path).exists() else None
        except Exception:
            names = {
                "edge": ("msedgedriver", "msedgedriver.exe"),
                "chrome": ("chromedriver", "chromedriver.exe"),
            }
            return next((shutil.which(name) for name in names.get(browser, ())), None)

    @staticmethod
    def _edge_driver_path() -> Optional[str]:
        """Resolve the EdgeDriver Selenium will use without starting Edge."""
        return AppAPI._driver_path("edge")

    @staticmethod
    def _chrome_driver_path() -> Optional[str]:
        return AppAPI._driver_path("chrome")

    @staticmethod
    def _edge_version(edge_path: str | None) -> str:
        return AppAPI._browser_version(edge_path, "Microsoft Edge")

    @staticmethod
    def _chrome_version(chrome_path: str | None) -> str:
        return AppAPI._browser_version(chrome_path, "Google Chrome")

    @staticmethod
    def _browser_version(browser_path: str | None, label: str) -> str:
        if not browser_path:
            return f"{label} was not found"
        if sys.platform == "darwin":
            info_plist = Path(browser_path).parents[1] / "Info.plist"
            try:
                with info_plist.open("rb") as stream:
                    version = plistlib.load(stream).get("CFBundleShortVersionString")
                if version:
                    return str(version)
            except (OSError, ValueError, KeyError):
                pass
        return AppAPI._command_version([browser_path, "--version"])

    @staticmethod
    def _version_components(value: str) -> tuple[int, ...]:
        import re

        match = re.search(r"\b(\d+(?:\.\d+){1,3})\b", value or "")
        return tuple(int(part) for part in match.group(1).split(".")) if match else ()

    @staticmethod
    def _command_version(command: list[str]) -> str:
        try:
            result = subprocess.run(command, capture_output=True, text=True, timeout=4, check=False)
            return (result.stdout or result.stderr).strip().splitlines()[0][:160]
        except Exception as exc:
            return f"unavailable ({exc})"

    @staticmethod
    def _clipboard_copy(text: str) -> tuple[bool, str]:
        commands = []
        if sys.platform == "darwin":
            commands = [["pbcopy"]]
        elif sys.platform.startswith("win"):
            commands = [["clip"]]
        else:
            commands = [["xclip", "-selection", "clipboard"], ["xsel", "--clipboard", "--input"]]
        for command in commands:
            try:
                subprocess.run(command, input=text, text=True, check=True, timeout=5)
                return True, "Clipboard updated"
            except (FileNotFoundError, subprocess.SubprocessError):
                continue
        return False, "Native clipboard utility unavailable"

    def run_diagnostics(self, input_path: str = "") -> Dict[str, Any]:
        """Run safe local diagnostics without reading credentials or cookies."""
        checks: list[Dict[str, str]] = []

        def check(name: str, status: str, detail: str) -> None:
            checks.append({"name": name, "status": status, "detail": detail})

        edge_path = self._find_edge()
        check("Operating system", "PASS", f"{platform.system()} {platform.release()} ({platform.machine()})")
        check("Python runtime", "PASS", platform.python_version())
        check("Application version", "PASS", self._application_version())
        edge_version = self._edge_version(edge_path)
        edge_status = "PASS" if edge_path and not edge_version.startswith("unavailable") else "WARN"
        check("Microsoft Edge", edge_status, edge_version)
        chrome_path = self._find_chrome()
        chrome_version = self._chrome_version(chrome_path)
        chrome_status = "PASS" if chrome_path and not chrome_version.startswith("unavailable") else "WARN"
        check("Google Chrome", chrome_status, chrome_version)

        # ---- Supported authentication browsers ----
        try:
            browser_options = self.auth_service.browser_options(self._read_settings().get("auth_browser"))
            available = browser_options.get("available", [])
            if available:
                names = ", ".join(str(item.get("name", "browser")) for item in available)
                selected_id = str(browser_options.get("selected") or "unknown")
                selected_name = next(
                    (str(item.get("name", selected_id)) for item in available if item.get("id") == selected_id),
                    selected_id,
                )
                source = str(browser_options.get("selection_source") or "settings")
                check("Coupa sign-in browsers", "PASS", f"{names}; selected={selected_name} ({source})")
            else:
                check("Coupa sign-in browsers", "WARN", "Microsoft Edge or Google Chrome was not found")
        except Exception as exc:
            check("Coupa sign-in browsers", "WARN", f"Could not detect supported browsers: {exc}")

        # ---- App-owned sign-in profiles ----
        try:
            browser_options = self.auth_service.browser_options(self._read_settings().get("auth_browser"))
            profiles = browser_options.get("profiles", {})
            for kind in BrowserKind:
                info = profiles.get(kind.value) if isinstance(profiles, dict) else None
                label = f"{kind.value.title()} profile"
                if isinstance(info, dict) and (info.get("exists") or info.get("registered")):
                    check(label, "PASS", self._safe_user_path(str(info.get("path", ""))))
                else:
                    check(label, "PASS", f"App-owned profile will be created if {kind.value.title()} is selected")

            selected = browser_options.get("selected")
            selected_kind = BrowserKind(str(selected)) if selected else None
            selected_info = profiles.get(selected_kind.value) if selected_kind and isinstance(profiles, dict) else None
            if isinstance(selected_info, dict) and selected_info.get("exists"):
                check("Coupa sign-in profile", "PASS", self._safe_user_path(str(selected_info.get("path", ""))))
            else:
                check("Coupa sign-in profile", "PASS", "Created on first Coupa sign-in")
        except Exception as exc:
            check("Edge profile", "WARN", f"Could not inspect the app-owned profile: {exc}")
            check("Chrome profile", "WARN", f"Could not inspect the app-owned profile: {exc}")
            check("Coupa sign-in profile", "WARN", f"Could not inspect the app-owned profile: {exc}")

        # ---- Coupa session ----
        try:
            result = asyncio.run(self.auth_service.check())
            if result.state is AuthState.VALID:
                check("Coupa session", "PASS", "Cached session is valid")
            elif result.state is AuthState.UNAVAILABLE and result.has_cached_session:
                check("Coupa session", "WARN", "Session validation unavailable — Coupa or network may be unreachable")
            elif result.state is AuthState.MISSING:
                check("Coupa session", "WARN", "No cached Coupa session; sign-in is required")
            else:
                check("Coupa session", "WARN", "Cached session expired; re-authentication is required")
        except Exception as exc:
            check("Coupa session", "WARN", f"Could not validate session: {exc}")

        # ---- Python portable edition ----
        if self._is_python_portable():
            check("Distribution", "PASS", "Python portable edition (no installation required)")

        for browser_label, browser_path, browser_version, driver_label, driver_path in (
            ("Microsoft Edge", edge_path, edge_version, "Microsoft EdgeDriver", self._edge_driver_path()),
            ("Google Chrome", chrome_path, chrome_version, "ChromeDriver", self._chrome_driver_path()),
        ):
            driver_version = self._command_version([driver_path, "--version"]) if driver_path else f"{driver_label} was not resolved by Selenium Manager"
            browser_numbers = self._version_components(browser_version)
            driver_numbers = self._version_components(driver_version)
            if not driver_path:
                driver_status = "WARN"
                driver_detail = driver_version
            elif browser_numbers and driver_numbers and browser_numbers[:3] != driver_numbers[:3]:
                driver_status = "WARN"
                driver_detail = f"{driver_version} — browser/driver versions differ"
            else:
                driver_status = "PASS"
                driver_detail = f"{driver_version} ({Path(driver_path).name})"
            check(driver_label, driver_status, driver_detail)

        webview_version = self._package_version("pywebview", "webview")
        check("WebView runtime", "PASS" if webview_version != "not installed" else "FAIL", f"pywebview {webview_version}")

        runtime_packages = {
            "pandas": ("pandas", "pandas"),
            "openpyxl": ("openpyxl", "openpyxl"),
            "httpx": ("httpx", "httpx"),
            "selenium": ("selenium", "selenium"),
            "beautifulsoup4": ("beautifulsoup4", "bs4"),
            "extract-msg": ("extract-msg", "extract_msg"),
            "fpdf2": ("fpdf2", "fpdf"),
        }
        for label, (distribution, import_name) in runtime_packages.items():
            version = self._package_version(distribution, import_name)
            check(f"Dependency: {label}", "PASS" if version != "not installed" else "FAIL", version)

        try:
            usage = shutil.disk_usage(Path.home())
            free_gb = usage.free / (1024 ** 3)
            check("Disk space", "PASS" if free_gb >= 1 else "WARN", f"{free_gb:.1f} GB free")
        except OSError as exc:
            check("Disk space", "WARN", str(exc))

        for label, directory in {
            "Download directory": self.default_download_dir,
            "Application data": str(Path.home() / ".contract_downloader"),
        }.items():
            try:
                Path(directory).mkdir(parents=True, exist_ok=True)
                probe = Path(directory) / ".diagnostic-write-test"
                probe.write_text("ok", encoding="utf-8")
                probe.unlink(missing_ok=True)
                check(label, "PASS", self._safe_user_path(directory))
            except OSError as exc:
                check(label, "FAIL", f"{self._safe_user_path(directory)} — {exc}")

        try:
            db_check = self.db.conn.execute("PRAGMA integrity_check").fetchone()[0]
            check("Session database", "PASS" if db_check == "ok" else "FAIL", str(db_check))
        except Exception as exc:
            check("Session database", "FAIL", str(exc))

        async def network_checks() -> None:
            import httpx
            for label, url in {
                "Coupa connectivity": "https://unilever.coupahost.com",
                "GitHub connectivity": "https://api.github.com",
            }.items():
                try:
                    async with httpx.AsyncClient(
                        timeout=5,
                        follow_redirects=True,
                        verify=system_ssl_context(),
                    ) as client:
                        response = await client.get(url)
                    check(label, "PASS" if response.status_code < 500 else "WARN", f"HTTP {response.status_code}")
                except Exception as exc:
                    check(label, "WARN", str(exc))

        try:
            asyncio.run(network_checks())
        except Exception as exc:
            check("Network diagnostics", "WARN", str(exc))

        if input_path:
            state = self.inspect_input_file(input_path)
            input_name = Path(input_path).name
            if not state.get("exists"):
                check("Selected input", "FAIL", f"{input_name}: file not found")
            elif state.get("open_detected"):
                check("Selected input", "WARN", f"{input_name}: Excel appears to be open")
            elif state.get("ready"):
                check("Selected input", "PASS", f"{input_name}: saved and readable")
            else:
                check("Selected input", "WARN", f"{input_name}: waiting for a stable save")

        passed = sum(item["status"] == "PASS" for item in checks)
        failed = sum(item["status"] == "FAIL" for item in checks)
        warnings = sum(item["status"] == "WARN" for item in checks)
        lines = [
            "CONTRACT DOWNLOADER - HOST DIAGNOSTIC REPORT",
            f"Generated: {datetime.now(timezone.utc).isoformat()}",
            "",
            f"Summary: {passed} pass | {warnings} warning(s) | {failed} failure(s)",
            "",
        ]
        for item in checks:
            lines.append(f"[{item['status']}] {item['name']}: {item['detail']}")
        lines.extend([
            "",
            "Privacy: credentials, cookies, PO contents, and full user paths are not included.",
            "Please send this report together with a screenshot of the error and the approximate run time.",
        ])
        report = "\n".join(lines)
        return {"success": True, "report": report, "checks": checks, "summary": {"passed": passed, "warnings": warnings, "failed": failed}}

    def copy_diagnostics_report(self, report: str) -> Dict[str, Any]:
        success, message = self._clipboard_copy(report or "")
        return {"success": success, "message": message}

    def save_diagnostics_report(self, report: str) -> Dict[str, Any]:
        try:
            destination = Path.home() / ".contract_downloader" / f"diagnostics_{time.strftime('%Y%m%d-%H%M%S')}.txt"
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_text(report or "", encoding="utf-8")
            return {"success": True, "path": str(destination)}
        except Exception as exc:
            return {"success": False, "error": str(exc)}

    def set_auth_cookies(self, cookies: Dict[str, str]) -> None:
        """Store authenticated cookies in the shared authentication service."""
        self.auth_service.set_cookies(cookies)
        self._cookies = dict(cookies)

    def reset_authentication(self) -> Dict[str, Any]:
        """Clear cached Coupa state and only app-owned browser profiles."""
        result = self.auth_service.reset()
        if result.get("success") or "cookies" in result.get("removed", []):
            self._cookies = None
        return result

    def check_auth(self) -> Dict[str, Any]:
        """Check the shared cache and preserve temporary-unavailable state."""
        try:
            result = asyncio.run(self.auth_service.check())
            if result.state is AuthState.VALID:
                self._cookies = dict(result.cookies)
                return {"authenticated": True, "state": "cached", "message": result.message}
            if result.state is AuthState.UNAVAILABLE and result.has_cached_session:
                self._cookies = dict(result.cookies)
                return {
                    "authenticated": False,
                    "state": "unavailable",
                    "has_cached_session": True,
                    "message": "Cached Coupa session found; live validation is temporarily unavailable. Keeping the cached session without requesting a new sign-in.",
                }
            self._cookies = None
            state = "login_required" if result.state is AuthState.MISSING else "expired"
            return {"authenticated": False, "state": state, "message": result.message}
        except Exception as exc:
            self._cookies = None
            return {"authenticated": False, "state": "unavailable", "message": str(exc)}

    def reset_new_run(self, filepath: str = "") -> Dict[str, Any]:
        """Reset the wizard and remove only a template created by the app."""
        import re

        value = str(filepath or "").strip()
        if not value:
            return {"success": True, "deleted": False, "preserved": False}
        path = Path(value).expanduser().resolve()
        template_dir = (Path.home() / "Documents" / "Contract Downloader" / "templates").resolve()
        is_generated_template = (
            path.parent == template_dir
            and re.fullmatch(r"input_template_\d{8}-\d{6}\.xlsx", path.name) is not None
        )
        if not is_generated_template:
            return {"success": True, "deleted": False, "preserved": True}
        try:
            path.unlink(missing_ok=True)
            return {"success": True, "deleted": True, "path": str(path)}
        except OSError as exc:
            return {"success": False, "error": f"Could not delete the generated template: {exc}"}

    def generate_input_template(self) -> Dict[str, Any]:
        """Generate a template CSV file ready for the user to populate with PO data.

        Creates a file next to the default input CSV location with a timestamp
        so the user can fill it out without overwriting anything.
        """
        template_dir = Path.home() / "Documents" / "Contract Downloader" / "templates"
        template_dir.mkdir(parents=True, exist_ok=True)

        timestamp = time.strftime("%Y%m%d-%H%M%S")
        template_path = template_dir / f"input_template_{timestamp}.xlsx"

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        input_headers = [
            "PO_NUMBER", "SUPPLIER", "<|>",
            "Company", "Year", "Quarter", "Business Unit",
        ]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Input"
        sheet.append(input_headers)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(input_headers))}1"

        header_fill = PatternFill("solid", fgColor="173B56")
        separator_fill = PatternFill("solid", fgColor="DCECF5")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        sheet[1][2].fill = separator_fill
        sheet[1][2].font = Font(bold=True, color="173B56")
        widths = [20, 28, 8, 24, 12, 14, 24]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

        instructions = workbook.create_sheet("Instructions")
        instructions.column_dimensions["A"].width = 24
        instructions.column_dimensions["B"].width = 100
        instructions.append(["Contract Downloader — input guide", ""])
        instructions.append(["Required fields", "PO_NUMBER and SUPPLIER must be filled for every row."])
        instructions.append(["The <|> separator", "Columns before <|> identify the PO. Columns after <|> define the destination folder hierarchy."])
        instructions.append(["Folder levels", "The first column after <|> is level 1, the next is level 2, and so on. Blank values become Unknown."])
        instructions.append(["Example", "Company=ACME, Year=2026, Quarter=Q3 → ACME/2026/Q3/PO_NUMBER"])
        instructions.append(["Workflow", "Fill the Input sheet, save and close Excel, choose the file in the app, validate, correct any errors, then start."])
        for cell in instructions[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for row in instructions.iter_rows(min_row=2):
            row[0].font = Font(bold=True, color="173B56")
            row[1].alignment = Alignment(wrap_text=True, vertical="top")
            row[0].alignment = Alignment(vertical="top")
        instructions.freeze_panes = "A2"
        workbook.save(template_path)

        template_path_str = str(template_path)
        try:
            self._open_path(template_path)
        except Exception:
            pass

        return {
            "success": True,
            "path": template_path_str,
            "format": "xlsx",
            "message": f"Excel template created at {template_path_str}",
        }

    def validate_input_file(self, filepath: str) -> Dict[str, Any]:
        """Validate a populated CSV/XLSX file before importing.

        Errors are grouped by cause so the UI can offer a targeted fix per
        group (remove blank rows, remove duplicate POs, clean invalid
        characters) or open the file for assisted editing when no safe
        automatic repair exists.
        """
        import re
        from pathlib import Path
        from src.engine.input_schema import (
            canonicalize_po_value,
            clean_scalar,
            columns_of_dataframe,
            detect_po_parts,
            header_comparison_key,
            is_excel_numeric_coercion,
            is_placeholder_po,
            is_placeholder_supplier,
            is_valid_canonical_po,
            normalize_po_value,
            normalize_supplier_value,
            parse_mapping_env,
            resolve_data_mapping,
            resolve_mapping,
            SUPPORTED_INPUT_SUFFIXES,
        )

        path = Path(filepath)
        if not path.exists():
            return {"valid": False, "errors": [f"File not found: {filepath}"], "warnings": [], "file_state": {"ready": False}}
        if path.suffix.lower() not in SUPPORTED_INPUT_SUFFIXES:
            return {"valid": False, "errors": ["Unsupported input format. Use CSV, XLS, XLSX or XLSM."], "warnings": []}

        file_state = self.inspect_input_file(filepath)
        if file_state.get("open_detected"):
            return {
                "valid": False,
                "errors": ["The input file appears to be open in Excel. Save and close it before validating."],
                "warnings": [],
                "file_state": file_state,
            }
        if not file_state.get("readable", True):
            return {"valid": False, "errors": ["The input file is not readable."], "warnings": [], "file_state": file_state}

        try:
            df = self._read_input_dataframe(path)
        except Exception as e:
            return {"valid": False, "errors": [f"Failed to parse input file: {e}"], "warnings": []}

        headers = columns_of_dataframe(df)
        header_keys: Dict[str, list[str]] = {}
        for header in headers:
            key = header_comparison_key(header)
            if key:
                header_keys.setdefault(key, []).append(str(header))
        empty_headers = [
            str(header) for header in headers
            if not str(header).strip() or str(header).strip().casefold().startswith("unnamed:")
        ]
        duplicate_headers = [values for values in header_keys.values() if len(values) > 1]
        mapping_values = self._mapping_for(str(path))
        mapping_values.update(parse_mapping_env() or {})
        mapping, mapping_suggestions = resolve_data_mapping(df, headers, mapping_values)

        hierarchy_columns = []
        empty_hierarchy_columns = []
        sep_column = next((column for column in df.columns if str(column).strip() == "<|>"), None)
        if sep_column is not None:
            hierarchy_columns = [str(column) for column in list(df.columns)[list(df.columns).index(sep_column) + 1:]]
        else:
            # Non-standard inputs without the <|> separator: every column
            # except the mapped PO/Supplier is a hierarchy candidate.
            excluded = {mapping.get("po"), mapping.get("supplier")}
            hierarchy_columns = [str(column) for column in df.columns if str(column) not in excluded]
        for column in hierarchy_columns:
            series = df[column].map(clean_scalar)
            series = series[series != ""]
            if series.empty:
                empty_hierarchy_columns.append(str(column))

        def folder_part_preview(value: Any) -> str:
            text = clean_scalar(value)
            if not text:
                return "Unknown"
            text = text.replace("/", "_").replace("\\", "_")
            text = "_".join(text.split()).strip("._")
            return text or "Unknown"

        reserved_names = {"con", "prn", "aux", "nul"} | {f"com{index}" for index in range(1, 10)} | {f"lpt{index}" for index in range(1, 10)}
        folder_safety_details: list[Dict[str, Any]] = []
        folder_columns = [column for column in [mapping.get("supplier"), *hierarchy_columns] if column]

        def add_folder_safety_detail(row: int, column: str, value: str, cleaned: str, reason: str) -> None:
            folder_safety_details.append({
                "row": row,
                "column": str(column),
                "value": value,
                "cleaned": cleaned,
                "reason": reason,
                "parts": [f"{column}: {value} → {cleaned} ({reason})"],
            })

        for column in dict.fromkeys(folder_columns):
            if column not in df.columns:
                continue
            sanitized: Dict[str, set[str]] = {}
            value_rows: Dict[str, list[int]] = {}
            for index, raw in df[column].items():
                raw_value = clean_scalar(raw)
                if not raw_value:
                    continue
                row_number = int(index) + 2
                cleaned_value = folder_part_preview(raw_value)
                key = cleaned_value.casefold()
                sanitized.setdefault(key, set()).add(raw_value)
                value_rows.setdefault(key, []).append(row_number)
                stem = cleaned_value.split(".", 1)[0].casefold()
                reasons = []
                if "/" in raw_value or "\\" in raw_value:
                    reasons.append("contains a path separator ('/' or backslash); it will be replaced with '_' in the folder name")
                if ".." in raw_value:
                    reasons.append("contains a traversal sequence")
                if stem in reserved_names:
                    reasons.append("uses a reserved folder name")
                if len(cleaned_value) > 100:
                    reasons.append("exceeds the 100-character folder limit")
                if reasons:
                    add_folder_safety_detail(row_number, str(column), raw_value, cleaned_value, "; ".join(reasons))
            for cleaned_value, raw_group in sanitized.items():
                if len(raw_group) > 1:
                    reason = f"different values collapse to '{cleaned_value}'"
                    for raw_value in raw_group:
                        for row_number in value_rows.get(cleaned_value, []):
                            row_raw = clean_scalar(df.at[row_number - 2, column])
                            if row_raw == raw_value:
                                add_folder_safety_detail(row_number, str(column), raw_value, folder_part_preview(raw_value), reason)

        folder_safety_issues = [
            f"row {item['row']} · {item['column']}: {item['value']} ({item['reason']})"
            for item in folder_safety_details
        ]

        errors: list[str] = []
        warnings: list[str] = []
        fixes: list[Dict[str, Any]] = []
        groups: list[Dict[str, Any]] = []
        valid_count = 0

        if empty_headers:
            errors.append(f"Empty or unnamed header(s): {empty_headers[:10]}")
            groups.append({
                "id": "empty_headers",
                "severity": "error",
                "title": "Empty or unnamed headers",
                "count": len(empty_headers),
                "fixable": False,
                "message": "Rename empty or unnamed columns before continuing.",
            })
        required_keys = {header_comparison_key(mapping.get("po")), header_comparison_key(mapping.get("supplier"))}
        hierarchy_header_set = set(hierarchy_columns)
        required_duplicates = []
        hierarchy_duplicates = []
        other_duplicates = []
        for values in duplicate_headers:
            if header_comparison_key(values[0]) in required_keys:
                non_hierarchy = [value for value in values if value not in hierarchy_header_set]
                if len(non_hierarchy) > 1:
                    required_duplicates.append(values)
                else:
                    hierarchy_duplicates.append(values)
            else:
                other_duplicates.append(values)
        similar_duplicates = hierarchy_duplicates + other_duplicates
        if required_duplicates:
            errors.append(f"Required column headers are ambiguous: {required_duplicates[:5]}")
            groups.append({
                "id": "duplicate_required_headers",
                "severity": "error",
                "title": "Duplicate required headers",
                "count": len(required_duplicates),
                "fixable": False,
                "message": "The PO or Supplier column appears more than once before the hierarchy separator.",
            })
        if similar_duplicates:
            warnings.append(f"Similar headers found: {similar_duplicates[:5]}")
            groups.append({
                "id": "duplicate_headers",
                "severity": "warning",
                "title": "Similar column headers",
                "count": len(similar_duplicates),
                "fixable": False,
                "message": "Review columns with the same base name before choosing folder levels.",
            })
        if empty_hierarchy_columns:
            warnings.append(f"Empty hierarchy column(s): {empty_hierarchy_columns[:10]}")
            groups.append({
                "id": "empty_hierarchy_columns",
                "severity": "warning",
                "title": "Empty hierarchy columns",
                "count": len(empty_hierarchy_columns),
                "fixable": False,
                "message": f"These columns contain no values and will not create folder levels: {empty_hierarchy_columns[:10]}.",
            })
        if folder_safety_details:
            warnings.append(f"Folder value(s) will be sanitized or collide after cleanup (showing first 10): {folder_safety_issues[:10]}")
            groups.append({
                "id": "folder_value_safety",
                "severity": "warning",
                "title": "Folder names will be sanitized",
                "count": len(folder_safety_details),
                "rows": sorted({int(item["row"]) for item in folder_safety_details}),
                "row_details": folder_safety_details,
                "rows_are_excel_rows": True,
                "fixable": False,
                "message": "Warning only: the download can continue. Unsafe characters such as '/' or backslash become '_' in folder names; reserved names, long values and collisions are also normalized.",
            })

        # Required: PO column
        po_col = mapping.get("po")
        if not po_col:
            errors.append(
                f"Missing PO Number column. Expected one of: PO_NUMBER, PO, Pedido. "
                f"Found: {list(df.columns)}"
            )
            groups.append({
                "id": "missing_po_column",
                "severity": "error",
                "title": "Missing PO Number column",
                "count": 1,
                "fixable": False,
                "mapping": True,
            })

        # Required: company/supplier column
        company_col = mapping.get("supplier")
        if not company_col:
            errors.append(
                f"Missing Supplier/Company column. Expected one of: SUPPLIER, LegalEntity, CompanyCode, Empresa. "
                f"Found: {list(df.columns)}"
            )
            groups.append({
                "id": "missing_supplier_column",
                "severity": "error",
                "title": "Missing Supplier/Company column",
                "count": 1,
                "fixable": False,
                "mapping": True,
            })

        if po_col and company_col and header_comparison_key(po_col) == header_comparison_key(company_col):
            errors.append("PO Number and Supplier must be mapped to different columns.")
            groups.append({
                "id": "mapping_same_column",
                "severity": "error",
                "title": "PO and Supplier use the same column",
                "count": 1,
                "fixable": False,
                "mapping": True,
                "message": "Choose two different columns for PO Number and Supplier.",
            })

        if not errors:
            # Validate rows. Empty required values are errors because the user
            # should be able to correct the same file before starting a run.
            values = df[[po_col, company_col]].apply(lambda col: col.map(clean_scalar))
            empty_mask = values[po_col].eq("") | values[company_col].eq("")
            normalized_frame = df.apply(lambda col: col.map(clean_scalar))
            fully_blank_mask = normalized_frame.eq("").all(axis=1)
            # Pandas no longer accepts a Series as a multidimensional indexer
            # for Index.__getitem__. Convert masks to plain NumPy booleans.
            blank_mask = fully_blank_mask.to_numpy(dtype=bool)
            partial_mask = (empty_mask & ~fully_blank_mask).to_numpy(dtype=bool)
            blank_rows = [int(index) + 2 for index in df.index[blank_mask]]
            partial_rows = [int(index) + 2 for index in df.index[partial_mask]]
            if blank_rows:
                errors.append(f"Blank row(s) found: {blank_rows[:10]}")
                fixes.append({
                    "action": "remove_blank_rows",
                    "count": len(blank_rows),
                    "description": f"Remove {len(blank_rows)} completely blank row(s)",
                })
                groups.append({
                    "id": "blank_rows",
                    "severity": "error",
                    "title": "Blank rows",
                    "count": len(blank_rows),
                    "rows": blank_rows,
                    "fix_action": "remove_blank_rows",
                    "fixable": True,
                    "message": f"Row(s) completely empty (showing first 10): {blank_rows[:10]}",
                })
            if partial_rows:
                errors.append(f"Missing PO Number or Supplier on row(s): {partial_rows[:10]}")
                groups.append({
                    "id": "partial_rows",
                    "severity": "error",
                    "title": "Rows with missing PO or Supplier",
                    "count": len(partial_rows),
                    "rows": partial_rows,
                    "fixable": False,
                    "message": f"Row(s) without PO Number or Supplier (showing first 10): {partial_rows[:10]}",
                })

            excel_error_tokens = {"#ref!", "#div/0!", "#value!", "#n/a", "#name?", "#null!", "#num!"}
            excel_error_rows = [
                int(index) + 2
                for index, row in normalized_frame.iterrows()
                if any(str(value).casefold() in excel_error_tokens for value in row.tolist())
            ]
            if excel_error_rows:
                errors.append(f"Excel error value(s) found on row(s): {excel_error_rows[:10]}")
                groups.append({
                    "id": "excel_cell_errors",
                    "severity": "error",
                    "title": "Excel error values",
                    "count": len(excel_error_rows),
                    "rows": excel_error_rows,
                    "fixable": False,
                    "message": "Replace formula errors such as #REF!, #VALUE! or #N/A before continuing.",
                })

            required_whitespace_rows = []
            for index, row in df[[po_col, company_col]].iterrows():
                for column in (po_col, company_col):
                    raw_value = row[column]
                    cleaned_value = clean_scalar(raw_value)
                    if cleaned_value and str(raw_value) != cleaned_value:
                        required_whitespace_rows.append(int(index) + 2)
                        break
            if required_whitespace_rows:
                required_whitespace_rows = list(dict.fromkeys(required_whitespace_rows))
                warnings.append(f"Leading or trailing whitespace found on row(s): {required_whitespace_rows[:10]}")
                fixes.append({
                    "action": "normalize_required_values",
                    "count": len(required_whitespace_rows),
                    "description": f"Trim whitespace from required values on {len(required_whitespace_rows)} row(s)",
                })
                groups.append({
                    "id": "required_value_whitespace",
                    "severity": "warning",
                    "title": "Whitespace around PO or Supplier values",
                    "count": len(required_whitespace_rows),
                    "rows": required_whitespace_rows,
                    "fix_action": "normalize_required_values",
                    "fixable": True,
                    "message": "Leading and trailing whitespace is ignored by the pipeline, but can be removed from the source file.",
                })

            supplier_placeholder_mask = values[company_col].map(is_placeholder_supplier) & values[po_col].ne("")
            supplier_placeholder_rows = [int(index) + 2 for index in df.index[supplier_placeholder_mask.to_numpy(dtype=bool)]]
            if supplier_placeholder_rows:
                errors.append(f"Placeholder Supplier value(s) found on row(s): {supplier_placeholder_rows[:10]}")
                groups.append({
                    "id": "placeholder_supplier",
                    "severity": "error",
                    "title": "Placeholder Supplier values",
                    "count": len(supplier_placeholder_rows),
                    "rows": supplier_placeholder_rows,
                    "fixable": False,
                    "message": "Replace placeholder Supplier values such as Unknown, N/A or TBD.",
                })

            clean = df.loc[~empty_mask].copy()
            po_analysis: Dict[int, Dict[str, Any]] = {}
            multiple_po_rows: list[int] = []
            multiple_po_details: list[Dict[str, Any]] = []
            ambiguous_po_rows: list[int] = []
            for index, raw_value in clean[po_col].items():
                analysis = detect_po_parts(raw_value)
                po_analysis[index] = analysis
                if analysis["multiple"]:
                    row_number = int(index) + 2
                    multiple_po_rows.append(row_number)
                    multiple_po_details.append({"row": row_number, "parts": [part["canonical"] for part in analysis["parts"]], "raw": clean_scalar(raw_value)})
                elif analysis["ambiguous"]:
                    ambiguous_po_rows.append(int(index) + 2)
            if multiple_po_rows:
                errors.append(f"Multiple PO values found in cell(s) on row(s): {multiple_po_rows[:10]}")
                groups.append({
                    "id": "multiple_pos_in_cell",
                    "severity": "error",
                    "title": "Multiple POs in one cell",
                    "count": len(multiple_po_rows),
                    "rows": multiple_po_rows,
                    "row_details": multiple_po_details,
                    "fix_action": "split_multiple_pos",
                    "fixable": True,
                    "message": "Each row must contain one PO. Review the proposed split before applying it.",
                })
            if ambiguous_po_rows:
                errors.append(f"Ambiguous PO separator found on row(s): {ambiguous_po_rows[:10]}")
                groups.append({
                    "id": "ambiguous_po_value",
                    "severity": "error",
                    "title": "Ambiguous PO value",
                    "count": len(ambiguous_po_rows),
                    "rows": ambiguous_po_rows,
                    "fixable": False,
                    "message": "The value contains a separator but not enough complete PO prefixes to split it safely.",
                })
            clean["__po_key"] = clean[po_col].map(normalize_po_value)
            clean["__supplier_key"] = clean[company_col].map(normalize_supplier_value)
            grouped_pos = clean.groupby("__po_key", sort=False, dropna=False)
            duplicate_keys = [key for key, frame in grouped_pos if len(frame) > 1]
            conflict_keys = [
                key for key in duplicate_keys
                if grouped_pos.get_group(key)["__supplier_key"].nunique(dropna=False) > 1
            ]
            safe_duplicate_keys = [key for key in duplicate_keys if key not in conflict_keys]

            if conflict_keys:
                conflict_pos = [clean_scalar(grouped_pos.get_group(key).iloc[0][po_col]) for key in conflict_keys]
                errors.append(f"PO Number(s) linked to multiple Suppliers (showing first 10): {conflict_pos[:10]}")
                groups.append({
                    "id": "po_supplier_conflict",
                    "severity": "error",
                    "title": "POs linked to multiple Suppliers",
                    "count": len(conflict_keys),
                    "rows": conflict_pos,
                    "fixable": False,
                    "message": "Review these POs manually; automatic deduplication could discard a supplier relationship.",
                })

            if safe_duplicate_keys:
                duplicate_pos = [clean_scalar(grouped_pos.get_group(key).iloc[0][po_col]) for key in safe_duplicate_keys]
                errors.append(f"Duplicate PO Number(s) (showing first 10): {duplicate_pos[:10]}")
                fixes.append({
                    "action": "remove_duplicate_pos",
                    "count": len(safe_duplicate_keys),
                    "description": "Keep the first row for each duplicated PO Number with the same Supplier",
                })
                groups.append({
                    "id": "duplicate_pos",
                    "severity": "error",
                    "title": "Duplicate PO numbers",
                    "count": len(safe_duplicate_keys),
                    "rows": duplicate_pos,
                    "fix_action": "remove_duplicate_pos",
                    "fixable": True,
                    "message": f"PO(s) repeated with the same Supplier (showing first 10): {duplicate_pos[:10]}",
                })

            # Validate identifier semantics before offering any automatic cleanup.
            char_pattern = re.compile(r"[^A-Za-z0-9]")
            placeholder_pos: list[str] = []
            numeric_pos: list[str] = []
            dirty_pos: list[str] = []
            normalization_changes: list[Dict[str, Any]] = []
            invalid_pos: list[str] = []
            unusual_length_pos: list[str] = []
            for index, value in clean[po_col].map(clean_scalar).items():
                analysis = po_analysis.get(index, {})
                if analysis.get("multiple") or analysis.get("ambiguous"):
                    continue
                if is_placeholder_po(value):
                    placeholder_pos.append(value)
                    continue
                if is_excel_numeric_coercion(value):
                    numeric_pos.append(value)
                    continue
                canonical = canonicalize_po_value(value)
                if char_pattern.search(value) or value != canonical:
                    dirty_pos.append(value)
                    if value != canonical:
                        normalization_changes.append({
                            "row": int(index) + 2,
                            "column": str(po_col),
                            "old": value,
                            "new": canonical,
                        })
                if not is_valid_canonical_po(canonical):
                    invalid_pos.append(value)
                elif len(canonical[2:]) != 8:
                    unusual_length_pos.append(value)

            if placeholder_pos:
                errors.append(f"Placeholder PO value(s) found (showing first 10): {placeholder_pos[:10]}")
                groups.append({
                    "id": "placeholder_pos",
                    "severity": "error",
                    "title": "Placeholder PO values",
                    "count": len(set(normalize_po_value(value) for value in placeholder_pos)),
                    "rows": list(dict.fromkeys(placeholder_pos)),
                    "fixable": False,
                    "message": "Replace placeholders such as UNK, N/A or TBD with real PO numbers.",
                })
            if numeric_pos:
                errors.append(f"PO value(s) look Excel-coerced (showing first 10): {numeric_pos[:10]}")
                groups.append({
                    "id": "excel_numeric_coercion",
                    "severity": "error",
                    "title": "PO values may have been converted by Excel",
                    "count": len(set(normalize_po_value(value) for value in numeric_pos)),
                    "rows": list(dict.fromkeys(numeric_pos)),
                    "fixable": False,
                    "message": "Format the PO column as text and restore any lost leading zeros before continuing.",
                })
            if dirty_pos:
                errors.append(f"PO Number(s) with unusual characters (showing first 5): {dirty_pos[:5]}")
                fixes.append({
                    "action": "clean_invalid_chars",
                    "count": len(dirty_pos),
                    "description": f"Remove unusual characters from {len(dirty_pos)} PO Number(s)",
                })
                groups.append({
                    "id": "invalid_chars",
                    "severity": "error",
                    "title": "PO numbers with unusual characters",
                    "count": len(set(normalize_po_value(value) for value in dirty_pos)),
                    "rows": list(dict.fromkeys(dirty_pos)),
                    "fix_action": "clean_invalid_chars",
                    "fixable": True,
                    "changes": normalization_changes,
                    "message": "PO values must be uppercase and contain only letters and numbers; the proposed cleanup is shown before saving.",
                })

            format_pos = [
                value for value in invalid_pos
                if not is_placeholder_po(value) and not is_excel_numeric_coercion(value)
            ]
            if format_pos:
                errors.append(
                    f"{len(format_pos)} PO(s) have an invalid prefix or format (showing first 5): "
                    f"{format_pos[:5]}"
                )
                groups.append({
                    "id": "unusual_format",
                    "severity": "error",
                    "title": "PO numbers with an invalid format",
                    "count": len(set(normalize_po_value(value) for value in format_pos)),
                    "rows": list(dict.fromkeys(format_pos)),
                    "fixable": False,
                    "message": "PO values must start with PO or PM and contain digits only after the prefix.",
                })

            if unusual_length_pos:
                warnings.append(f"PO(s) with a non-standard numeric length (expected 8 digits; showing first 5): {unusual_length_pos[:5]}")
                groups.append({
                    "id": "unusual_po_length",
                    "severity": "warning",
                    "title": "PO numbers with an unusual length",
                    "count": len(set(normalize_po_value(value) for value in unusual_length_pos)),
                    "rows": list(dict.fromkeys(unusual_length_pos)),
                    "fixable": False,
                    "message": "Most PO numbers contain 8 digits after PO or PM. Confirm values with another length.",
                })

            valid_count = clean["__po_key"].nunique()
            if valid_count == 0:
                errors.append("No valid PO entries found after cleaning.")

        return {
            "valid": len(errors) == 0,
            "errors": errors,
            "warnings": warnings,
            "groups": groups,
            "total_rows": len(df),
            "valid_po_count": valid_count if not errors else 0,
            "hierarchy_columns": hierarchy_columns,
            "empty_hierarchy_columns": empty_hierarchy_columns,
            "mapping": mapping,
            "mapping_suggestions": mapping_suggestions,
            "fixes": fixes,
            "file_state": file_state,
        }

    def preview_repair_input_file(self, filepath: str, action: str) -> Dict[str, Any]:
        """Return old/new values before a repair is allowed to modify the input."""
        path = Path(filepath)
        if not path.exists():
            return {"success": False, "error": "File not found."}
        validation = self.validate_input_file(str(path))
        if action == "remove_duplicate_pos" and any(group.get("id") == "po_supplier_conflict" for group in validation.get("groups", [])):
            return {"success": False, "error": "Duplicate removal is blocked because a PO is linked to multiple Suppliers."}
        mapping = validation.get("mapping") or self._mapping_for(str(path))
        try:
            frame = self._read_input_dataframe(path)
        except Exception as exc:
            return {"success": False, "error": f"Could not read the input file: {exc}"}
        po_column = mapping.get("po")
        supplier_column = mapping.get("supplier")
        changes: list[Dict[str, Any]] = []

        def add_change(row: int, column: str, old: Any, new: Any, reason: str) -> None:
            old_text = "" if old is None else str(old)
            new_text = "" if new is None else str(new)
            if old_text != new_text:
                changes.append({"row": row, "column": column, "old": old_text, "new": new_text, "reason": reason})

        if action in {"normalize_required_values", "clean_invalid_chars"} and po_column:
            from src.engine.input_schema import canonicalize_po_value, clean_scalar
            for index, value in frame[po_column].items():
                new = canonicalize_po_value(value)
                add_change(int(index) + 2, str(po_column), value, new, "Canonical PO format")
            if action == "normalize_required_values" and supplier_column:
                for index, value in frame[supplier_column].items():
                    add_change(int(index) + 2, str(supplier_column), value, clean_scalar(value), "Trim Supplier value")
        elif action == "split_multiple_pos" and po_column:
            from src.engine.input_schema import detect_po_parts
            for index, value in frame[po_column].items():
                parts = detect_po_parts(value).get("parts", [])
                if len(parts) > 1:
                    for number, part in enumerate(parts, start=1):
                        changes.append({"row": int(index) + 2, "column": str(po_column), "old": str(value or ""), "new": part["canonical"], "reason": f"Split PO {number}/{len(parts)}"})
        elif action == "remove_blank_rows":
            for index, row in frame.iterrows():
                if all(str(value or "").strip() == "" for value in row.tolist()):
                    changes.append({"row": int(index) + 2, "column": "(entire row)", "old": "row", "new": "removed", "reason": "Blank row"})
        elif action == "remove_duplicate_pos" and po_column:
            seen: set[str] = set()
            from src.engine.input_schema import normalize_po_value
            for index, value in frame[po_column].items():
                key = normalize_po_value(value)
                if key and key in seen:
                    changes.append({"row": int(index) + 2, "column": str(po_column), "old": str(value or ""), "new": "removed", "reason": "Duplicate PO"})
                elif key:
                    seen.add(key)

        state = validation.get("file_state") or self.inspect_input_file(str(path))
        fingerprint = f"{state.get('mtime_ns')}:{state.get('size')}" if state.get("mtime_ns") and state.get("size") else None
        return {
            "success": True,
            "action": action,
            "changes": changes[:500],
            "total_changes": len(changes),
            "expected_fingerprint": fingerprint,
            "message": "Review these changes before saving the input.",
        }

    def repair_input_file(self, filepath: str, actions: List[str], expected_fingerprint: Optional[str] = None) -> Dict[str, Any]:
        """Apply only explicitly requested safe repairs, preserving a backup."""
        path = Path(filepath)
        if not path.exists():
            return {"success": False, "error": "File not found."}
        if expected_fingerprint:
            current = self.inspect_input_file(filepath)
            current_fingerprint = f"{current.get('mtime_ns')}:{current.get('size')}"
            if current_fingerprint != expected_fingerprint:
                return {"success": False, "error": "The input changed after the preview. Validate it again before applying the repair."}
        from src.engine.input_schema import SUPPORTED_INPUT_SUFFIXES
        if path.suffix.lower() not in SUPPORTED_INPUT_SUFFIXES:
            return {"success": False, "error": "Unsupported input format. Use CSV, XLS, XLSX or XLSM."}
        state = self.inspect_input_file(filepath)
        if state.get("open_detected"):
            return {"success": False, "error": "Save and close the input file before applying repairs."}

        actions = set(actions or [])
        if not actions.intersection({"remove_blank_rows", "remove_duplicate_pos", "clean_invalid_chars", "normalize_required_values", "split_multiple_pos"}):
            return {"success": False, "error": "No supported repair was selected."}
        if "remove_duplicate_pos" in actions:
            validation = self.validate_input_file(filepath)
            if any(group.get("id") == "po_supplier_conflict" for group in validation.get("groups", [])):
                return {"success": False, "error": "Duplicate removal is blocked because at least one PO is linked to multiple Suppliers."}

        if path.suffix.lower() == ".xls":
            return {"success": False, "error": "Automatic repair for legacy .xls is unavailable. Save it as .xlsx and try again."}

        removed_blank = 0
        removed_duplicates = 0
        cleaned_chars = 0
        normalized_required_values = 0
        split_rows = 0

        import copy
        import re
        from src.engine.input_schema import canonicalize_po_value, clean_scalar, detect_po_parts, normalize_po_value, resolve_mapping

        char_pattern = re.compile(r"[^A-Za-z0-9_\-]")

        def clean_value(value: str) -> str:
            return canonicalize_po_value(value)

        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook
            workbook = load_workbook(path, keep_vba=path.suffix.lower() == ".xlsm")
            sheet = workbook.worksheets[0]
            headers = [cell.value for cell in sheet[1]]
            header_names = [str(value or "") for value in headers]
            mapping = resolve_mapping(header_names, self._mapping_for(str(path)))
            po_header = mapping.get("po")
            supplier_header = mapping.get("supplier")
            po_column = next(
                (index + 1 for index, value in enumerate(header_names) if value == po_header),
                None,
            )
            supplier_column = next(
                (index + 1 for index, value in enumerate(header_names) if value == supplier_header),
                None,
            )
            if not po_column:
                return {"success": False, "error": "PO Number column not found."}

            split_plan: list[tuple[int, list[str]]] = []
            if "split_multiple_pos" in actions:
                for row_number in range(2, sheet.max_row + 1):
                    details = detect_po_parts(sheet.cell(row_number, po_column).value)
                    if details["multiple"]:
                        canonical_parts = [part["canonical"] for part in details["parts"]]
                        if not all(canonicalize_po_value(part) and re.fullmatch(r"(PO|PM)[0-9]+", part) for part in canonical_parts):
                            return {"success": False, "error": f"Could not safely split PO values on row {row_number}."}
                        split_plan.append((row_number, canonical_parts))

            timestamp = time.strftime("%Y%m%d-%H%M%S")
            backup = path.with_name(f"{path.stem}.backup_{timestamp}{path.suffix}")
            shutil.copy2(path, backup)
            rows_to_remove: set[int] = set()
            if split_plan:
                for row_number, canonical_parts in reversed(split_plan):
                    source_values = [sheet.cell(row_number, column).value for column in range(1, sheet.max_column + 1)]
                    source_styles = [copy.copy(sheet.cell(row_number, column)._style) for column in range(1, sheet.max_column + 1)]
                    sheet.cell(row_number, po_column).value = canonical_parts[0]
                    for offset, po_value in enumerate(canonical_parts[1:], start=1):
                        target_row = row_number + offset
                        sheet.insert_rows(target_row)
                        for column, value in enumerate(source_values, start=1):
                            target = sheet.cell(target_row, column)
                            target.value = value
                            target._style = copy.copy(source_styles[column - 1])
                        sheet.cell(target_row, po_column).value = po_value
                        split_rows += 1
            seen_pos: set[str] = set()
            for row_number in range(2, sheet.max_row + 1):
                values = [sheet.cell(row_number, column).value for column in range(1, sheet.max_column + 1)]
                is_blank = all(value is None or str(value).strip() == "" for value in values)
                if is_blank and "remove_blank_rows" in actions:
                    rows_to_remove.add(row_number)
                    removed_blank += 1
                    continue
                po_value = clean_scalar(sheet.cell(row_number, po_column).value)
                if "normalize_required_values" in actions:
                    for column in (po_column, supplier_column):
                        if column:
                            cell = sheet.cell(row_number, column)
                            raw_value = cell.value
                            cleaned_value = canonicalize_po_value(raw_value) if column == po_column else clean_scalar(raw_value)
                            if cleaned_value and str(raw_value) != cleaned_value:
                                cell.value = cleaned_value
                                normalized_required_values += 1
                if not po_value:
                    continue
                if "clean_invalid_chars" in actions and po_value != clean_value(po_value):
                    sheet.cell(row_number, po_column).value = clean_value(po_value)
                    cleaned_chars += 1
                    po_value = clean_value(po_value)
                if "remove_duplicate_pos" in actions:
                    po_key = normalize_po_value(po_value)
                    if po_key in seen_pos:
                        rows_to_remove.add(row_number)
                        removed_duplicates += 1
                    else:
                        seen_pos.add(po_key)
            for row_number in sorted(rows_to_remove, reverse=True):
                sheet.delete_rows(row_number)
            workbook.save(path)
        else:
            frame = self._read_input_dataframe(path)
            separator = ";" if path.suffix.lower() != ".csv" else ";"
            if path.suffix.lower() == ".csv":
                raw = path.read_bytes()[:8192].decode(self._input_encoding(path), errors="replace")
                from src.engine.input_schema import detect_csv_separator
                separator = detect_csv_separator(raw)
            mapping = resolve_mapping([str(column) for column in frame.columns], self._mapping_for(str(path)))
            po_column = mapping.get("po")
            supplier_column = mapping.get("supplier")
            if not po_column:
                return {"success": False, "error": "PO Number column not found."}
            timestamp = time.strftime("%Y%m%d-%H%M%S")
            backup = path.with_name(f"{path.stem}.backup_{timestamp}{path.suffix}")
            shutil.copy2(path, backup)
            blank_mask = frame.apply(lambda row: all(clean_scalar(value) == "" for value in row), axis=1)
            if "remove_blank_rows" in actions:
                removed_blank = int(blank_mask.sum())
                frame = frame.loc[~blank_mask]
            if "split_multiple_pos" in actions:
                expanded_rows = []
                for _, row in frame.iterrows():
                    details = detect_po_parts(row.get(po_column, ""))
                    if not details["multiple"]:
                        expanded_rows.append(row.to_dict())
                        continue
                    canonical_parts = [part["canonical"] for part in details["parts"]]
                    for po_value in canonical_parts:
                        copied = row.to_dict()
                        copied[po_column] = po_value
                        expanded_rows.append(copied)
                        split_rows += 1 if po_value != canonical_parts[0] else 0
                frame = pd.DataFrame(expanded_rows, columns=frame.columns)
            if "normalize_required_values" in actions:
                for column in (po_column, supplier_column):
                    if column and column in frame.columns:
                        normalized_values = frame[column].map(canonicalize_po_value if column == po_column else clean_scalar)
                        raw_values = frame[column].map(lambda value: "" if value is None else str(value))
                        mask = normalized_values.ne("") & raw_values.ne(normalized_values)
                        frame.loc[mask, column] = normalized_values.loc[mask]
                        normalized_required_values += int(mask.sum())
            if "clean_invalid_chars" in actions:
                raw_values = frame[po_column].map(clean_scalar)
                normalized_values = raw_values.map(clean_value)
                mask = raw_values.ne(normalized_values)
                frame.loc[mask, po_column] = normalized_values.loc[mask]
                cleaned_chars = int(mask.sum())
            if "remove_duplicate_pos" in actions:
                duplicate_mask = frame[po_column].map(normalize_po_value).duplicated(keep="first")
                removed_duplicates = int(duplicate_mask.sum())
                frame = frame.loc[~duplicate_mask]
            frame.to_csv(path, sep=separator, index=False, encoding="utf-8-sig")

        return {
            "success": True,
            "backup_path": str(backup),
            "removed_blank_rows": removed_blank,
            "removed_duplicate_rows": removed_duplicates,
            "cleaned_invalid_chars": cleaned_chars,
            "normalized_required_values": normalized_required_values,
            "split_rows": split_rows,
            "message": "Safe repairs applied. The original file was backed up.",
        }

    def import_file(self, filepath: str) -> Dict[str, Any]:
        import re
        from src.engine.input_schema import canonicalize_po_value, clean_scalar, columns_of_dataframe, detect_po_parts, is_excel_numeric_coercion, is_placeholder_po, is_placeholder_supplier, is_valid_canonical_po, normalize_po_value, normalize_supplier_value, resolve_data_mapping, resolve_mapping
        try:
            if not os.path.exists(filepath):
                return {'success': False, 'error': f"File not found: {filepath}"}

            _, ext = os.path.splitext(filepath.lower())
            if ext not in {'.csv', '.xls', '.xlsx', '.xlsm'}:
                return {'success': False, 'error': "Unsupported file format. Use XLSX, XLSM, XLS or CSV."}
            df = self._read_input_dataframe(Path(filepath))

            def clean_folder_part(value: Any) -> str:
                text = str(value or '').strip()
                if text.lower() in {'', 'nan', 'none'}:
                    text = 'Unknown'
                text = text.replace('/', '_').replace('\\', '_')
                text = '_'.join(text.split())
                text = text.strip('._')
                return text or 'Unknown'

            def extract_hierarchy_columns(frame: pd.DataFrame, exclude: set[str]) -> tuple[list[str], bool]:
                sep_col = None
                for c in frame.columns:
                    if str(c).strip() == '<|>':
                        sep_col = c
                        break

                cols = [str(c) for c in frame.columns]
                if sep_col is not None:
                    hierarchy_cols = cols[cols.index(str(sep_col)) + 1:]
                else:
                    # Non-standard inputs: all columns except PO/Supplier.
                    hierarchy_cols = [c for c in cols if c not in exclude]
                if not hierarchy_cols:
                    return [], False

                populated = []
                for col in hierarchy_cols:
                    series = frame[col].fillna('').astype(str).str.strip()
                    series = series[(series != '') & (series.str.lower() != 'nan')]
                    if not series.empty:
                        populated.append(col)
                return hierarchy_cols, bool(populated)

            mapping, _ = resolve_data_mapping(df, columns_of_dataframe(df), self._mapping_for(filepath))
            po_col = mapping.get('po')
            company_col = mapping.get('supplier')
            hierarchy_cols, has_hierarchy_data = extract_hierarchy_columns(df, {po_col, company_col})

            if not po_col:
                return {'success': False, 'error': f"Could not find PO Number column. Found columns: {list(df.columns)}"}
            if not company_col:
                return {'success': False, 'error': f"Could not find Company Code/Legal Entity/Supplier column. Found columns: {list(df.columns)}"}

            df = df.copy()
            df[po_col] = df[po_col].map(clean_scalar)
            df[company_col] = df[company_col].map(clean_scalar)
            df = df[(df[po_col] != "") & (df[company_col] != "")]
            invalid_identifier = df[po_col].map(is_placeholder_po) | df[po_col].map(is_excel_numeric_coercion)
            invalid_multi = df[po_col].map(lambda value: detect_po_parts(value)["multiple"] or detect_po_parts(value)["ambiguous"])
            invalid_supplier = df[company_col].map(is_placeholder_supplier)
            invalid_format = df[po_col].map(lambda value: not is_valid_canonical_po(value))
            if invalid_identifier.any() or invalid_supplier.any() or invalid_multi.any() or invalid_format.any():
                return {'success': False, 'error': 'The input contains invalid, ambiguous or placeholder required values.'}
            df[po_col] = df[po_col].map(canonicalize_po_value)
            df["__po_key"] = df[po_col].map(normalize_po_value)
            df["__supplier_key"] = df[company_col].map(normalize_supplier_value)
            conflicts = df.groupby("__po_key")["__supplier_key"].nunique()
            if any(conflicts > 1):
                return {'success': False, 'error': 'At least one PO is linked to multiple Suppliers.'}
            df = df.drop_duplicates(subset=["__po_key"], keep="first")
            df = df.sort_values(by=company_col, kind='stable')
            session_id = self.db.create_session(os.path.basename(filepath))

            total_pos = 0
            for _, row in df.iterrows():
                po_val = str(row[po_col]).strip()
                company_val = str(row[company_col]).strip()
                if po_val and po_val.lower() != 'nan' and company_val and company_val.lower() != 'nan':
                    # Supplier is always the first folder level; optional
                    # hierarchy columns come next (PO stays at the file level).
                    parts = [clean_folder_part(company_val)]
                    if has_hierarchy_data and hierarchy_cols:
                        parts.extend(clean_folder_part(row.get(col, '')) for col in hierarchy_cols)
                    output_subdir = PurePosixPath(*parts).as_posix()

                    self.db.add_po(PODownload(
                        session_id=session_id,
                        po_number=po_val,
                        company_code=company_val,
                        output_subdir=output_subdir,
                        status="PENDING"
                    ))
                    total_pos += 1

            return {
                'success': True,
                'session_id': session_id,
                'total_pos': total_pos
            }

        except Exception as e:
            return {'success': False, 'error': str(e)}

    def _session_summary(self, session_id: int) -> Dict[str, int]:
        cursor = self.db.conn.cursor()
        row = cursor.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN status = 'SUCCESS' THEN 1 ELSE 0 END) AS success,
                SUM(CASE WHEN status IN ('ERROR', 'SKIPPED_VERIFICATION_REQUIRED') THEN 1 ELSE 0 END) AS errors,
                SUM(CASE WHEN status = 'PENDING' THEN 1 ELSE 0 END) AS pending
            FROM po_downloads WHERE session_id = ?
            """,
            (session_id,),
        ).fetchone()
        return {
            "total": int(row["total"] or 0),
            "success": int(row["success"] or 0),
            "errors": int(row["errors"] or 0),
            "pending": int(row["pending"] or 0),
        }

    def get_session_history(self) -> List[Dict[str, Any]]:
        cursor = self.db.conn.cursor()
        cursor.execute("SELECT * FROM sessions ORDER BY id DESC")
        history: list[Dict[str, Any]] = []
        for row in cursor.fetchall():
            item = dict(row)
            stats = self._session_summary(int(item["id"]))
            item.update({
                "total_pos": stats["total"],
                "success_count": stats["success"],
                "error_count": stats["errors"],
                "pending_count": stats["pending"],
            })
            if item["status"] in {"PENDING", "RUNNING"} and stats["total"] > 0 and stats["pending"] == 0:
                if stats["errors"] == 0:
                    item["status"] = "SUCCESS"
                elif stats["success"] > 0:
                    item["status"] = "PARTIAL"
                else:
                    item["status"] = "FAILED"
            history.append(item)
        return history

    def get_session_details(self, session_id: int) -> Dict[str, Any]:
        cursor = self.db.conn.cursor()
        cursor.execute("SELECT * FROM po_downloads WHERE session_id = ?", (session_id,))
        pos = [dict(row) for row in cursor.fetchall()]
        session = self.db.get_session(session_id)
        return {
            'session': session,
            'pos': pos
        }

    def confirm_and_retry_company(self, session_id: int, company_code: str) -> Dict[str, Any]:
        try:
            cursor = self.db.conn.cursor()
            cursor.execute('''
                UPDATE po_downloads
                SET status = 'PENDING', updated_at = strftime('%Y-%m-%d %H:%M:%f', 'now')
                WHERE session_id = ? AND company_code = ? AND status = 'SKIPPED_VERIFICATION_REQUIRED'
            ''', (session_id, company_code))
            self.db.conn.commit()
            return {'success': True}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def export_session_report(self, session_id: int, dest_filepath: str) -> Dict[str, Any]:
        try:
            cursor = self.db.conn.cursor()
            cursor.execute('''
                SELECT po_number, company_code, status, download_folder, attachment_count, error_message, updated_at
                FROM po_downloads
                WHERE session_id = ?
            ''', (session_id,))
            rows = cursor.fetchall()
            df = pd.DataFrame([dict(r) for r in rows])

            df.to_excel(dest_filepath, index=False)
            from src.db.coupa_metadata import CoupaMetadataRepository
            from src.reports.coupa_excel import enrich_excel_report

            metadata_repository = CoupaMetadataRepository(self.db)
            enrich_excel_report(
                dest_filepath,
                metadata_repository.list_po_metadata(session_id),
                metadata_repository.list_line_metadata(session_id),
            )
            return {'success': True, 'filepath': dest_filepath}
        except Exception as e:
            return {'success': False, 'error': str(e)}


def _check_runtime_stop(
    session_id: int,
    runtime: Dict[int, Dict[str, Any]],
    lock: threading.Lock,
) -> None:
    with lock:
        rt = runtime.get(session_id)
        if rt and rt.get("stop_requested"):
            rt["status"] = "STOPPED"
            raise SessionStoppedError(session_id)


async def _check_runtime_pause(
    session_id: int,
    runtime: Dict[int, Dict[str, Any]],
    lock: threading.Lock,
) -> None:
    while True:
        with lock:
            rt = runtime.get(session_id)
            if not rt:
                return
            if rt.get("stop_requested"):
                rt["status"] = "STOPPED"
                raise SessionStoppedError(session_id)
            if not rt.get("paused", False):
                return
        await asyncio.sleep(0.2)


class SessionStoppedError(Exception):
    def __init__(self, session_id: int):
        self.session_id = session_id
        super().__init__(f"Session {session_id} stopped by user")
