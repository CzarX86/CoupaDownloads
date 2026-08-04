"""Excel enrichment for Coupa metadata.

This module receives plain mappings and knows nothing about crawling or SQLite.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SUMMARY_COLUMNS = (
    "COUPA_COMPANY_CODE",
    "COUPA_SHIP_TO_USER",
    "COUPA_PAYMENT_TERM",
    "COUPA_CRG_CODES",
    "COUPA_CRG_RAW",
    "COUPA_CURRENCY_CODES",
    "COUPA_METADATA_STATUS",
    "COUPA_METADATA_SCRAPED_AT",
    "COUPA_METADATA_ERROR",
)
LINE_COLUMNS = (
    "PO_NUMBER",
    "LINE_NUMBER",
    "COUPA_CRG_CODE",
    "COUPA_CRG_RAW",
    "COUPA_CURRENCY_CODE",
)


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _po_key(value: Any) -> str:
    return _text(value).upper()


def _unique_join(values: Iterable[Any]) -> str:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = _text(value)
        if not text or text in seen:
            continue
        result.append(text)
        seen.add(text)
    return "; ".join(result)


def _style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="173B56")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    sheet.freeze_panes = "A2"


def _ensure_columns(sheet, columns: Iterable[str]) -> dict[str, int]:
    headers = {_text(cell.value): cell.column for cell in sheet[1] if cell.value is not None}
    for column in columns:
        if column not in headers:
            index = sheet.max_column + 1
            sheet.cell(1, index, column)
            headers[column] = index
    return headers


def enrich_excel_report(
    report_path: str | Path,
    po_metadata: Iterable[Mapping[str, Any]],
    line_metadata: Iterable[Mapping[str, Any]],
) -> str:
    """Append Coupa summary columns and a normalized ``COUPA_LINES`` sheet."""
    path = Path(report_path)
    workbook = load_workbook(path)
    summary = workbook.worksheets[0]
    summary_headers = _ensure_columns(summary, SUMMARY_COLUMNS)

    metadata_by_po = {
        _po_key(row.get("po_number")): row
        for row in po_metadata
        if _po_key(row.get("po_number"))
    }
    lines_by_po: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    all_lines = list(line_metadata)
    for row in all_lines:
        key = _po_key(row.get("po_number"))
        if key:
            lines_by_po[key].append(row)

    po_column = next(
        (index for name, index in {_text(cell.value): cell.column for cell in summary[1]}.items() if name.upper() == "PO_NUMBER"),
        None,
    )
    if po_column is not None:
        for row_number in range(2, summary.max_row + 1):
            po_key = _po_key(summary.cell(row_number, po_column).value)
            metadata = metadata_by_po.get(po_key, {})
            lines = lines_by_po.get(po_key, [])
            values = {
                "COUPA_COMPANY_CODE": metadata.get("company_code", ""),
                "COUPA_SHIP_TO_USER": metadata.get("ship_to_user", ""),
                "COUPA_PAYMENT_TERM": metadata.get("payment_term", ""),
                "COUPA_CRG_CODES": _unique_join(line.get("crg_code") for line in lines),
                "COUPA_CRG_RAW": _unique_join(line.get("crg_raw") for line in lines),
                "COUPA_CURRENCY_CODES": _unique_join(line.get("currency_code") for line in lines),
                "COUPA_METADATA_STATUS": metadata.get("metadata_status", "NOT_EXTRACTED"),
                "COUPA_METADATA_SCRAPED_AT": metadata.get("scraped_at", ""),
                "COUPA_METADATA_ERROR": metadata.get("metadata_error", ""),
            }
            for name, value in values.items():
                summary.cell(row_number, summary_headers[name], value or "")

    if "COUPA_LINES" in workbook.sheetnames:
        del workbook["COUPA_LINES"]
    lines_sheet = workbook.create_sheet("COUPA_LINES")
    lines_sheet.append(LINE_COLUMNS)
    for row in all_lines:
        lines_sheet.append(tuple(row.get(column.removeprefix("COUPA_").lower(), "") for column in LINE_COLUMNS))

    # The line repository keys are named without the COUPA_ prefix, while the
    # report columns intentionally use the prefix to avoid collisions with the
    # user's input columns.
    for row_number, row in enumerate(all_lines, start=2):
        lines_sheet.cell(row_number, 1, row.get("po_number", ""))
        lines_sheet.cell(row_number, 2, row.get("line_number", ""))
        lines_sheet.cell(row_number, 3, row.get("crg_code", ""))
        lines_sheet.cell(row_number, 4, row.get("crg_raw", ""))
        lines_sheet.cell(row_number, 5, row.get("currency_code", ""))

    _style_header(summary)
    _style_header(lines_sheet)
    summary.auto_filter.ref = f"A1:{get_column_letter(summary.max_column)}{max(summary.max_row, 1)}"
    lines_sheet.auto_filter.ref = f"A1:{get_column_letter(lines_sheet.max_column)}{max(lines_sheet.max_row, 1)}"
    for sheet in (summary, lines_sheet):
        for column_cells in sheet.columns:
            values = [len(_text(cell.value)) for cell in column_cells[:200]]
            width = min(max(max(values, default=10) + 2, 12), 55)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    workbook.save(path)
    return str(path)
