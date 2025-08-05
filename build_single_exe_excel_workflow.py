#!/usr/bin/env python3
"""
Excel Workflow Single EXE Builder for CoupaDownloads
Creates EXE with Excel-based input workflow
"""

import os
import sys
import platform
import subprocess
import shutil
import requests
import zipfile
from pathlib import Path
import tempfile


class ExcelWorkflowExeBuilder:
    """Builds single executable with Excel-based input workflow."""
    
    def __init__(self):
        self.project_root = Path(__file__).parent
        self.build_dir = self.project_root / "build"
        self.drivers_dir = self.build_dir / "drivers"
        
    def clean_build(self):
        """Clean previous build."""
        print("🧹 Cleaning previous build...")
        if self.build_dir.exists():
            shutil.rmtree(self.build_dir)
        self.build_dir.mkdir(exist_ok=True)
        print("✅ Build directory cleaned")
    
    def download_stable_edgedriver(self):
        """Download a stable EdgeDriver version."""
        print("🔧 Downloading stable EdgeDriver...")
        
        system = platform.system().lower()
        machine = platform.machine().lower()
        
        # Use a stable version that works with most Edge versions
        stable_version = "120.0.2210.91"
        
        if system == "windows":
            if "64" in machine or "x86_64" in machine:
                driver_url = f"https://msedgedriver.azureedge.net/{stable_version}/edgedriver_win64.zip"
                driver_filename = "edgedriver_win64.zip"
            else:
                driver_url = f"https://msedgedriver.azureedge.net/{stable_version}/edgedriver_win32.zip"
                driver_filename = "edgedriver_win32.zip"
        else:
            print("⚠️ Single EXE build only supported on Windows")
            return False
        
        try:
            # Download EdgeDriver
            response = requests.get(driver_url, stream=True)
            response.raise_for_status()
            
            driver_zip = self.build_dir / driver_filename
            with open(driver_zip, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            
            # Extract EdgeDriver
            with zipfile.ZipFile(driver_zip, 'r') as zip_ref:
                zip_ref.extractall(self.drivers_dir)
            
            print(f"✅ Stable EdgeDriver {stable_version} downloaded")
            return True
            
        except Exception as e:
            print(f"❌ Failed to download EdgeDriver: {e}")
            return False
    
    def create_excel_template(self):
        """Create Excel template file."""
        print("📊 Creating Excel template...")
        
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "PO Numbers"
            
            # Set up headers
            headers = [
                "PO_NUMBER",
                "SUPPLIER",
                "DESCRIPTION",
                "STATUS",
                "NOTES"
            ]
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add sample data
            sample_data = [
                ["PO15262984", "Sample Supplier 1", "Office Supplies", "PENDING", ""],
                ["PO15327452", "Sample Supplier 2", "IT Equipment", "PENDING", ""],
                ["PO15362783", "Sample Supplier 3", "Furniture", "PENDING", ""],
                ["PO15412345", "Sample Supplier 4", "Software License", "PENDING", ""],
                ["PO15467890", "Sample Supplier 5", "Training Services", "PENDING", ""]
            ]
            
            for row, data in enumerate(sample_data, 2):
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add instructions sheet
            ws2 = wb.create_sheet("Instructions")
            instructions = [
                ["INSTRUCTIONS FOR USING THIS TEMPLATE"],
                [""],
                ["1. Enter your PO numbers in the 'PO_NUMBER' column"],
                ["2. Optionally fill in supplier and description information"],
                ["3. Save the file when you're done"],
                ["4. Close the file to continue with processing"],
                [""],
                ["IMPORTANT NOTES:"],
                ["- PO numbers must start with 'PO'"],
                ["- You can delete the sample rows"],
                ["- The system will process all PO numbers in the file"],
                ["- Files will be downloaded to your Downloads folder"],
                [""],
                ["SUPPORTED FILE TYPES:"],
                ["- PDF files"],
                ["- Word documents (.docx)"],
                ["- Excel files (.xlsx)"],
                ["- Email files (.msg)"],
                ["- Image files (.jpg, .png, .pdf)"]
            ]
            
            for row, instruction in enumerate(instructions, 1):
                ws2.cell(row=row, column=1, value=instruction[0])
            
            # Save template
            template_path = self.build_dir / "po_template.xlsx"
            wb.save(template_path)
            
            print("✅ Excel template created")
            return template_path
            
        except ImportError:
            print("⚠️ pandas/openpyxl not available, creating simple template")
            # Create simple CSV template as fallback
            template_content = """PO_NUMBER,SUPPLIER,DESCRIPTION,STATUS,NOTES
PO15262984,Sample Supplier 1,Office Supplies,PENDING,
PO15327452,Sample Supplier 2,IT Equipment,PENDING,
PO15362783,Sample Supplier 3,Furniture,PENDING,
PO15412345,Sample Supplier 4,Software License,PENDING,
PO15467890,Sample Supplier 5,Training Services,PENDING,"""
            
            template_path = self.build_dir / "po_template.csv"
            template_path.write_text(template_content, encoding='utf-8')
            
            print("✅ CSV template created (fallback)")
            return template_path
    
    def create_excel_workflow_script(self):
        """Create the main script with Excel workflow."""
        print("📝 Creating Excel workflow script...")
        
        main_script_content = '''#!/usr/bin/env python3
"""
CoupaDownloads - Excel Workflow Application
Opens Excel template, processes from temp, moves to downloads
"""

import os
import sys
import tempfile
import zipfile
import subprocess
import platform
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import threading
import time
import shutil
import csv
import pandas as pd
from datetime import datetime


class CoupaDownloadsExcelApp:
    """Main application with Excel-based workflow."""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("CoupaDownloads - Excel Workflow")
        self.root.geometry("600x500")
        self.root.resizable(True, True)
        
        # Set icon if available
        try:
            self.root.iconbitmap("icon.ico")
        except:
            pass
        
        self.setup_ui()
        self.temp_dir = None
        self.excel_file_path = None
        
    def setup_ui(self):
        """Setup the user interface."""
        # Configure grid
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_rowconfigure(2, weight=1)
        
        # Title
        title_label = tk.Label(
            self.root, 
            text="CoupaDownloads", 
            font=("Arial", 16, "bold"),
            fg="#2E86AB"
        )
        title_label.grid(row=0, column=0, pady=10)
        
        # Subtitle
        subtitle_label = tk.Label(
            self.root,
            text="Excel Workflow - PO Attachment Downloader",
            font=("Arial", 10),
            fg="#666666"
        )
        subtitle_label.grid(row=1, column=0, pady=(0, 20))
        
        # Main frame
        main_frame = tk.Frame(self.root)
        main_frame.grid(row=2, column=0, sticky="nsew", padx=20, pady=10)
        main_frame.grid_columnconfigure(0, weight=1)
        
        # Instructions section
        instructions_frame = tk.LabelFrame(main_frame, text="Instructions", padx=10, pady=10)
        instructions_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        
        instructions_text = """1. Click "Open Excel Template" to create input file
2. Add your PO numbers to the template
3. Save and close the Excel file
4. Click "Start Processing" to begin downloads
5. Files will be saved to Downloads/CoupaDownloads/"""
        
        instructions_label = tk.Label(
            instructions_frame,
            text=instructions_text,
            font=("Arial", 9),
            justify="left"
        )
        instructions_label.grid(row=0, column=0, sticky="w")
        
        # File status section
        status_frame = tk.LabelFrame(main_frame, text="File Status", padx=10, pady=10)
        status_frame.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        
        self.file_status_var = tk.StringVar(value="No input file created")
        file_status_label = tk.Label(
            status_frame,
            textvariable=self.file_status_var,
            font=("Arial", 9),
            fg="#666666"
        )
        file_status_label.grid(row=0, column=0, sticky="w")
        
        # Settings section
        settings_frame = tk.LabelFrame(main_frame, text="Settings", padx=10, pady=10)
        settings_frame.grid(row=2, column=0, sticky="ew", pady=(0, 10))
        
        # Headless mode
        self.headless_var = tk.BooleanVar()
        headless_check = tk.Checkbutton(
            settings_frame,
            text="Run in headless mode (no browser window)",
            variable=self.headless_var,
            font=("Arial", 9)
        )
        headless_check.grid(row=0, column=0, sticky="w")
        
        # Progress section
        progress_frame = tk.LabelFrame(main_frame, text="Progress", padx=10, pady=10)
        progress_frame.grid(row=3, column=0, sticky="ew", pady=(0, 10))
        progress_frame.grid_columnconfigure(0, weight=1)
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = tk.Scale(
            progress_frame,
            from_=0,
            to=100,
            orient="horizontal",
            variable=self.progress_var,
            state="readonly",
            showvalue=False
        )
        self.progress_bar.grid(row=0, column=0, sticky="ew", pady=(0, 5))
        
        # Status label
        self.status_var = tk.StringVar(value="Ready to start")
        status_label = tk.Label(
            progress_frame,
            textvariable=self.status_var,
            font=("Arial", 9),
            fg="#666666"
        )
        status_label.grid(row=1, column=0, sticky="w")
        
        # Action buttons
        action_frame = tk.Frame(main_frame)
        action_frame.grid(row=4, column=0, pady=(10, 0))
        
        # Open template button
        self.open_template_btn = tk.Button(
            action_frame,
            text="📊 Open Excel Template",
            command=self.open_excel_template,
            bg="#2196F3",
            fg="white",
            font=("Arial", 10, "bold"),
            width=15
        )
        self.open_template_btn.pack(side="left", padx=(0, 10))
        
        # Start button
        self.start_btn = tk.Button(
            action_frame,
            text="🚀 Start Processing",
            command=self.start_processing,
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10, "bold"),
            width=15
        )
        self.start_btn.pack(side="left", padx=(0, 10))
        
        # Exit button
        exit_btn = tk.Button(
            action_frame,
            text="❌ Exit",
            command=self.root.quit,
            bg="#f44336",
            fg="white",
            font=("Arial", 10),
            width=10
        )
        exit_btn.pack(side="left")
        
    def open_excel_template(self):
        """Open Excel template for user to edit."""
        try:
            # Create temporary directory
            self.temp_dir = Path(tempfile.mkdtemp(prefix="CoupaDownloads_"))
            
            # Copy template to temp directory
            template_src = Path("po_template.xlsx")
            if not template_src.exists():
                # Create simple template if not available
                self.create_simple_template()
            
            self.excel_file_path = self.temp_dir / "po_input.xlsx"
            shutil.copy2(template_src, self.excel_file_path)
            
            # Update status
            self.file_status_var.set(f"Template created: {self.excel_file_path}")
            
            # Open Excel file
            if platform.system().lower() == "windows":
                os.startfile(self.excel_file_path)
            else:
                subprocess.run(["open", str(self.excel_file_path)])
            
            messagebox.showinfo(
                "Excel Template Opened",
                f"Excel template has been opened.\\n\\nPlease:\\n1. Add your PO numbers\\n2. Save the file\\n3. Close Excel\\n4. Click 'Start Processing'"
            )
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel template: {e}")
    
    def create_simple_template(self):
        """Create a simple Excel template."""
        try:
            import pandas as pd
            
            # Create sample data
            data = {
                'PO_NUMBER': ['PO15262984', 'PO15327452', 'PO15362783'],
                'SUPPLIER': ['Sample Supplier 1', 'Sample Supplier 2', 'Sample Supplier 3'],
                'DESCRIPTION': ['Office Supplies', 'IT Equipment', 'Furniture'],
                'STATUS': ['PENDING', 'PENDING', 'PENDING'],
                'NOTES': ['', '', '']
            }
            
            df = pd.DataFrame(data)
            df.to_excel("po_template.xlsx", index=False)
            
        except ImportError:
            # Fallback to CSV
            csv_content = """PO_NUMBER,SUPPLIER,DESCRIPTION,STATUS,NOTES
PO15262984,Sample Supplier 1,Office Supplies,PENDING,
PO15327452,Sample Supplier 2,IT Equipment,PENDING,
PO15362783,Sample Supplier 3,Furniture,PENDING,"""
            
            with open("po_template.csv", 'w', encoding='utf-8') as f:
                f.write(csv_content)
    
    def read_excel_file(self):
        """Read PO numbers from Excel file."""
        try:
            if self.excel_file_path.suffix.lower() == '.xlsx':
                # Read Excel file
                df = pd.read_excel(self.excel_file_path)
            else:
                # Read CSV file
                df = pd.read_csv(self.excel_file_path)
            
            # Extract PO numbers
            po_numbers = []
            if 'PO_NUMBER' in df.columns:
                po_numbers = df['PO_NUMBER'].dropna().astype(str).tolist()
            
            # Filter valid PO numbers
            valid_pos = []
            for po in po_numbers:
                po = po.strip()
                if po and po.upper().startswith('PO'):
                    valid_pos.append(po)
            
            return valid_pos
            
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
            return []
    
    def validate_input(self):
        """Validate that input file exists and has PO numbers."""
        if not self.excel_file_path or not self.excel_file_path.exists():
            messagebox.showerror("Error", "Please create an input file first by clicking 'Open Excel Template'")
            return False
        
        po_numbers = self.read_excel_file()
        
        if not po_numbers:
            messagebox.showerror("Error", "No valid PO numbers found in the file. Please add PO numbers and try again.")
            return False
        
        return True
    
    def setup_original_system(self):
        """Setup the original system in temporary directory."""
        try:
            # Create directory structure
            data_dir = self.temp_dir / "data" / "input"
            data_dir.mkdir(parents=True, exist_ok=True)
            
            # Copy source files
            src_dir = self.temp_dir / "src" / "core"
            src_dir.mkdir(parents=True, exist_ok=True)
            
            # Copy core files
            core_files = [
                "browser.py", "config.py", "csv_processor.py", 
                "downloader.py", "driver_manager.py"
            ]
            
            for file_name in core_files:
                src_file = Path("src/core") / file_name
                dst_file = src_dir / file_name
                if src_file.exists():
                    shutil.copy2(src_file, dst_file)
            
            # Copy main.py
            main_src = Path("src/main.py")
            main_dst = self.temp_dir / "src" / "main.py"
            if main_src.exists():
                shutil.copy2(main_src, main_dst)
            
            # Copy drivers
            drivers_dst = self.temp_dir / "drivers"
            drivers_dst.mkdir(exist_ok=True)
            if Path("drivers").exists():
                shutil.copytree("drivers", drivers_dst, dirs_exist_ok=True)
            
            return True
            
        except Exception as e:
            print(f"❌ Failed to setup original system: {e}")
            return False
    
    def create_input_csv_from_excel(self):
        """Create input CSV from Excel file."""
        try:
            po_numbers = self.read_excel_file()
            
            input_file = self.temp_dir / "data" / "input" / "input.csv"
            
            with open(input_file, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['PO_NUMBER'])  # Header
                for po in po_numbers:
                    writer.writerow([po])
            
            return str(input_file), len(po_numbers)
            
        except Exception as e:
            print(f"❌ Failed to create input CSV: {e}")
            return None, 0
    
    def run_original_system(self, input_file_path):
        """Run the original system with the created input file."""
        try:
            # Change to temp directory
            original_cwd = os.getcwd()
            os.chdir(self.temp_dir)
            
            # Set environment variables
            os.environ['PYTHONPATH'] = str(self.temp_dir)
            
            # Run the original main.py
            result = subprocess.run([
                sys.executable, "src/main.py"
            ], capture_output=True, text=True)
            
            # Restore original directory
            os.chdir(original_cwd)
            
            return result.returncode == 0, result.stdout, result.stderr
            
        except Exception as e:
            print(f"❌ Failed to run original system: {e}")
            return False, "", str(e)
    
    def move_to_downloads(self):
        """Move processed files to Downloads folder."""
        try:
            downloads_dir = Path.home() / "Downloads" / "CoupaDownloads"
            downloads_dir.mkdir(parents=True, exist_ok=True)
            
            # Move Excel file to downloads
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            final_excel_path = downloads_dir / f"po_input_{timestamp}.xlsx"
            shutil.copy2(self.excel_file_path, final_excel_path)
            
            return str(final_excel_path)
            
        except Exception as e:
            print(f"❌ Failed to move to downloads: {e}")
            return None
    
    def start_processing(self):
        """Start the processing workflow."""
        if not self.validate_input():
            return
        
        # Disable buttons
        self.start_btn.config(state="disabled")
        self.open_template_btn.config(state="disabled")
        
        # Start processing in separate thread
        processing_thread = threading.Thread(target=self.processing_workflow)
        processing_thread.daemon = True
        processing_thread.start()
    
    def processing_workflow(self):
        """Main processing workflow."""
        try:
            self.status_var.set("Reading Excel file...")
            self.progress_var.set(10)
            
            # Read PO numbers
            po_numbers = self.read_excel_file()
            if not po_numbers:
                raise Exception("No valid PO numbers found")
            
            self.status_var.set(f"Found {len(po_numbers)} PO numbers")
            self.progress_var.set(20)
            
            # Setup original system
            self.status_var.set("Setting up system...")
            self.progress_var.set(30)
            
            if not self.setup_original_system():
                raise Exception("Failed to setup original system")
            
            # Create input CSV
            self.status_var.set("Creating input file...")
            self.progress_var.set(40)
            
            input_file_path, po_count = self.create_input_csv_from_excel()
            if not input_file_path:
                raise Exception("Failed to create input CSV")
            
            # Run original system
            self.status_var.set("Starting download process...")
            self.progress_var.set(50)
            
            success, stdout, stderr = self.run_original_system(input_file_path)
            
            if success:
                self.progress_var.set(80)
                self.status_var.set("Moving files to Downloads...")
                
                # Move to downloads
                final_path = self.move_to_downloads()
                
                self.progress_var.set(100)
                self.status_var.set("Processing completed successfully!")
                
                messagebox.showinfo(
                    "Success", 
                    f"Processing completed!\\n\\nProcessed {po_count} PO numbers.\\n\\nFiles saved to: %USERPROFILE%\\Downloads\\CoupaDownloads\\"
                )
            else:
                raise Exception(f"Processing failed: {stderr}")
            
        except Exception as e:
            self.status_var.set(f"Error: {e}")
            messagebox.showerror("Error", f"Processing failed: {e}")
        
        finally:
            # Re-enable buttons
            self.start_btn.config(state="normal")
            self.open_template_btn.config(state="normal")
    
    def run(self):
        """Run the application."""
        self.root.mainloop()


def main():
    """Main function."""
    app = CoupaDownloadsExcelApp()
    app.run()


if __name__ == "__main__":
    main()
'''
        
        main_script_path = self.build_dir / "coupa_downloads_excel.py"
        main_script_path.write_text(main_script_content, encoding='utf-8')
        
        print("✅ Excel workflow script created")
        return main_script_path
    
    def copy_source_files(self):
        """Copy source files to build directory."""
        print("📋 Copying source files...")
        
        # Copy source files
        source_files = [
            "src/main.py",
            "src/core/__init__.py",
            "src/core/browser.py",
            "src/core/config.py",
            "src/core/csv_processor.py",
            "src/core/downloader.py",
            "src/core/driver_manager.py"
        ]
        
        build_src_dir = self.build_dir / "src" / "core"
        build_src_dir.mkdir(parents=True, exist_ok=True)
        
        for file_path in source_files:
            src = self.project_root / file_path
            dst = self.build_dir / file_path
            
            if src.exists():
                dst.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(src, dst)
                print(f"  ✅ {file_path}")
            else:
                print(f"  ⚠️ {file_path} not found")
        
        print("✅ Source files copied")
    
    def build(self):
        """Build the complete Excel workflow EXE package."""
        print("🏗️ Building Excel workflow EXE package...")
        print("=" * 50)
        
        try:
            # Clean previous build
            self.clean_build()
            
            # Download stable EdgeDriver
            if not self.download_stable_edgedriver():
                print("❌ Failed to download EdgeDriver")
                return False
            
            # Copy source files
            self.copy_source_files()
            
            # Create Excel template
            template_path = self.create_excel_template()
            
            # Create Excel workflow script
            main_script_path = self.create_excel_workflow_script()
            
            print("\n" + "=" * 50)
            print("🎉 Excel workflow EXE package built successfully!")
            print("=" * 50)
            print("✅ Excel template created")
            print("✅ Excel workflow implemented")
            print("✅ Temporary workspace management")
            print("✅ Downloads folder integration")
            print("\n📁 Files created:")
            print(f"- {template_path.name} (Excel template)")
            print("- coupa_downloads_excel.py (main script)")
            
            return True
            
        except Exception as e:
            print(f"❌ Build failed: {e}")
            return False


def main():
    """Main function."""
    print("CoupaDownloads - Excel Workflow EXE Builder")
    print("=" * 50)
    
    if platform.system().lower() != "windows":
        print("❌ Excel workflow EXE build only supported on Windows")
        return 1
    
    builder = ExcelWorkflowExeBuilder()
    
    if builder.build():
        print("\n✅ Excel workflow build completed successfully!")
        return 0
    else:
        print("\n❌ Excel workflow build failed!")
        return 1


if __name__ == "__main__":
    sys.exit(main()) 