import os
import pandas as pd
from datetime import datetime
from typing import List, Tuple, Dict, Any
import threading
import time
import shutil
from pathlib import Path

from config import Config

class ExcelProcessor:
    """Enhanced Excel processor for parallel testing with full functionality from main project."""
    
    def __init__(self):
        self.excel_path = os.path.join(Config.INPUT_DIR, "input.xlsx")
        self.lock = threading.Lock()  # Thread safety for Excel operations
        self.backup_dir = os.path.join(Config.INPUT_DIR, "backups")
        os.makedirs(self.backup_dir, exist_ok=True)
    
    def get_excel_file_path(self) -> str:
        """Get the path to the Excel file."""
        return self.excel_path
    
    def read_po_numbers_from_excel(self, excel_path: str) -> Tuple[List[Dict], List[str], List[str], bool]:
        """Read PO numbers from Excel file."""
        try:
            # Read Excel file (try different sheet names)
            try:
                df = pd.read_excel(excel_path, sheet_name='PO_Data')
            except:
                try:
                    df = pd.read_excel(excel_path, sheet_name='Sheet1')
                except:
                    df = pd.read_excel(excel_path)  # Read first sheet
            
            # Define hierarchy columns
            hierarchy_cols = ['Priority', 'Supplier Segment', 'Spend Type', 'L1 UU Supplier Name']
            
            # Check if hierarchy columns exist
            has_hierarchy_data = all(col in df.columns for col in hierarchy_cols)
            
            # Convert to list of dictionaries with all columns
            po_entries = []
            for _, row in df.iterrows():
                entry = {
                    'po_number': str(row['PO_NUMBER']),
                    'status': row.get('STATUS', 'Pending'),
                    'supplier': row.get('SUPPLIER', ''),
                    'attachments_found': row.get('ATTACHMENTS_FOUND', 0),
                    'attachments_downloaded': row.get('ATTACHMENTS_DOWNLOADED', 0),
                    'attachment_names': row.get('AttachmentName', ''),
                    'last_processed': row.get('LAST_PROCESSED', ''),
                    'error_message': row.get('ERROR_MESSAGE', ''),
                    'download_folder': row.get('DOWNLOAD_FOLDER', ''),
                    'coupa_url': row.get('COUPA_URL', '')
                }
                
                # Add hierarchy data
                for col in hierarchy_cols:
                    if col in df.columns:
                        entry[col] = str(row[col]) if pd.notna(row[col]) else ''
                
                po_entries.append(entry)
            
            # Get original columns
            original_cols = list(df.columns)
            
            print(f"📊 Loaded {len(po_entries)} POs from Excel")
            return po_entries, original_cols, hierarchy_cols, has_hierarchy_data
            
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
            return [], [], [], False
    
    def _create_backup(self) -> str:
        """Create a backup of the Excel file before modification."""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
            backup_filename = f"input_backup_{timestamp}.xlsx"
            backup_path = os.path.join(self.backup_dir, backup_filename)
            shutil.copy2(self.excel_path, backup_path)
            return backup_path
        except Exception as e:
            print(f"⚠️ Warning: Could not create backup: {e}")
            return None
    
    def _wait_for_file_access(self, file_path: str, max_wait: int = 30) -> bool:
        """Wait for file to be accessible (not locked by another process)."""
        wait_time = 0
        while wait_time < max_wait:
            try:
                # Try to open file in append mode to test if it's accessible
                with open(file_path, 'a'):
                    pass
                return True
            except (PermissionError, IOError):
                time.sleep(1)
                wait_time += 1
        return False
    
    def _safe_excel_update(self, po_number: str, status: str, supplier: str = '', 
                          attachments_found: int = 0, attachments_downloaded: int = 0,
                          error_message: str = '', download_folder: str = '', 
                          coupa_url: str = '', attachment_names: str = '') -> bool:
        """
        Safely update Excel file with retry logic and backup creation.
        Returns True if successful, False otherwise.
        """
        max_retries = 3
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                # Wait for file access
                if not self._wait_for_file_access(self.excel_path):
                    print(f"⚠️ Excel file is locked, waiting... (attempt {retry_count + 1}/{max_retries})")
                    retry_count += 1
                    time.sleep(2)
                    continue
                
                # Create backup before modification
                backup_path = self._create_backup()
                
                # Use openpyxl to update specific cells without rewriting the entire file
                from openpyxl import load_workbook
                
                # Load the workbook with error handling
                try:
                    workbook = load_workbook(self.excel_path)
                except Exception as load_error:
                    print(f"❌ Error loading Excel file: {load_error}")
                    return False
                
                # IMPORTANT: Only work on the 'PO_Data' sheet, never touch backup sheets
                if 'PO_Data' in workbook.sheetnames:
                    worksheet = workbook['PO_Data']
                else:
                    print(f"❌ 'PO_Data' sheet not found in Excel file")
                    workbook.close()
                    return False
                
                # Find the PO number in the first column (PO_NUMBER)
                po_found = False
                for row_num in range(2, worksheet.max_row + 1):  # Start from row 2 (skip header)
                    cell_value = worksheet.cell(row=row_num, column=1).value  # PO_NUMBER column
                    if str(cell_value).strip() == str(po_number).strip():
                        # Update only the specific cells for this PO
                        worksheet.cell(row=row_num, column=2).value = status  # STATUS column
                        worksheet.cell(row=row_num, column=3).value = supplier  # SUPPLIER column
                        worksheet.cell(row=row_num, column=4).value = attachments_found  # ATTACHMENTS_FOUND column
                        worksheet.cell(row=row_num, column=5).value = attachments_downloaded  # ATTACHMENTS_DOWNLOADED column
                        worksheet.cell(row=row_num, column=6).value = attachment_names  # AttachmentName column
                        worksheet.cell(row=row_num, column=7).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # LAST_PROCESSED column
                        worksheet.cell(row=row_num, column=8).value = error_message  # ERROR_MESSAGE column
                        worksheet.cell(row=row_num, column=9).value = download_folder  # DOWNLOAD_FOLDER column
                        worksheet.cell(row=row_num, column=10).value = coupa_url  # COUPA_URL column
                        po_found = True
                        break
                
                if not po_found:
                    print(f"⚠️ PO {po_number} not found in Excel for status update")
                    workbook.close()
                    return False
                
                # Save the workbook with proper error handling
                try:
                    # Create a temporary file first to avoid corruption
                    temp_path = self.excel_path + '.tmp'
                    workbook.save(temp_path)
                    
                    # Verify the temporary file is valid
                    test_df = pd.read_excel(temp_path)
                    if len(test_df) == 0:
                        raise Exception("Saved file is empty")
                    
                    # Replace original file only if temporary file is valid
                    shutil.move(temp_path, self.excel_path)
                    print(f"✅ Excel file saved successfully: {len(test_df)} rows preserved")
                    
                except Exception as save_error:
                    print(f"❌ Error saving Excel file with openpyxl: {save_error}")
                    # Clean up temporary file if it exists
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
                    workbook.close()
                    
                    # Fallback to pandas approach
                    print("🔄 Attempting fallback with pandas...")
                    return self._update_excel_with_pandas(po_number, status, supplier, attachments_found, 
                                                        attachments_downloaded, error_message, download_folder, 
                                                        coupa_url, attachment_names)
                
                workbook.close()
                
                # Print status update
                status_emoji = {
                    'COMPLETED': '✅',
                    'FAILED': '❌', 
                    'PARTIAL': '⚠️',
                    'PENDING': '⏳',
                    'SKIPPED': '⏭️',
                    'NO_ATTACHMENTS': '📭',
                    'Success': '✅',
                    'Error': '❌'
                }.get(status, '📋')
                
                print(f"  {status_emoji} Updated {po_number}: {status}")
                
                # Clean up old backups (keep only last 10)
                self._cleanup_old_backups()
                
                return True
                
            except Exception as e:
                retry_count += 1
                print(f"❌ Error updating PO status (attempt {retry_count}/{max_retries}): {e}")
                
                if retry_count < max_retries:
                    time.sleep(2)  # Wait before retry
                else:
                    print(f"❌ Failed to update PO {po_number} after {max_retries} attempts")
                    return False
        
        return False
    
    def _cleanup_old_backups(self):
        """Clean up old backup files, keeping only the most recent 10."""
        try:
            backup_files = []
            for file in os.listdir(self.backup_dir):
                if file.startswith("input_backup_") and file.endswith(".xlsx"):
                    file_path = os.path.join(self.backup_dir, file)
                    backup_files.append((file_path, os.path.getmtime(file_path)))
            
            # Sort by modification time (newest first)
            backup_files.sort(key=lambda x: x[1], reverse=True)
            
            # Keep only the 10 most recent backups
            for file_path, _ in backup_files[10:]:
                try:
                    os.remove(file_path)
                except:
                    pass  # Ignore errors when cleaning up
                    
        except Exception as e:
            print(f"⚠️ Warning: Could not cleanup old backups: {e}")
    
    def _update_excel_with_pandas(self, po_number: str, status: str, supplier: str = '', 
                                 attachments_found: int = 0, attachments_downloaded: int = 0,
                                 error_message: str = '', download_folder: str = '', 
                                 coupa_url: str = '', attachment_names: str = '') -> bool:
        """
        Fallback method to update Excel using pandas when openpyxl fails.
        """
        try:
            # Read the Excel file
            df = pd.read_excel(self.excel_path)
            
            # Find and update the PO row
            po_mask = df['PO_NUMBER'].astype(str).str.strip() == str(po_number).strip()
            if not po_mask.any():
                print(f"⚠️ PO {po_number} not found in Excel for pandas update")
                return False
            
            # Update the specific row
            df.loc[po_mask, 'STATUS'] = status
            df.loc[po_mask, 'SUPPLIER'] = supplier
            df.loc[po_mask, 'ATTACHMENTS_FOUND'] = attachments_found
            df.loc[po_mask, 'ATTACHMENTS_DOWNLOADED'] = attachments_downloaded
            df.loc[po_mask, 'AttachmentName'] = attachment_names
            df.loc[po_mask, 'LAST_PROCESSED'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            df.loc[po_mask, 'ERROR_MESSAGE'] = error_message
            df.loc[po_mask, 'DOWNLOAD_FOLDER'] = download_folder
            df.loc[po_mask, 'COUPA_URL'] = coupa_url
            
            # Save with pandas (safer for complex Excel files)
            temp_path = self.excel_path + '.pandas.tmp'
            df.to_excel(temp_path, index=False)
            
            # Verify the file is valid
            test_df = pd.read_excel(temp_path)
            if len(test_df) == 0:
                raise Exception("Pandas saved file is empty")
            
            # Replace original file
            shutil.move(temp_path, self.excel_path)
            print(f"✅ Excel file saved successfully with pandas: {len(test_df)} rows preserved")
            
            return True
            
        except Exception as e:
            print(f"❌ Error updating Excel with pandas: {e}")
            # Clean up temporary file
            temp_path = self.excel_path + '.pandas.tmp'
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return False
    
    def update_po_status(self, po_number: str, status: str, supplier: str = '', 
                        attachments_found: int = 0, attachments_downloaded: int = 0,
                        error_message: str = '', download_folder: str = '', 
                        coupa_url: str = '', attachment_names: str = '') -> None:
        """
        Update the status of a specific PO in the Excel file with thread-safe operations.
        
        Args:
            po_number: PO number to update
            status: New status (PENDING, COMPLETED, FAILED, etc.)
            supplier: Supplier name
            attachments_found: Number of attachments found
            attachments_downloaded: Number of attachments downloaded
            error_message: Error message if applicable
            download_folder: Folder where files were saved
            coupa_url: Full URL to the PO page
            attachment_names: Formatted attachment names
        """
        # Use thread lock to ensure only one thread updates Excel at a time
        with self.lock:
            success = self._safe_excel_update(
                po_number, status, supplier, attachments_found, 
                attachments_downloaded, error_message, download_folder, 
                coupa_url, attachment_names
            )
            
            if not success:
                print(f"❌ Failed to update Excel for PO {po_number}")
