#!/usr/bin/env python3
"""
Excel Export Example
Demonstrates how to add Excel export functionality to the existing CSV processor.
"""

import os
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any
from core.csv_processor import CSVProcessor


class ExcelExporter:
    """Excel export functionality that works alongside CSVProcessor."""
    
    @staticmethod
    def export_csv_to_excel(csv_file_path: str, excel_file_path: str) -> None:
        """
        Export CSV data to formatted Excel file.
        
        Args:
            csv_file_path: Path to input CSV file
            excel_file_path: Path to output Excel file
        """
        try:
            # Read CSV data using existing CSVProcessor
            po_entries = CSVProcessor.read_po_numbers_from_csv(csv_file_path)
            
            if not po_entries:
                print("❌ No data found in CSV file")
                return
            
            # Create Excel file with multiple sheets
            with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
                # Main PO data sheet
                po_df = pd.DataFrame([
                    {
                        'PO_NUMBER': entry['po_number'],
                        'STATUS': entry['status'],
                        'SUPPLIER': entry['supplier'],
                        'ATTACHMENTS_FOUND': entry['attachments_found'],
                        'ATTACHMENTS_DOWNLOADED': entry['attachments_downloaded'],
                        'LAST_PROCESSED': entry['last_processed'],
                        'ERROR_MESSAGE': entry['error_message'],
                        'DOWNLOAD_FOLDER': entry['download_folder'],
                        'COUPA_URL': entry['coupa_url']
                    }
                    for entry in po_entries
                ])
                
                po_df.to_excel(writer, sheet_name='PO_Data', index=False)
                
                # Summary sheet
                summary_data = ExcelExporter._generate_summary_data(po_entries)
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Supplier breakdown sheet
                supplier_data = ExcelExporter._generate_supplier_data(po_entries)
                if supplier_data:
                    supplier_df = pd.DataFrame(supplier_data)
                    supplier_df.to_excel(writer, sheet_name='Suppliers', index=False)
            
            # Apply formatting
            ExcelExporter._apply_excel_formatting(excel_file_path)
            
            print(f"✅ Excel file created: {excel_file_path}")
            
        except Exception as e:
            print(f"❌ Error creating Excel file: {e}")
    
    @staticmethod
    def _generate_summary_data(po_entries: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Generate summary statistics for Excel report."""
        total_pos = len(po_entries)
        status_counts = {}
        total_attachments_found = 0
        total_attachments_downloaded = 0
        suppliers = set()
        
        for entry in po_entries:
            status = entry['status']
            status_counts[status] = status_counts.get(status, 0) + 1
            total_attachments_found += entry['attachments_found']
            total_attachments_downloaded += entry['attachments_downloaded']
            
            if entry['supplier']:
                suppliers.add(entry['supplier'])
        
        success_rate = round((status_counts.get('COMPLETED', 0) / total_pos) * 100, 1) if total_pos > 0 else 0
        
        return [
            {'METRIC': 'Total POs', 'VALUE': total_pos},
            {'METRIC': 'Completed', 'VALUE': status_counts.get('COMPLETED', 0)},
            {'METRIC': 'Failed', 'VALUE': status_counts.get('FAILED', 0)},
            {'METRIC': 'Pending', 'VALUE': status_counts.get('PENDING', 0)},
            {'METRIC': 'Success Rate', 'VALUE': f'{success_rate}%'},
            {'METRIC': 'Total Attachments Found', 'VALUE': total_attachments_found},
            {'METRIC': 'Total Attachments Downloaded', 'VALUE': total_attachments_downloaded},
            {'METRIC': 'Unique Suppliers', 'VALUE': len(suppliers)},
            {'METRIC': 'Report Generated', 'VALUE': datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        ]
    
    @staticmethod
    def _generate_supplier_data(po_entries: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Generate supplier breakdown data."""
        supplier_stats = {}
        
        for entry in po_entries:
            supplier = entry['supplier'] or 'Unknown'
            if supplier not in supplier_stats:
                supplier_stats[supplier] = {
                    'total_pos': 0,
                    'completed': 0,
                    'failed': 0,
                    'attachments_found': 0,
                    'attachments_downloaded': 0
                }
            
            supplier_stats[supplier]['total_pos'] += 1
            supplier_stats[supplier]['attachments_found'] += entry['attachments_found']
            supplier_stats[supplier]['attachments_downloaded'] += entry['attachments_downloaded']
            
            if entry['status'] == 'COMPLETED':
                supplier_stats[supplier]['completed'] += 1
            elif entry['status'] == 'FAILED':
                supplier_stats[supplier]['failed'] += 1
        
        return [
            {
                'SUPPLIER': supplier,
                'TOTAL_POS': stats['total_pos'],
                'COMPLETED': stats['completed'],
                'FAILED': stats['failed'],
                'SUCCESS_RATE': f"{round((stats['completed'] / stats['total_pos']) * 100, 1)}%" if stats['total_pos'] > 0 else "0%",
                'ATTACHMENTS_FOUND': stats['attachments_found'],
                'ATTACHMENTS_DOWNLOADED': stats['attachments_downloaded']
            }
            for supplier, stats in supplier_stats.items()
        ]
    
    @staticmethod
    def _apply_excel_formatting(excel_file_path: str) -> None:
        """Apply professional formatting to Excel file."""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.formatting.rule import CellIsRule
            
            wb = load_workbook(excel_file_path)
            
            # Format PO_Data sheet
            if 'PO_Data' in wb.sheetnames:
                ws = wb['PO_Data']
                
                # Style header row
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Conditional formatting for status column
                green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
                
                # Apply conditional formatting to status column (column B)
                status_col = 'B'
                for row in range(2, ws.max_row + 1):
                    cell = ws[f'{status_col}{row}']
                    if cell.value == 'COMPLETED':
                        cell.fill = green_fill
                    elif cell.value == 'FAILED':
                        cell.fill = red_fill
                    elif cell.value == 'PENDING':
                        cell.fill = yellow_fill
            
            # Format Summary sheet
            if 'Summary' in wb.sheetnames:
                ws = wb['Summary']
                
                # Style header row
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
            
            wb.save(excel_file_path)
            print("✅ Excel formatting applied successfully")
            
        except ImportError:
            print("⚠️ openpyxl not available - skipping formatting")
        except Exception as e:
            print(f"⚠️ Error applying formatting: {e}")


def main():
    """Example usage of Excel export functionality."""
    print("📊 Excel Export Example")
    print("=" * 40)
    
    # Get CSV file path
    csv_file_path = CSVProcessor.get_csv_file_path()
    
    if not os.path.exists(csv_file_path):
        print(f"❌ CSV file not found: {csv_file_path}")
        print("💡 Please create an input.csv file first")
        return
    
    # Generate Excel file path
    script_dir = os.path.dirname(os.path.abspath(__file__))
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_file_path = os.path.join(script_dir, f"coupa_report_{timestamp}.xlsx")
    
    # Export to Excel
    print(f"📁 Reading CSV: {csv_file_path}")
    print(f"📊 Creating Excel: {excel_file_path}")
    
    ExcelExporter.export_csv_to_excel(csv_file_path, excel_file_path)
    
    # Show summary
    if os.path.exists(excel_file_path):
        file_size = os.path.getsize(excel_file_path) / 1024  # KB
        print(f"\n📈 Report Summary:")
        print(f"   File: {os.path.basename(excel_file_path)}")
        print(f"   Size: {file_size:.1f} KB")
        print(f"   Sheets: PO_Data, Summary, Suppliers")
        print(f"   Features: Conditional formatting, professional styling")


if __name__ == "__main__":
    main() 