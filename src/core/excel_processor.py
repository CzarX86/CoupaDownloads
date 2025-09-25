"""
Excel/CSV processor module for Coupa Downloads automation.
Handles spreadsheet (CSV/XLSX) reading and PO number processing with enhanced hierarchy support.
"""

import os
import csv
import pandas as pd
from datetime import datetime
from typing import List, Tuple, Dict, Any, Optional
from .config import Config
from .folder_hierarchy import FolderHierarchyManager


class ExcelProcessor:
    """Handles reading PO numbers from Excel and updating progress with hierarchy support."""

    def __init__(self):
        self.folder_hierarchy = FolderHierarchyManager()

    @staticmethod
    def get_excel_file_path() -> str:
        """Get the full path to the input spreadsheet file.

        Supports both Excel (.xlsx) and CSV (.csv):
        - Honors env vars: `EXCEL_FILE_PATH`, `INPUT_XLSX`, or `INPUT_CSV` when set
        - Otherwise, prefers `<project_root>/data/input/input.csv` if it exists
        - Falls back to `<project_root>/data/input/input.xlsx`
        """
        # Allow override via environment for safer test/runs without touching the real file
        override = (
            os.environ.get("EXCEL_FILE_PATH")
            or os.environ.get("INPUT_XLSX")
            or os.environ.get("INPUT_CSV")
        )
        if override:
            return os.path.expanduser(override)

        # Navigate from src/core/ to project root, then to data/input/
        current_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(os.path.dirname(current_dir))
        csv_path = os.path.join(project_root, "data", "input", "input.csv")
        xlsx_path = os.path.join(project_root, "data", "input", "input.xlsx")
        return csv_path if os.path.exists(csv_path) else xlsx_path

    def read_po_numbers_from_excel(self, excel_file_path: str) -> Tuple[List[Dict[str, Any]], List[str], List[str], bool]:
        """
        Read PO numbers from spreadsheet (Excel or CSV) and analyze hierarchy structure.
        
        Returns:
            Tuple of (po_entries, original_columns, hierarchy_columns, has_hierarchy_data)
        """
        try:
            # Determine file type and read accordingly
            _, ext = os.path.splitext(excel_file_path.lower())
            if ext == ".csv":
                # Robust CSV read: handle BOM and auto-detect delimiter (comma/semicolon)
                df, detected_sep, used_encoding = self._read_csv_auto(excel_file_path)
                print(f"🧭 Detected CSV (sep='{detected_sep}', encoding='{used_encoding}')")
            else:
                # Read Excel file using pandas - specifically target 'PO_Data' sheet
                df = pd.read_excel(excel_file_path, sheet_name='PO_Data', engine='openpyxl')
                
                # Safety check: warn about multiple sheets
                xl_file = pd.ExcelFile(excel_file_path)
                if len(xl_file.sheet_names) > 1:
                    print(f"⚠️ WARNING: Excel file contains {len(xl_file.sheet_names)} sheets: {xl_file.sheet_names}")
                    print(f"   Only working on 'PO_Data' sheet. Other sheets will NOT be modified.")
            
            # Analyze Excel structure for hierarchy
            original_cols, hierarchy_cols, has_hierarchy_data = self.folder_hierarchy.analyze_excel_structure(df)
            
            print(f"📊 Input structure analysis:")
            print(f"   Original columns: {original_cols}")
            print(f"   Hierarchy columns: {hierarchy_cols}")
            print(f"   Has hierarchy data: {has_hierarchy_data}")
            
            # Create a copy for processing to avoid modifying the original DataFrame
            df_processed = df.copy()
            
            # Fill NaN values with appropriate defaults only for processing columns
            df_processed = df_processed.fillna({
                'STATUS': 'PENDING',
                'SUPPLIER': '',
                'ATTACHMENTS_FOUND': 0,
                'ATTACHMENTS_DOWNLOADED': 0,
                'LAST_PROCESSED': '',
                'ERROR_MESSAGE': '',
                'DOWNLOAD_FOLDER': '',
                'COUPA_URL': '',
                'AttachmentName': ''
            })
            
            po_entries = []
            
            # Process each row
            for _, row in df_processed.iterrows():
                po_number = str(row.get('PO_NUMBER', '')).strip()
                if not po_number or po_number == 'nan':
                    continue
                
                # Create PO entry with all available data
                po_entry = {
                    'po_number': po_number,
                    'status': str(row.get('STATUS', 'PENDING')).strip(),
                    'supplier': str(row.get('SUPPLIER', '')).strip(),
                    'attachments_found': int(row.get('ATTACHMENTS_FOUND', 0)),
                    'attachments_downloaded': int(row.get('ATTACHMENTS_DOWNLOADED', 0)),
                    'last_processed': str(row.get('LAST_PROCESSED', '')).strip(),
                    'error_message': str(row.get('ERROR_MESSAGE', '')).strip(),
                    'download_folder': str(row.get('DOWNLOAD_FOLDER', '')).strip(),
                    'coupa_url': str(row.get('COUPA_URL', '')).strip(),
                    'attachment_names': str(row.get('AttachmentName', '')).strip()
                }
                
                # Add hierarchy column data if available
                for col in hierarchy_cols:
                    if col in row:
                        po_entry[col] = str(row.get(col, '')).strip()
                
                po_entries.append(po_entry)
                            
        except FileNotFoundError:
            raise Exception(f"Input file not found: {excel_file_path}")
        except Exception as e:
            raise Exception(f"Error reading input file: {e}")

        print(f"📊 Read {len(po_entries)} PO entries from input file")
        
        # Safety check: warn if no PO numbers found
        if len(po_entries) == 0:
            print("⚠️ WARNING: No PO numbers found in input file!")
            print("   Please ensure the PO_NUMBER column contains valid PO numbers.")
            print("   The application will not process any files without PO numbers.")
        
        return po_entries, original_cols, hierarchy_cols, has_hierarchy_data

    @staticmethod
    def _read_csv_auto(path: str) -> Tuple[pd.DataFrame, str, str]:
        """Read CSV with auto delimiter and BOM handling.

        Returns: (df, delimiter_used, encoding_used)
        """
        # Try UTF-8 with BOM handling first
        encodings = ["utf-8-sig", "utf-8", "latin1"]
        last_err: Exception | None = None
        for enc in encodings:
            try:
                # sep=None requires engine='python' for detection
                df = pd.read_csv(path, sep=None, engine='python', encoding=enc)
                # Heuristic: detect actual delimiter from header if possible
                with open(path, 'r', encoding=enc, errors='ignore') as f:
                    header_line = f.readline()
                detected_sep = ';' if header_line.count(';') > header_line.count(',') else ','
                return df, detected_sep, enc
            except Exception as e:
                last_err = e
                continue
        # If all failed, raise the last error
        raise Exception(f"CSV read failed: {last_err}")

    @staticmethod
    def _find_column(df: pd.DataFrame, target: str) -> Optional[str]:
        """Find a column by name, ignoring case and surrounding spaces."""
        if target in df.columns:
            return target
        low = {str(c).strip().lower(): str(c) for c in df.columns}
        return low.get(target.lower())

    @staticmethod
    def process_po_numbers(po_entries: List[Dict[str, Any]]) -> List[Tuple[str, str]]:
        """
        Process and validate PO numbers, filtering out already completed ones.
        
        Returns:
            List of tuples: (display_po, clean_po) for processing
        """
        valid_entries = []
        skipped_count = 0
        
        for entry in po_entries:
            po_number = entry['po_number']
            status = entry['status']
            
            if not po_number or po_number == 'nan':
                continue
                
            # Skip already completed POs unless forced reprocessing
            if status == 'COMPLETED':
                print(f"  ⏭️ Skipping {po_number} (already completed)")
                skipped_count += 1
                continue
            
            # Clean PO number (remove PO or PM prefix if present)
            display_po = po_number
            clean_po = po_number
            if po_number.startswith("PO"):
                clean_po = po_number.replace("PO", "")
            elif po_number.startswith("PM"):
                clean_po = po_number.replace("PM", "")
            
            # Validate PO number format - accept both PO and PM prefixes
            if clean_po.isdigit() and len(clean_po) >= 6:
                valid_entries.append((display_po, clean_po))
                print(f"  ✅ {display_po} → Will process")
            else:
                print(f"  ❌ Invalid PO format: {po_number}")
                # Update status to FAILED
                # Generate URL even for invalid PO for reference
                clean_po_attempt = po_number
                if po_number.startswith("PO"):
                    clean_po_attempt = po_number.replace("PO", "")
                elif po_number.startswith("PM"):
                    clean_po_attempt = po_number.replace("PM", "")
                invalid_url = Config.BASE_URL.format(clean_po_attempt)
                ExcelProcessor.update_po_status(po_number, 'FAILED', error_message='Invalid PO format', coupa_url=invalid_url)
        
        if skipped_count > 0:
            print(f"📊 Skipped {skipped_count} already completed POs")
            
        print(f"📊 {len(valid_entries)} POs ready for processing")
        return valid_entries

    @staticmethod
    def update_po_status(po_number: str, status: str, supplier: str = '', 
                        attachments_found: int = 0, attachments_downloaded: int = 0,
                        error_message: str = '', download_folder: str = '', 
                        coupa_url: str = '', attachment_names: str = '') -> None:
        """
        Update the status of a specific PO in the Excel file.
        
        Args:
            po_number: PO number to update
            status: New status (PENDING, COMPLETED, FAILED, etc.)
            supplier: Supplier name
            attachments_found: Number of attachments found
            attachments_downloaded: Number of attachments downloaded
            error_message: Error message if applicable
            download_folder: Folder where files were saved
            coupa_url: Full URL to the PO page
            attachment_names: Formatted attachment names
        """
        excel_file_path = ExcelProcessor.get_excel_file_path()
        _, ext = os.path.splitext(excel_file_path.lower())
        
        try:
            if ext == ".csv":
                # Update CSV by reading (auto), modifying the row, and writing back preserving delimiter.
                # Always write in UTF-8 with BOM (utf-8-sig) to avoid encoding issues with emojis/special chars in Excel.
                df, detected_sep, _enc = ExcelProcessor._read_csv_auto(excel_file_path)

                # Resolve PO column robustly
                po_col = ExcelProcessor._find_column(df, 'PO_NUMBER')
                if not po_col:
                    print("❌ 'PO_NUMBER' column not found in CSV for status update")
                    return

                # Ensure expected columns exist
                string_cols = ['STATUS', 'SUPPLIER', 'AttachmentName', 'LAST_PROCESSED', 'ERROR_MESSAGE', 'DOWNLOAD_FOLDER', 'COUPA_URL']
                int_cols = ['ATTACHMENTS_FOUND', 'ATTACHMENTS_DOWNLOADED']

                # Add missing cols and coerce dtypes to avoid pandas warnings
                for col in string_cols:
                    if col not in df.columns:
                        df[col] = ""
                    else:
                        df[col] = df[col].astype('string').fillna("")
                for col in int_cols:
                    if col not in df.columns:
                        df[col] = 0
                    else:
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

                target = str(po_number).strip()
                # Also accept numeric-only variant (strip PO/PM prefix)
                target_clean = target
                if target.upper().startswith("PO"):
                    target_clean = target[2:]
                elif target.upper().startswith("PM"):
                    target_clean = target[2:]

                series = df[po_col].astype(str).str.strip()
                mask = series.isin([target, target_clean])
                if not mask.any():
                    print(f"⚠️ PO {po_number} not found in CSV for status update")
                    return

                df.loc[mask, 'STATUS'] = status
                df.loc[mask, 'SUPPLIER'] = supplier
                df.loc[mask, 'ATTACHMENTS_FOUND'] = attachments_found
                df.loc[mask, 'ATTACHMENTS_DOWNLOADED'] = attachments_downloaded
                df.loc[mask, 'AttachmentName'] = attachment_names
                df.loc[mask, 'LAST_PROCESSED'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                df.loc[mask, 'ERROR_MESSAGE'] = error_message
                df.loc[mask, 'DOWNLOAD_FOLDER'] = download_folder
                df.loc[mask, 'COUPA_URL'] = coupa_url

                # Ensure a stable, human-friendly column order matching Excel layout
                desired_order = [
                    'PO_NUMBER', 'STATUS', 'SUPPLIER', 'ATTACHMENTS_FOUND',
                    'ATTACHMENTS_DOWNLOADED', 'AttachmentName', 'LAST_PROCESSED',
                    'ERROR_MESSAGE', 'DOWNLOAD_FOLDER', 'COUPA_URL'
                ]
                existing = [c for c in desired_order if c in df.columns]
                remaining = [c for c in df.columns if c not in desired_order]
                df = df[existing + remaining]

                # Write back with stable settings for Excel compatibility
                df.to_csv(
                    excel_file_path,
                    index=False,
                    sep=detected_sep,
                    encoding="utf-8-sig",
                    lineterminator="\n",
                    quoting=csv.QUOTE_MINIMAL
                )
            else:
                # Use openpyxl to update specific cells without rewriting the entire file
                from openpyxl import load_workbook
                
                # Load the workbook
                workbook = load_workbook(excel_file_path)
                
                # IMPORTANT: Only work on the 'PO_Data' sheet, never touch backup sheets
                if 'PO_Data' in workbook.sheetnames:
                    worksheet = workbook['PO_Data']
                else:
                    print(f"❌ 'PO_Data' sheet not found in Excel file")
                    return
                
                # Find the PO number in the first column (PO_NUMBER)
                po_found = False
                for row_num in range(2, worksheet.max_row + 1):  # Start from row 2 (skip header)
                    cell_value = worksheet.cell(row=row_num, column=1).value  # PO_NUMBER column
                    if str(cell_value).strip() == str(po_number).strip():
                        # Update only the specific cells for this PO
                        worksheet.cell(row=row_num, column=2).value = status  # STATUS column
                        worksheet.cell(row=row_num, column=3).value = supplier  # SUPPLIER column
                        worksheet.cell(row=row_num, column=4).value = attachments_found  # ATTACHMENTS_FOUND column
                        worksheet.cell(row=row_num, column=5).value = attachments_downloaded  # ATTACHMENTS_DOWNLOADED column
                        worksheet.cell(row=row_num, column=6).value = attachment_names  # AttachmentName column
                        worksheet.cell(row=row_num, column=7).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # LAST_PROCESSED column
                        worksheet.cell(row=row_num, column=8).value = error_message  # ERROR_MESSAGE column
                        worksheet.cell(row=row_num, column=9).value = download_folder  # DOWNLOAD_FOLDER column
                        worksheet.cell(row=row_num, column=10).value = coupa_url  # COUPA_URL column
                        po_found = True
                        break
                
                if not po_found:
                    print(f"⚠️ PO {po_number} not found in Excel for status update")
                    return
                
                # Save the workbook
                workbook.save(excel_file_path)
                workbook.close()
            
            # Print status update
            status_emoji = {
                'COMPLETED': '✅',
                'FAILED': '❌', 
                'PARTIAL': '⚠️',
                'PENDING': '⏳',
                'SKIPPED': '⏭️',
                'NO_ATTACHMENTS': '📭',
                'PO_NOT_FOUND': '🚫'
            }.get(status, '📋')
            
            print(f"  {status_emoji} Updated {po_number}: {status}")
            
        except Exception as e:
            print(f"❌ Error updating PO status: {e}")

    @staticmethod
    def write_enhanced_excel(excel_file_path: str, po_entries: List[Dict[str, Any]], 
                           original_cols: List[str], hierarchy_cols: List[str]) -> None:
        """
        Write the enhanced Excel format with all tracking columns and hierarchy support.
        WARNING: This method overwrites the entire file. Use update_po_status for safer updates.
        
        Args:
            excel_file_path: Path to Excel file
            po_entries: List of PO entry dictionaries
            original_cols: List of original column names
            hierarchy_cols: List of hierarchy column names
        """
        try:
            # Create DataFrame with all columns
            df_data = []
            for entry in po_entries:
                row_data = {}
                
                # Add original columns
                for col in original_cols:
                    if col in entry:
                        row_data[col] = entry[col]
                    else:
                        row_data[col] = ''
                
                # Add separator column if not present
                if '<|>' not in row_data:
                    row_data['<|>'] = ''
                
                # Add hierarchy columns
                for col in hierarchy_cols:
                    if col in entry:
                        row_data[col] = entry[col]
                    else:
                        row_data[col] = ''
                
                df_data.append(row_data)
            
            df = pd.DataFrame(df_data)
            
            # Write to Excel with formatting
            with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='PO_Data', index=False)
                
                # Apply formatting
                ExcelProcessor._apply_excel_formatting(writer, df)
                    
        except Exception as e:
            print(f"❌ Error writing enhanced Excel: {e}")

    @staticmethod
    def _apply_excel_formatting(writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
        """Apply professional formatting to Excel file."""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['PO_Data']
            
            # Style header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Conditional formatting for status column
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            
            # Apply conditional formatting to status column (column B)
            status_col = 'B'
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f'{status_col}{row}']
                if cell.value == 'COMPLETED':
                    cell.fill = green_fill
                elif cell.value == 'FAILED':
                    cell.fill = red_fill
                elif cell.value == 'PENDING':
                    cell.fill = yellow_fill
                    
        except ImportError:
            print("⚠️ openpyxl not available - skipping formatting")
        except Exception as e:
            print(f"⚠️ Error applying formatting: {e}")

    @staticmethod
    def generate_summary_report(excel_file_path: Optional[str] = None) -> Dict[str, Any]:
        """
        Generate a summary report of processing results.
        
        Returns:
            Dictionary with summary statistics
        """
        if excel_file_path is None:
            excel_file_path = ExcelProcessor.get_excel_file_path()
            
        processor = ExcelProcessor()
        po_entries, original_cols, hierarchy_cols, has_hierarchy_data = processor.read_po_numbers_from_excel(excel_file_path)
        
        # Count by status
        status_counts = {}
        total_attachments_found = 0
        total_attachments_downloaded = 0
        suppliers = set()
        
        for entry in po_entries:
            status = entry['status']
            status_counts[status] = status_counts.get(status, 0) + 1
            total_attachments_found += entry['attachments_found']
            total_attachments_downloaded += entry['attachments_downloaded']
            
            if entry['supplier']:
                suppliers.add(entry['supplier'])
        
        return {
            'total_pos': len(po_entries),
            'status_counts': status_counts,
            'total_attachments_found': total_attachments_found,
            'total_attachments_downloaded': total_attachments_downloaded,
            'unique_suppliers': len(suppliers),
            'supplier_list': sorted(list(suppliers)),
            'success_rate': round((status_counts.get('COMPLETED', 0) / len(po_entries)) * 100, 1) if po_entries else 0,
            'has_hierarchy_data': has_hierarchy_data,
            'hierarchy_columns': hierarchy_cols
        }

    @staticmethod
    def print_summary_report() -> None:
        """Print a formatted summary report to console."""
        report = ExcelProcessor.generate_summary_report()
        
        print("\n" + "="*60)
        print("📊 PROCESSING SUMMARY REPORT (Excel)")
        print("="*60)
        
        print(f"📋 Total POs: {report['total_pos']}")
        print(f"🎯 Success Rate: {report['success_rate']}%")
        
        print(f"\n📈 Status Breakdown:")
        for status, count in report['status_counts'].items():
            emoji = {
                'COMPLETED': '✅', 'FAILED': '❌', 'PARTIAL': '⚠️',
                'PENDING': '⏳', 'SKIPPED': '⏭️', 'NO_ATTACHMENTS': '📭',
                'PO_NOT_FOUND': '🚫'
            }.get(status, '📋')
            print(f"  {emoji} {status}: {count}")
        
        print(f"\n📎 Attachments:")
        print(f"  🔍 Found: {report['total_attachments_found']}")
        print(f"  💾 Downloaded: {report['total_attachments_downloaded']}")
        
        if report['unique_suppliers'] > 0:
            print(f"\n🏢 Suppliers: {report['unique_suppliers']} unique")
            for supplier in report['supplier_list'][:10]:  # Show first 10
                print(f"  📁 {supplier}")
            if len(report['supplier_list']) > 10:
                print(f"  ... and {len(report['supplier_list']) - 10} more")
        
        # Hierarchy information
        if report['has_hierarchy_data']:
            print(f"\n📁 Hierarchy Structure:")
            print(f"  ✅ Using hierarchy columns: {report['hierarchy_columns']}")
        else:
            print(f"\n📁 Folder Structure:")
            print(f"  🔄 Using fallback structure (Supplier/PO)")
        
        print("="*60)

    @staticmethod
    def backup_excel() -> str:
        """
        Create a backup of the current Excel file.
        
        Returns:
            Path to backup file
        """
        excel_file_path = ExcelProcessor.get_excel_file_path()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Create backups directory if it doesn't exist
        backups_dir = os.path.join(os.path.dirname(excel_file_path), 'backups')
        os.makedirs(backups_dir, exist_ok=True)
        
        # Create backup filename and path
        backup_filename = f'input_backup_{timestamp}.xlsx'
        backup_path = os.path.join(backups_dir, backup_filename)
        
        try:
            import shutil
            shutil.copy2(excel_file_path, backup_path)
            print(f"💾 Excel backup created: {backup_filename}")
            return backup_path
        except Exception as e:
            print(f"⚠️ Could not create Excel backup: {e}")
            return ""

    @staticmethod
    def convert_csv_to_excel(csv_file_path: str, excel_file_path: str) -> None:
        """
        Convert CSV file to Excel format.
        
        Args:
            csv_file_path: Path to input CSV file
            excel_file_path: Path to output Excel file
        """
        try:
            # Read CSV using pandas
            df = pd.read_csv(csv_file_path)
            
            # Write to Excel with formatting
            with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='PO_Data', index=False)
                
                # Apply formatting
                ExcelProcessor._apply_excel_formatting(writer, df)
            
            print(f"✅ Converted CSV to Excel: {excel_file_path}")
            
        except Exception as e:
            print(f"❌ Error converting CSV to Excel: {e}") 
