"""
Excel Processor Testing
Tests Excel processor implementation that mirrors CSV processor interface.
"""

import os
import tempfile
import pytest
import pandas as pd
from datetime import datetime
from unittest.mock import Mock, patch
from typing import List, Dict, Any, Optional, Tuple


class ExcelProcessor:
    """
    Mock Excel processor that mirrors CSVProcessor interface.
    This allows us to test Excel integration without breaking existing code.
    """
    
    def __init__(self, excel_file_path: str):
        self.excel_file_path = excel_file_path
        self.sheet_name = 'PO_Data'
    
    @staticmethod
    def get_excel_file_path() -> str:
        """Get the full path to the input Excel file."""
        script_dir = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(script_dir, "input.xlsx")
    
    def read_po_numbers_from_excel(self) -> List[Dict[str, Any]]:
        """
        Read PO numbers from Excel file and return enhanced data structure.
        Mirrors CSVProcessor.read_po_numbers_from_csv interface.
        """
        po_entries = []
        
        try:
            # Check if file exists
            if not os.path.exists(self.excel_file_path):
                print(f"❌ Excel file not found: {self.excel_file_path}")
                return []
            
            # Read Excel file
            df = pd.read_excel(self.excel_file_path, sheet_name=self.sheet_name)
            
            # Convert DataFrame to list of dictionaries
            for _, row in df.iterrows():
                po_entries.append({
                    'po_number': str(row.get('PO_NUMBER', '')).strip(),
                    'status': str(row.get('STATUS', 'PENDING')).strip(),
                    'supplier': str(row.get('SUPPLIER', '')).strip(),
                    'attachments_found': int(row.get('ATTACHMENTS_FOUND', 0)),
                    'attachments_downloaded': int(row.get('ATTACHMENTS_DOWNLOADED', 0)),
                    'last_processed': str(row.get('LAST_PROCESSED', '')).strip(),
                    'error_message': str(row.get('ERROR_MESSAGE', '')).strip(),
                    'download_folder': str(row.get('DOWNLOAD_FOLDER', '')).strip(),
                    'coupa_url': str(row.get('COUPA_URL', '')).strip()
                })
                
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
            return []
        
        print(f"📊 Read {len(po_entries)} PO entries from Excel")
        return po_entries
    
    def process_po_numbers(self, po_entries: List[Dict[str, Any]]) -> List[Tuple[str, str]]:
        """
        Process and validate PO numbers, filtering out already completed ones.
        Mirrors CSVProcessor.process_po_numbers interface.
        """
        valid_entries = []
        skipped_count = 0
        
        for entry in po_entries:
            po_number = entry['po_number']
            status = entry['status']
            
            if not po_number:
                continue
                
            # Skip already completed POs unless forced reprocessing
            if status == 'COMPLETED':
                print(f"  ⏭️ Skipping {po_number} (already completed)")
                skipped_count += 1
                continue
            
            # Clean PO number (remove PO prefix if present)
            display_po = po_number
            clean_po = po_number.replace("PO", "") if po_number.startswith("PO") else po_number
            
            # Validate PO number format
            if clean_po.isdigit() and len(clean_po) >= 6:
                valid_entries.append((display_po, clean_po))
                print(f"  ✅ {display_po} → Will process")
            else:
                print(f"  ❌ Invalid PO format: {po_number}")
        
        if skipped_count > 0:
            print(f"📊 Skipped {skipped_count} already completed POs")
            
        print(f"📊 {len(valid_entries)} POs ready for processing")
        return valid_entries
    
    def update_po_status(self, po_number: str, status: str, supplier: str = '', 
                        attachments_found: int = 0, attachments_downloaded: int = 0,
                        error_message: str = '', download_folder: str = '', 
                        coupa_url: str = '') -> None:
        """
        Update the status of a specific PO in the Excel file.
        Mirrors CSVProcessor.update_po_status interface.
        """
        try:
            # Read all entries
            po_entries = self.read_po_numbers_from_excel()
            
            # Find and update the specific PO
            updated = False
            for entry in po_entries:
                if entry['po_number'] == po_number:
                    entry['status'] = status
                    entry['supplier'] = supplier
                    entry['attachments_found'] = attachments_found
                    entry['attachments_downloaded'] = attachments_downloaded
                    entry['last_processed'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    entry['error_message'] = error_message
                    entry['download_folder'] = download_folder
                    entry['coupa_url'] = coupa_url
                    updated = True
                    break
            
            if not updated:
                print(f"⚠️ PO {po_number} not found in Excel for status update")
                return
            
            # Write back to Excel
            self.write_enhanced_excel(po_entries)
            
            # Print status update
            status_emoji = {
                'COMPLETED': '✅',
                'FAILED': '❌', 
                'PARTIAL': '⚠️',
                'PENDING': '⏳',
                'SKIPPED': '⏭️',
                'NO_ATTACHMENTS': '📭'
            }.get(status, '📋')
            
            print(f"  {status_emoji} Updated {po_number}: {status}")
            
        except Exception as e:
            print(f"❌ Error updating PO status: {e}")
    
    def write_enhanced_excel(self, po_entries: List[Dict[str, Any]]) -> None:
        """
        Write the enhanced Excel format with all tracking columns.
        """
        headers = [
            'PO_NUMBER', 'STATUS', 'SUPPLIER', 'ATTACHMENTS_FOUND', 
            'ATTACHMENTS_DOWNLOADED', 'LAST_PROCESSED', 'ERROR_MESSAGE', 'DOWNLOAD_FOLDER', 'COUPA_URL'
        ]
        
        try:
            # Convert to DataFrame
            df = pd.DataFrame([
                {
                    'PO_NUMBER': entry['po_number'],
                    'STATUS': entry['status'],
                    'SUPPLIER': entry['supplier'],
                    'ATTACHMENTS_FOUND': entry['attachments_found'],
                    'ATTACHMENTS_DOWNLOADED': entry['attachments_downloaded'],
                    'LAST_PROCESSED': entry['last_processed'],
                    'ERROR_MESSAGE': entry['error_message'],
                    'DOWNLOAD_FOLDER': entry['download_folder'],
                    'COUPA_URL': entry['coupa_url']
                }
                for entry in po_entries
            ])
            
            # Write to Excel
            df.to_excel(self.excel_file_path, index=False, sheet_name=self.sheet_name)
                    
        except Exception as e:
            print(f"❌ Error writing enhanced Excel: {e}")
    
    def generate_summary_report(self) -> Dict[str, Any]:
        """
        Generate a summary report of processing results.
        """
        po_entries = self.read_po_numbers_from_excel()
        
        # Count by status
        status_counts = {}
        total_attachments_found = 0
        total_attachments_downloaded = 0
        suppliers = set()
        
        for entry in po_entries:
            status = entry['status']
            status_counts[status] = status_counts.get(status, 0) + 1
            total_attachments_found += entry['attachments_found']
            total_attachments_downloaded += entry['attachments_downloaded']
            
            if entry['supplier']:
                suppliers.add(entry['supplier'])
        
        return {
            'total_pos': len(po_entries),
            'status_counts': status_counts,
            'total_attachments_found': total_attachments_found,
            'total_attachments_downloaded': total_attachments_downloaded,
            'unique_suppliers': len(suppliers),
            'supplier_list': sorted(list(suppliers)),
            'success_rate': round((status_counts.get('COMPLETED', 0) / len(po_entries)) * 100, 1) if po_entries else 0
        }
    
    def print_summary_report(self) -> None:
        """Print a formatted summary report to console."""
        report = self.generate_summary_report()
        
        print("\n" + "="*60)
        print("📊 EXCEL PROCESSING SUMMARY REPORT")
        print("="*60)
        
        print(f"📋 Total POs: {report['total_pos']}")
        print(f"🎯 Success Rate: {report['success_rate']}%")
        
        print(f"\n📈 Status Breakdown:")
        for status, count in report['status_counts'].items():
            emoji = {
                'COMPLETED': '✅', 'FAILED': '❌', 'PARTIAL': '⚠️',
                'PENDING': '⏳', 'SKIPPED': '⏭️', 'NO_ATTACHMENTS': '📭'
            }.get(status, '📋')
            print(f"  {emoji} {status}: {count}")
        
        print(f"\n📎 Attachments:")
        print(f"  🔍 Found: {report['total_attachments_found']}")
        print(f"  💾 Downloaded: {report['total_attachments_downloaded']}")
        
        if report['unique_suppliers'] > 0:
            print(f"\n🏢 Suppliers: {report['unique_suppliers']} unique")
            for supplier in report['supplier_list'][:10]:  # Show first 10
                print(f"  📁 {supplier}")
            if len(report['supplier_list']) > 10:
                print(f"  ... and {len(report['supplier_list']) - 10} more")
        
        print("="*60)


class TestExcelProcessorIntegration:
    """Test Excel processor integration with existing codebase"""
    
    def test_excel_processor_interface_compatibility(self):
        """Test that ExcelProcessor has the same interface as CSVProcessor"""
        # Create temporary Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            # Create sample data
            sample_data = [
                {
                    'PO_NUMBER': 'PO15826591',
                    'STATUS': 'PENDING',
                    'SUPPLIER': '',
                    'ATTACHMENTS_FOUND': 0,
                    'ATTACHMENTS_DOWNLOADED': 0,
                    'LAST_PROCESSED': '',
                    'ERROR_MESSAGE': '',
                    'DOWNLOAD_FOLDER': '',
                    'COUPA_URL': ''
                },
                {
                    'PO_NUMBER': 'PO15873456',
                    'STATUS': 'COMPLETED',
                    'SUPPLIER': 'Test_Supplier_Inc',
                    'ATTACHMENTS_FOUND': 3,
                    'ATTACHMENTS_DOWNLOADED': 3,
                    'LAST_PROCESSED': '2024-01-15 14:30:25',
                    'ERROR_MESSAGE': '',
                    'DOWNLOAD_FOLDER': 'Test_Supplier_Inc/',
                    'COUPA_URL': 'https://coupa.company.com/requisition_lines/15873456'
                }
            ]
            
            # Write sample data to Excel
            df = pd.DataFrame(sample_data)
            df.to_excel(tmp_file.name, index=False, sheet_name='PO_Data')
            
            # Test ExcelProcessor
            excel_processor = ExcelProcessor(tmp_file.name)
            
            # Test reading
            po_entries = excel_processor.read_po_numbers_from_excel()
            assert len(po_entries) == 2
            assert po_entries[0]['po_number'] == 'PO15826591'
            assert po_entries[1]['status'] == 'COMPLETED'
            
            # Test processing
            valid_entries = excel_processor.process_po_numbers(po_entries)
            assert len(valid_entries) == 1  # Only PENDING should be processed
            assert valid_entries[0][0] == 'PO15826591'
            
            # Test status update
            excel_processor.update_po_status(
                'PO15826591', 'COMPLETED', 'Test_Supplier', 
                3, 3, '', 'Test_Supplier/', 'https://test.url'
            )
            
            # Verify update
            updated_entries = excel_processor.read_po_numbers_from_excel()
            updated_entry = next(e for e in updated_entries if e['po_number'] == 'PO15826591')
            assert updated_entry['status'] == 'COMPLETED'
            assert updated_entry['supplier'] == 'Test_Supplier'
            
            # Test summary report
            report = excel_processor.generate_summary_report()
            assert report['total_pos'] == 2
            assert report['success_rate'] == 100.0  # Both should be completed now
            
            # Cleanup
            os.unlink(tmp_file.name)
    
    def test_excel_processor_error_handling(self):
        """Test Excel processor error handling"""
        # Test with non-existent file
        excel_processor = ExcelProcessor('non_existent.xlsx')
        po_entries = excel_processor.read_po_numbers_from_excel()
        assert len(po_entries) == 0
        
        # Test with corrupted file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            # Write invalid data
            with open(tmp_file.name, 'w') as f:
                f.write('invalid excel data')
            
            excel_processor = ExcelProcessor(tmp_file.name)
            po_entries = excel_processor.read_po_numbers_from_excel()
            assert len(po_entries) == 0
            
            # Cleanup
            os.unlink(tmp_file.name)
    
    def test_excel_processor_performance(self):
        """Test Excel processor performance with larger datasets"""
        import time
        
        # Create larger dataset
        sample_data = [
            {
                'PO_NUMBER': f'PO{15826591 + i}',
                'STATUS': 'PENDING' if i % 2 == 0 else 'COMPLETED',
                'SUPPLIER': f'Supplier_{i % 10}',
                'ATTACHMENTS_FOUND': i % 5,
                'ATTACHMENTS_DOWNLOADED': i % 3,
                'LAST_PROCESSED': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'ERROR_MESSAGE': '',
                'DOWNLOAD_FOLDER': f'Supplier_{i % 10}/',
                'COUPA_URL': f'https://coupa.company.com/requisition_lines/{15826591 + i}'
            }
            for i in range(500)  # 500 records
        ]
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            # Write data
            df = pd.DataFrame(sample_data)
            df.to_excel(tmp_file.name, index=False, sheet_name='PO_Data')
            
            # Test performance
            excel_processor = ExcelProcessor(tmp_file.name)
            
            start_time = time.time()
            po_entries = excel_processor.read_po_numbers_from_excel()
            read_time = time.time() - start_time
            
            start_time = time.time()
            valid_entries = excel_processor.process_po_numbers(po_entries)
            process_time = time.time() - start_time
            
            start_time = time.time()
            excel_processor.update_po_status('PO15826591', 'COMPLETED')
            update_time = time.time() - start_time
            
            print(f"\nExcel Processor Performance (500 records):")
            print(f"Read time: {read_time:.4f}s")
            print(f"Process time: {process_time:.4f}s")
            print(f"Update time: {update_time:.4f}s")
            
            assert len(po_entries) == 500
            assert len(valid_entries) == 250  # Half should be PENDING
            
            # Cleanup
            os.unlink(tmp_file.name)


class TestExcelProcessorAdvancedFeatures:
    """Test advanced Excel features"""
    
    def test_excel_with_multiple_sheets(self):
        """Test Excel processor with multiple sheets"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            # Create data for multiple sheets
            po_data = [
                {'PO_NUMBER': 'PO15826591', 'STATUS': 'PENDING'},
                {'PO_NUMBER': 'PO15873456', 'STATUS': 'COMPLETED'}
            ]
            
            summary_data = [
                {'METRIC': 'Total POs', 'VALUE': 2},
                {'METRIC': 'Completed', 'VALUE': 1},
                {'METRIC': 'Success Rate', 'VALUE': '50%'}
            ]
            
            # Write to Excel with multiple sheets
            with pd.ExcelWriter(tmp_file.name, engine='openpyxl') as writer:
                pd.DataFrame(po_data).to_excel(writer, sheet_name='PO_Data', index=False)
                pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
            
            # Test reading from specific sheet
            excel_processor = ExcelProcessor(tmp_file.name)
            excel_processor.sheet_name = 'PO_Data'
            
            po_entries = excel_processor.read_po_numbers_from_excel()
            assert len(po_entries) == 2
            
            # Test reading summary sheet
            excel_processor.sheet_name = 'Summary'
            summary_entries = excel_processor.read_po_numbers_from_excel()
            assert len(summary_entries) == 3
            
            # Cleanup
            os.unlink(tmp_file.name)
    
    def test_excel_with_formatting(self):
        """Test Excel processor with formatting"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font
        except ImportError:
            pytest.skip("openpyxl not available for formatting tests")
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            # Create sample data
            data = [
                {'PO_NUMBER': 'PO15826591', 'STATUS': 'COMPLETED'},
                {'PO_NUMBER': 'PO15873456', 'STATUS': 'FAILED'}
            ]
            
            df = pd.DataFrame(data)
            df.to_excel(tmp_file.name, index=False)
            
            # Apply formatting
            wb = load_workbook(tmp_file.name)
            ws = wb.active
            
            # Style header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            wb.save(tmp_file.name)
            
            # Test reading formatted file
            excel_processor = ExcelProcessor(tmp_file.name)
            po_entries = excel_processor.read_po_numbers_from_excel()
            assert len(po_entries) == 2
            
            # Cleanup
            os.unlink(tmp_file.name)


if __name__ == "__main__":
    # Run specific tests for manual exploration
    print("🧪 Excel Processor Integration Testing")
    print("=" * 50)
    
    # Test basic integration
    test_integration = TestExcelProcessorIntegration()
    test_integration.test_excel_processor_interface_compatibility()
    test_integration.test_excel_processor_performance()
    
    print("\n✅ Excel processor integration tests completed!") 