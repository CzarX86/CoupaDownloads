#!/usr/bin/env python3
"""
Excel Sandbox Test Runner
Execute Excel integration tests to evaluate migration decision.
"""

import os
import sys
import subprocess
import tempfile
from pathlib import Path

def install_required_packages():
    """Install required packages for Excel testing"""
    print("📦 Installing required packages...")
    
    packages = [
        'pandas>=1.5.0',
        'openpyxl>=3.0.0',
        'psutil>=5.8.0'  # For memory testing
    ]
    
    for package in packages:
        try:
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', package])
            print(f"✅ Installed {package}")
        except subprocess.CalledProcessError:
            print(f"⚠️ Failed to install {package}")

def run_basic_excel_tests():
    """Run basic Excel functionality tests"""
    print("\n🧪 Running Basic Excel Tests")
    print("=" * 40)
    
    try:
        import pandas as pd
        import tempfile
        
        # Test 1: Basic Excel creation
        print("Test 1: Basic Excel file creation...")
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            data = [
                {'PO_NUMBER': 'PO15826591', 'STATUS': 'PENDING'},
                {'PO_NUMBER': 'PO15873456', 'STATUS': 'COMPLETED'}
            ]
            df = pd.DataFrame(data)
            df.to_excel(tmp_file.name, index=False)
            
            # Verify file was created
            assert os.path.exists(tmp_file.name)
            print("✅ Excel file created successfully")
            
            # Read back and verify
            read_df = pd.read_excel(tmp_file.name)
            assert len(read_df) == 2
            print("✅ Excel file read successfully")
            
            os.unlink(tmp_file.name)
        
        # Test 2: Multiple sheets
        print("Test 2: Multiple sheets...")
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            po_data = [{'PO_NUMBER': 'PO15826591', 'STATUS': 'PENDING'}]
            summary_data = [{'METRIC': 'Total POs', 'VALUE': 1}]
            
            with pd.ExcelWriter(tmp_file.name, engine='openpyxl') as writer:
                pd.DataFrame(po_data).to_excel(writer, sheet_name='PO_Data', index=False)
                pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
            
            # Verify both sheets
            po_df = pd.read_excel(tmp_file.name, sheet_name='PO_Data')
            summary_df = pd.read_excel(tmp_file.name, sheet_name='Summary')
            
            assert len(po_df) == 1
            assert len(summary_df) == 1
            print("✅ Multiple sheets created successfully")
            
            os.unlink(tmp_file.name)
        
        print("✅ All basic Excel tests passed!")
        return True
        
    except Exception as e:
        print(f"❌ Basic Excel tests failed: {e}")
        return False

def run_performance_comparison():
    """Run performance comparison between CSV and Excel"""
    print("\n⚡ Performance Comparison")
    print("=" * 40)
    
    try:
        import pandas as pd
        import csv
        import time
        import tempfile
        
        # Generate test data
        test_data = [
            {
                'PO_NUMBER': f'PO{15826591 + i}',
                'STATUS': 'PENDING' if i % 2 == 0 else 'COMPLETED',
                'SUPPLIER': f'Supplier_{i}',
                'ATTACHMENTS_FOUND': i % 5,
                'ATTACHMENTS_DOWNLOADED': i % 3
            }
            for i in range(100)
        ]
        
        # CSV performance
        csv_start = time.time()
        with tempfile.NamedTemporaryFile(suffix='.csv', mode='w', delete=False) as csv_file:
            writer = csv.DictWriter(csv_file, fieldnames=test_data[0].keys())
            writer.writeheader()
            writer.writerows(test_data)
            csv_path = csv_file.name
        csv_write_time = time.time() - csv_start
        
        # Excel performance
        excel_start = time.time()
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as excel_file:
            df = pd.DataFrame(test_data)
            df.to_excel(excel_file.name, index=False)
            excel_path = excel_file.name
        excel_write_time = time.time() - excel_start
        
        # Read performance
        csv_read_start = time.time()
        with open(csv_path, 'r') as f:
            reader = csv.DictReader(f)
            csv_data = list(reader)
        csv_read_time = time.time() - csv_read_start
        
        excel_read_start = time.time()
        excel_data = pd.read_excel(excel_path).to_dict('records')
        excel_read_time = time.time() - excel_read_start
        
        # Report results
        print(f"Dataset: 100 records")
        print(f"CSV Write:  {csv_write_time:.4f}s")
        print(f"Excel Write: {excel_write_time:.4f}s")
        print(f"CSV Read:   {csv_read_time:.4f}s")
        print(f"Excel Read:  {excel_read_time:.4f}s")
        print(f"Write ratio: {excel_write_time/csv_write_time:.2f}x")
        print(f"Read ratio:  {excel_read_time/csv_read_time:.2f}x")
        
        # Cleanup
        os.unlink(csv_path)
        os.unlink(excel_path)
        
        return True
        
    except Exception as e:
        print(f"❌ Performance comparison failed: {e}")
        return False

def run_formatting_tests():
    """Test Excel formatting capabilities"""
    print("\n🎨 Excel Formatting Tests")
    print("=" * 40)
    
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
        import tempfile
        
        # Create sample data
        data = [
            {'PO_NUMBER': 'PO15826591', 'STATUS': 'COMPLETED'},
            {'PO_NUMBER': 'PO15873456', 'STATUS': 'FAILED'}
        ]
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            df = pd.DataFrame(data)
            df.to_excel(tmp_file.name, index=False)
            
            # Apply formatting
            wb = load_workbook(tmp_file.name)
            ws = wb.active
            
            # Style header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            wb.save(tmp_file.name)
            
            # Verify file still readable
            read_df = pd.read_excel(tmp_file.name)
            assert len(read_df) == 2
            print("✅ Excel formatting applied successfully")
            
            os.unlink(tmp_file.name)
        
        return True
        
    except ImportError:
        print("⚠️ openpyxl not available for formatting tests")
        return False
    except Exception as e:
        print(f"❌ Formatting tests failed: {e}")
        return False

def run_memory_usage_test():
    """Test memory usage of Excel operations"""
    print("\n💾 Memory Usage Test")
    print("=" * 40)
    
    try:
        import psutil
        import pandas as pd
        import tempfile
        import gc
        
        process = psutil.Process()
        
        # Generate larger dataset
        test_data = [
            {
                'PO_NUMBER': f'PO{15826591 + i}',
                'STATUS': 'PENDING' if i % 2 == 0 else 'COMPLETED',
                'SUPPLIER': f'Supplier_{i % 10}',
                'ATTACHMENTS_FOUND': i % 5,
                'ATTACHMENTS_DOWNLOADED': i % 3
            }
            for i in range(1000)
        ]
        
        # Test Excel memory usage
        gc.collect()
        initial_memory = process.memory_info().rss / 1024 / 1024  # MB
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as excel_file:
            df = pd.DataFrame(test_data)
            df.to_excel(excel_file.name, index=False)
            
            peak_memory = process.memory_info().rss / 1024 / 1024  # MB
            excel_path = excel_file.name
        
        # Read back
        read_df = pd.read_excel(excel_path)
        read_memory = process.memory_info().rss / 1024 / 1024  # MB
        
        print(f"Initial memory: {initial_memory:.2f} MB")
        print(f"Write peak: {peak_memory:.2f} MB (+{peak_memory - initial_memory:.2f} MB)")
        print(f"Read peak: {read_memory:.2f} MB (+{read_memory - initial_memory:.2f} MB)")
        
        # Cleanup
        os.unlink(excel_path)
        del df, read_df
        gc.collect()
        
        return True
        
    except ImportError:
        print("⚠️ psutil not available for memory testing")
        return False
    except Exception as e:
        print(f"❌ Memory usage test failed: {e}")
        return False

def generate_recommendation(test_results):
    """Generate recommendation based on test results"""
    print("\n🎯 Migration Recommendation")
    print("=" * 40)
    
    passed_tests = sum(test_results.values())
    total_tests = len(test_results)
    
    print(f"Tests passed: {passed_tests}/{total_tests}")
    
    if passed_tests == total_tests:
        print("✅ All tests passed! Excel integration is viable.")
        print("💡 Recommendation: Consider implementing Excel support")
        print("   - Add Excel export functionality alongside CSV")
        print("   - Use Excel for reporting and user-facing features")
        print("   - Keep CSV for core processing operations")
    elif passed_tests >= total_tests * 0.7:
        print("⚠️ Most tests passed. Excel integration is mostly viable.")
        print("💡 Recommendation: Implement with caution")
        print("   - Test thoroughly in staging environment")
        print("   - Consider hybrid approach")
    else:
        print("❌ Many tests failed. Excel integration may be problematic.")
        print("💡 Recommendation: Stick with CSV for now")
        print("   - Focus on improving CSV functionality")
        print("   - Reconsider Excel integration later")

def main():
    """Main test runner"""
    print("🧪 Excel Integration Sandbox Testing")
    print("=" * 50)
    
    # Install required packages
    install_required_packages()
    
    # Run tests
    test_results = {
        'Basic Excel Tests': run_basic_excel_tests(),
        'Performance Comparison': run_performance_comparison(),
        'Formatting Tests': run_formatting_tests(),
        'Memory Usage Test': run_memory_usage_test()
    }
    
    # Generate recommendation
    generate_recommendation(test_results)
    
    print("\n✅ Sandbox testing completed!")

if __name__ == "__main__":
    main() 