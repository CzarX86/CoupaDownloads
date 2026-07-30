from collections import deque
from pathlib import Path

import pytest

from src.gui.cli_supervisor import CliProcessSupervisor


def test_recent_progress_metrics_use_rolling_completion_window():
    supervisor = CliProcessSupervisor()
    supervisor.started_at = 1_000.0
    supervisor._progress_samples = deque([(1_000.0, 0)], maxlen=180)
    supervisor._last_processed = 0
    supervisor._last_progress_at = 1_000.0

    speed, eta, stalled = supervisor._recent_progress_metrics(10, 40, now=1_030.0)

    assert speed == pytest.approx(20.0)
    assert eta == "01:30"
    assert stalled == 0

    speed, eta, stalled = supervisor._recent_progress_metrics(10, 40, now=1_100.0)

    assert speed == 0.0
    assert eta == "--:--"
    assert stalled == 70


def test_eta_format_does_not_wrap_after_one_hour():
    assert CliProcessSupervisor._format_eta(65) == "01:05"
    assert CliProcessSupervisor._format_eta(3_661) == "1h 01m"


def test_actionable_logs_keep_progress_and_drop_timing_noise():
    assert CliProcessSupervisor._actionable_log_entry(
        "[TIMING] _fetch_html po: total=25ms http=20ms"
    ) is None

    entry = CliProcessSupervisor._actionable_log_entry(
        "[  25/100] ok=23 err=2 files=41  speed=50/min  eta=1m30s"
    )

    assert entry == {
        "type": "Warning",
        "message": "Progress: 25/100 POs · 23 succeeded · 2 failed · 41 attachments",
    }


def test_sidebar_owns_viewport_and_main_content_owns_scroll():
    css = (Path(__file__).parents[1] / "src" / "gui" / "web" / "style.css").read_text(encoding="utf-8")

    assert "height: 100vh" in css
    assert "overflow: hidden" in css
    assert ".main-content" in css and "overflow-y: auto" in css
    assert "zoom: calc(1 / var(--font-scale))" in css


def test_pythonw_gui_uses_console_python_for_cli_worker(monkeypatch, tmp_path):
    import src.gui.cli_supervisor as supervisor_module

    runtime = tmp_path / "runtime"
    runtime.mkdir()
    pythonw = runtime / "pythonw.exe"
    python = runtime / "python.exe"
    pythonw.touch()
    python.touch()
    supervisor = CliProcessSupervisor()
    monkeypatch.setattr(supervisor_module.sys, "executable", str(pythonw))
    monkeypatch.setattr(supervisor_module.sys, "platform", "win32")

    command = supervisor._command(concurrency=4)

    assert command[0] == str(python)
    assert command[1].endswith("process_all_pos.py")


def test_retry_file_update_replaces_po_and_adds_remark(tmp_path):
    from openpyxl import Workbook, load_workbook

    path = tmp_path / "input.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Input"
    sheet.append(["PO_NUMBER", "SUPPLIER"])
    sheet.append(["12345678", "ACME"])
    workbook.save(path)

    CliProcessSupervisor._replace_po_in_file(
        path,
        "12345678",
        "1234567",
        "Corrected from 12345678 to 1234567; retry succeeded.",
    )

    row = list(load_workbook(path, read_only=True).active.iter_rows(values_only=True))[1]
    assert row == ("1234567", "ACME", "Corrected from 12345678 to 1234567; retry succeeded.")

    # Existing archived values may have a PO prefix or numeric Excel formatting.
    assert CliProcessSupervisor._po_values_equal("168012998", "PO168012998")
    assert CliProcessSupervisor._po_values_equal("168012998.0", "PO168012998")


def test_history_status_translation_preserves_filter_inputs():
    web_root = Path(__file__).parents[1] / "src" / "gui" / "web"
    html = (web_root / "index.html").read_text(encoding="utf-8")
    javascript = (web_root / "app.js").read_text(encoding="utf-8")

    assert 'data-status="SUCCESS" checked><span>Success</span>' in html
    assert '"#status-filter legend, #status-filter label span"' in javascript
    assert '"#status-filter legend, #status-filter label"' not in javascript
    assert 'class="coupa-column"' not in html
    assert 'class="coupa-link po-number-link"' in javascript
    assert 'id="btn-check-updates"' in html
    assert 'id="btn-start-over"' in html
    assert 'reset_new_run' in javascript
    assert 'get_authentication_status' in javascript
    assert 'Action required' in javascript
    assert 'checkForUpdates(true)' in javascript
    assert '$("#modal-pos-tbody").addEventListener("click"' in javascript
    assert 'retry_po_with_edit' in javascript
    assert 'save_retry_attempt' in javascript
    assert 'discard_retry_attempt' in javascript
    assert 'id="retry-result-modal"' in html
    assert 'id="btn-save-retry-result"' in html
    assert 'id="btn-discard-retry-result"' in html
    assert '#retry-edit-modal, #retry-result-modal' in (web_root / "style.css").read_text(encoding="utf-8")
