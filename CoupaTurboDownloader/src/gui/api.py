import json
import os
import sys
import time
import asyncio
import platform
import plistlib
import shutil
import subprocess
import threading
from datetime import datetime, timezone
from pathlib import Path, PurePosixPath
from typing import Dict, Any, List, Optional

import pandas as pd

from src.db.session_db import SessionDB, PODownload
from src.engine.crawler import CoupaCrawler
from src.engine.tls import system_ssl_context
from src.engine.authenticator import clear_cached_authentication, load_cached_cookies, validate_cookies_detailed


class AppAPI:
    DEFAULT_SETTINGS = {
        "concurrency": 4,
        "retry_attempts": 1,
        "auto_updates": True,
        "retention": "all",
        "msg_processing": "convert_extract",
        "deduplicate_files": True,
        "language": "en",
        "font_scale": 1.1,
    }

    def __init__(self, db: SessionDB, default_download_dir: str):
        self.db = db
        self._settings_path = Path.home() / ".coupa_turbo" / "gui_settings.json"
        self.default_download_dir = self._load_download_root(default_download_dir)
        self._runtime_lock = threading.Lock()
        self._cookies: Optional[Dict[str, str]] = None
        self._runtime: Dict[int, Dict[str, Any]] = {}
        self._file_observations: Dict[str, tuple[tuple[int, int], float]] = {}

    def _set_session_status(self, session_id: int, status: str):
        cursor = self.db.conn.cursor()
        cursor.execute(
            "UPDATE sessions SET status = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (status, session_id),
        )
        self.db.conn.commit()

    @staticmethod
    def _is_python_portable() -> bool:
        return os.environ.get("COUPA_PYTHON_PORTABLE") == "1"

    def _read_settings(self) -> Dict[str, Any]:
        settings = dict(self.DEFAULT_SETTINGS)
        if self._is_python_portable():
            settings["auto_updates"] = False
        try:
            stored = json.loads(self._settings_path.read_text(encoding="utf-8"))
            if isinstance(stored, dict):
                settings.update(stored)
        except (OSError, ValueError):
            pass
        settings["concurrency"] = max(1, min(8, int(settings.get("concurrency", 4))))
        settings["retry_attempts"] = max(1, min(3, int(settings.get("retry_attempts", 1))))
        settings["auto_updates"] = bool(settings.get("auto_updates", True))
        if settings.get("retention") not in {"all", "10", "30", "90"}:
            settings["retention"] = "all"
        if settings.get("msg_processing") not in {"disabled", "convert", "convert_extract"}:
            settings["msg_processing"] = "convert_extract"
        settings["deduplicate_files"] = bool(settings.get("deduplicate_files", True))
        settings["language"] = "pt-BR" if settings.get("language") == "pt-BR" else "en"
        try:
            settings["font_scale"] = max(1.0, min(1.3, float(settings.get("font_scale", 1.1))))
        except (TypeError, ValueError):
            settings["font_scale"] = 1.1
        return settings

    @staticmethod
    def _absolute_user_path(value: str) -> str:
        path = Path(str(value or "")).expanduser()
        if not path.is_absolute():
            path = Path.home() / path
        return str(path.resolve(strict=False))

    def _load_download_root(self, fallback: str) -> str:
        try:
            settings = self._read_settings()
            value = str(settings.get("download_root", "")).strip()
            if value:
                return self._absolute_user_path(value)
        except (OSError, ValueError):
            pass
        return self._absolute_user_path(fallback)

    def _persist_download_root(self, value: str) -> str:
        root = self._absolute_user_path(value)
        try:
            self._settings_path.parent.mkdir(parents=True, exist_ok=True)
            settings = {}
            if self._settings_path.exists():
                settings = json.loads(self._settings_path.read_text(encoding="utf-8"))
            settings["download_root"] = root
            self._settings_path.write_text(json.dumps(settings, indent=2), encoding="utf-8")
        except (OSError, ValueError):
            pass
        self.default_download_dir = root
        return root

    def get_app_settings(self) -> Dict[str, Any]:
        settings = self._read_settings()
        settings["download_root"] = self.default_download_dir
        settings["python_portable"] = self._is_python_portable()
        return settings

    def set_app_settings(self, values: Dict[str, Any]) -> Dict[str, Any]:
        current = self._read_settings()
        try:
            current["concurrency"] = max(1, min(8, int(values.get("concurrency", current["concurrency"]))))
            current["retry_attempts"] = max(1, min(3, int(values.get("retry_attempts", current["retry_attempts"]))))
        except (TypeError, ValueError):
            return {"success": False, "error": "Concurrency and retry values must be numeric."}
        current["auto_updates"] = bool(values.get("auto_updates", current["auto_updates"]))
        retention = str(values.get("retention", current["retention"]))
        if retention not in {"all", "10", "30", "90"}:
            return {"success": False, "error": "Invalid history retention option."}
        current["retention"] = retention
        msg_processing = str(values.get("msg_processing", current["msg_processing"]))
        if msg_processing not in {"disabled", "convert", "convert_extract"}:
            return {"success": False, "error": "Invalid MSG processing option."}
        current["msg_processing"] = msg_processing
        current["deduplicate_files"] = bool(values.get("deduplicate_files", current["deduplicate_files"]))
        current["language"] = "pt-BR" if values.get("language") == "pt-BR" else "en"
        try:
            current["font_scale"] = max(1.0, min(1.3, float(values.get("font_scale", current["font_scale"]))))
        except (TypeError, ValueError):
            return {"success": False, "error": "Text size must be numeric."}
        root = str(values.get("download_root", self.default_download_dir)).strip()
        if not root:
            return {"success": False, "error": "Download folder cannot be empty."}
        current["download_root"] = self._absolute_user_path(root)
        try:
            self._settings_path.parent.mkdir(parents=True, exist_ok=True)
            self._settings_path.write_text(json.dumps(current, indent=2), encoding="utf-8")
        except OSError as exc:
            return {"success": False, "error": f"Could not save settings: {exc}"}
        self.default_download_dir = current["download_root"]
        return {"success": True, "settings": current}

    def _new_run_directory(self, root: str):
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        return str(Path(root).expanduser() / f"run_{timestamp}")

    def get_default_download_directory(self) -> str:
        return self._new_run_directory(self.default_download_dir)

    def set_default_download_directory(self, value: str) -> Dict[str, Any]:
        if not str(value or "").strip():
            return {"success": False, "error": "Download folder cannot be empty."}
        root = self._persist_download_root(value)
        return {"success": True, "path": root}

    @staticmethod
    def _native_window():
        """Return the active pywebview window when running inside the desktop app."""
        try:
            import webview
            windows = getattr(webview, "windows", [])
            return windows[0] if windows else None
        except Exception:
            return None

    @staticmethod
    def _open_path(path: Path) -> None:
        """Open a file with the platform default application."""
        if sys.platform.startswith("win"):
            os.startfile(str(path))
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(path)])
        else:
            subprocess.Popen(["xdg-open", str(path)])

    def open_input_path(self, filepath: str) -> Dict[str, Any]:
        """Open a user-selected input file in the platform default editor."""
        path = Path(str(filepath or "")).expanduser().resolve()
        if not path.is_file():
            return {"success": False, "error": "The selected input file could not be found."}
        try:
            self._open_path(path)
            return {"success": True, "path": str(path)}
        except (OSError, ValueError) as exc:
            return {"success": False, "error": f"Could not open the input file: {exc}"}

    def select_file(self) -> Dict[str, Any]:
        """Open a native input-file dialog and return the selected path."""
        window = self._native_window()
        if window is None:
            return {"success": False, "error": "Native file dialog is unavailable."}
        try:
            import webview
            paths = window.create_file_dialog(
                webview.OPEN_DIALOG,
                allow_multiple=False,
                file_types=("Input files (*.csv;*.xlsx;*.xls)",),
            )
            path = paths[0] if paths else ""
            return {"success": bool(path), "path": path}
        except Exception as exc:
            return {"success": False, "error": str(exc)}

    def select_directory(self) -> str:
        """Open a native destination-folder dialog."""
        Path(self.default_download_dir).mkdir(parents=True, exist_ok=True)
        window = self._native_window()
        if window is None:
            return self.get_default_download_directory()
        try:
            import webview
            paths = window.create_file_dialog(webview.FOLDER_DIALOG)
            if paths:
                root = self._persist_download_root(str(paths[0]))
                return self._new_run_directory(root)
            return self.get_default_download_directory()
        except Exception:
            return self.default_download_dir

    def _append_log(self, session_id: int, log_type: str, message: str):
        with self._runtime_lock:
            runtime = self._runtime.get(session_id)
            if not runtime:
                return
            runtime["latest_logs"].append({"type": log_type, "message": message})
            runtime["latest_logs"] = runtime["latest_logs"][-30:]

    async def _run_session_async(self, session_id: int, download_dir: str):
        cursor = self.db.conn.cursor()
        cursor.execute(
            "SELECT po_number, company_code FROM po_downloads WHERE session_id = ? AND status = 'PENDING' ORDER BY id ASC",
            (session_id,),
        )
        rows = cursor.fetchall()

        runtime = self._runtime.get(session_id, {})
        concurrency = max(1, int(runtime.get("concurrency", 11)))

        # Load cached cookies if not already set (from authenticate step).
        cookies = self._cookies or load_cached_cookies()
        if cookies:
            self._cookies = cookies

        crawler = CoupaCrawler(
            self.db,
            session_id,
            download_dir,
            concurrency=concurrency,
            cookies=cookies,
        )
        self._set_session_status(session_id, "RUNNING")
        self._append_log(
            session_id,
            "System",
            f"Starting session {session_id} with {len(rows)} POs (workers={concurrency})",
        )

        semaphore = asyncio.Semaphore(concurrency)
        stopped = False

        async def process_one(po_number: str, company_code: str):
            async with semaphore:
                _check_runtime_stop(session_id, self._runtime, self._runtime_lock)
                await _check_runtime_pause(session_id, self._runtime, self._runtime_lock)

                self._append_log(session_id, "Info", f"Processing PO {po_number} ({company_code})")
                result = await crawler.process_po(po_number, company_code)
                po_row = self.db.get_po(session_id, po_number)
                po_status = (po_row or {}).get("status", "ERROR")
                failed = po_status in {"ERROR", "SKIPPED_VERIFICATION_REQUIRED"}

                with self._runtime_lock:
                    current = self._runtime.get(session_id)
                    if current:
                        current["processed"] += 1
                        if failed:
                            current["errors"] += 1

                if failed:
                    self._append_log(
                        session_id,
                        "Error",
                        f"PO {po_number} failed: {(po_row or {}).get('error_message') or po_status}",
                    )
                else:
                    self._append_log(session_id, "Success", f"PO {po_number} completed")
                return result

        try:
            tasks = [process_one(row["po_number"], row["company_code"]) for row in rows]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            for result in results:
                if isinstance(result, SessionStoppedError):
                    stopped = True
                    continue
                if isinstance(result, Exception):
                    self._append_log(session_id, "Error", f"Task error: {result}")
                    with self._runtime_lock:
                        current = self._runtime.get(session_id)
                        if current:
                            current["errors"] += 1

            with self._runtime_lock:
                current = self._runtime.get(session_id)
                processed = int(current.get("processed", 0)) if current else 0
                errors = int(current.get("errors", 0)) if current else 0
                pending = max(len(rows) - processed, 0)
                if stopped or (current and current.get("stop_requested")):
                    final_status = "STOPPED"
                elif errors == 0 and pending == 0:
                    final_status = "SUCCESS"
                elif processed > errors:
                    final_status = "PARTIAL"
                else:
                    final_status = "FAILED"
                if current:
                    current["status"] = final_status

            self._set_session_status(session_id, final_status)
            message = f"Session {session_id} finished with status: {final_status}"
            self._append_log(session_id, "System", message)

        finally:
            await crawler.close()

    def _run_session_thread(self, session_id: int, download_dir: str):
        try:
            asyncio.run(self._run_session_async(session_id, download_dir))
        except Exception as e:
            with self._runtime_lock:
                runtime = self._runtime.get(session_id)
                if runtime:
                    runtime["status"] = "FAILED"
                    runtime["errors"] += 1
            self._set_session_status(session_id, "FAILED")
            self._append_log(session_id, "Error", f"Fatal session error: {e}")

    def inspect_input_file(self, filepath: str) -> Dict[str, Any]:
        """Return a cross-platform best-effort state for an editable input file."""
        path = Path(filepath)
        if not path.exists():
            return {"exists": False, "ready": False, "open_detected": False, "stable": False, "error": "File not found."}

        try:
            stat = path.stat()
            fingerprint = (int(stat.st_mtime_ns), int(stat.st_size))
            now = time.monotonic()
            previous = self._file_observations.get(str(path))
            stable = bool(previous and previous[0] == fingerprint and now - previous[1] >= 0.8)
            if not previous or previous[0] != fingerprint:
                self._file_observations[str(path)] = (fingerprint, now)

            lock_names = [
                f"~${path.name}",
                f".~lock.{path.name}#",
                f".{path.name}.lock",
            ]
            lock_files = [str(path.parent / name) for name in lock_names if (path.parent / name).exists()]
            readable = os.access(path, os.R_OK)
            open_detected = bool(lock_files)
            return {
                "exists": True,
                "readable": readable,
                "stable": stable,
                "open_detected": open_detected,
                "lock_files": lock_files,
                "mtime_ns": fingerprint[0],
                "size": fingerprint[1],
                "ready": readable and stable and not open_detected,
                "error": None,
            }
        except OSError as exc:
            return {"exists": True, "ready": False, "open_detected": True, "stable": False, "error": str(exc)}

    def start_download(self, session_id: int, download_dir: str, concurrency: int = 11) -> Dict[str, Any]:
        try:
            session_id = int(session_id)
            download_dir = (download_dir or self.default_download_dir).strip() or self.default_download_dir
            os.makedirs(download_dir, exist_ok=True)
            self.default_download_dir = download_dir

            cursor = self.db.conn.cursor()
            cursor.execute(
                "SELECT COUNT(*) as total FROM po_downloads WHERE session_id = ? AND status = 'PENDING'",
                (session_id,),
            )
            total = int(cursor.fetchone()["total"])

            if total == 0:
                return {"success": False, "error": "No pending POs found in this session."}

            with self._runtime_lock:
                existing = self._runtime.get(session_id)
                if existing and existing.get("status") in {"RUNNING", "PAUSED"}:
                    return {"success": False, "error": "Session already running."}

                self._runtime[session_id] = {
                    "status": "RUNNING",
                    "started_at": time.time(),
                    "processed": 0,
                    "total": total,
                    "errors": 0,
                    "paused": False,
                    "stop_requested": False,
                    "concurrency": concurrency,
                    "latest_logs": [],
                }

            self._set_session_status(session_id, "RUNNING")

            worker = threading.Thread(
                target=self._run_session_thread,
                args=(session_id, download_dir),
                daemon=True,
            )
            with self._runtime_lock:
                self._runtime[session_id]["thread"] = worker

            worker.start()
            return {"success": True, "session_id": session_id, "total": total}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def get_active_session_status(self, session_id: int) -> Dict[str, Any]:
        session_id = int(session_id)
        with self._runtime_lock:
            runtime = self._runtime.get(session_id)

        if runtime:
            elapsed = max(time.time() - runtime.get("started_at", time.time()), 0.1)
            processed = int(runtime.get("processed", 0))
            total = int(runtime.get("total", 0))
            errors = int(runtime.get("errors", 0))
            speed = (processed / elapsed) * 60.0 if processed > 0 else 0.0
            remaining = max(total - processed, 0)
            eta_seconds = int((remaining / speed) * 60) if speed > 0 else 0
            eta = time.strftime("%M:%S", time.gmtime(eta_seconds)) if speed > 0 else "--:--"

            with self._runtime_lock:
                latest_logs = list(runtime.get("latest_logs", []))
                runtime["latest_logs"] = []
                status = runtime.get("status", "PENDING")

            return {
                "status": status,
                "total": total,
                "processed": processed,
                "speed": speed,
                "eta": eta,
                "errors": errors,
                "latest_logs": latest_logs,
            }

        cursor = self.db.conn.cursor()
        cursor.execute("SELECT status FROM sessions WHERE id = ?", (session_id,))
        session_row = cursor.fetchone()
        session_status = session_row["status"] if session_row else "UNKNOWN"

        cursor.execute(
            """
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN status != 'PENDING' THEN 1 ELSE 0 END) as processed,
                SUM(CASE WHEN status = 'ERROR' THEN 1 ELSE 0 END) as errors
            FROM po_downloads
            WHERE session_id = ?
            """,
            (session_id,),
        )
        stats = cursor.fetchone()
        total = int(stats["total"] or 0)
        processed = int(stats["processed"] or 0)
        errors = int(stats["errors"] or 0)

        return {
            "status": session_status,
            "total": total,
            "processed": processed,
            "speed": 0.0,
            "eta": "--:--",
            "errors": errors,
            "latest_logs": [],
        }

    def pause_download(self, session_id: int) -> Dict[str, Any]:
        try:
            session_id = int(session_id)
            with self._runtime_lock:
                runtime = self._runtime.get(session_id)
                if not runtime:
                    return {"success": False, "error": "Session not active."}
                runtime["paused"] = True
                runtime["status"] = "PAUSED"
            self._set_session_status(session_id, "PAUSED")
            self._append_log(session_id, "System", "Session paused")
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def resume_download(self, session_id: int) -> Dict[str, Any]:
        try:
            session_id = int(session_id)
            with self._runtime_lock:
                runtime = self._runtime.get(session_id)
                if not runtime:
                    return {"success": False, "error": "Session not active."}
                runtime["paused"] = False
                runtime["status"] = "RUNNING"
            self._set_session_status(session_id, "RUNNING")
            self._append_log(session_id, "System", "Session resumed")
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def stop_download(self, session_id: int) -> Dict[str, Any]:
        try:
            session_id = int(session_id)
            with self._runtime_lock:
                runtime = self._runtime.get(session_id)
                if not runtime:
                    return {"success": False, "error": "Session not active."}
                runtime["stop_requested"] = True
            self._append_log(session_id, "System", "Stop requested by user")
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": str(e)}

    @staticmethod
    def _safe_user_path(path: str) -> str:
        """Redact the user's home directory before a report is shared."""
        value = str(path or "")
        home = str(Path.home())
        return value.replace(home, "~")

    @staticmethod
    def _package_version(distribution: str, import_name: str | None = None) -> str:
        try:
            from importlib.metadata import version
            return version(distribution)
        except Exception:
            try:
                import importlib
                module = importlib.import_module(import_name or distribution.replace("-", "_"))
                return str(getattr(module, "__version__", "bundled"))
            except Exception:
                return "not installed"

    @staticmethod
    def _application_version() -> str:
        candidates = []
        if getattr(sys, "_MEIPASS", None):
            candidates.append(Path(sys._MEIPASS) / ".version")
        candidates.extend([
            Path(__file__).resolve().parents[2] / ".version",
            Path.cwd() / ".version",
        ])
        for path in candidates:
            try:
                value = path.read_text(encoding="utf-8").strip()
                if value:
                    return value.lstrip("v")
            except OSError:
                continue
        return "unknown"

    @staticmethod
    def _find_edge() -> Optional[str]:
        candidates = [
            shutil.which("msedge"),
            shutil.which("microsoft-edge"),
            "/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge",
            r"C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedge.exe",
            r"C:\\Program Files\\Microsoft\\Edge\\Application\\msedge.exe",
        ]
        return next((candidate for candidate in candidates if candidate and Path(candidate).exists()), None)

    @staticmethod
    def _edge_driver_path() -> Optional[str]:
        """Resolve the EdgeDriver Selenium will use without starting Edge."""
        try:
            from selenium.webdriver.common.selenium_manager import SeleniumManager

            result = SeleniumManager().binary_paths(["--browser", "edge"])
            path = result.get("driver_path") if isinstance(result, dict) else None
            return str(path) if path and Path(path).exists() else None
        except Exception:
            return shutil.which("msedgedriver") or shutil.which("msedgedriver.exe")

    @staticmethod
    def _edge_version(edge_path: str | None) -> str:
        if not edge_path:
            return "Microsoft Edge was not found"
        if sys.platform == "darwin":
            info_plist = Path(edge_path).parents[1] / "Info.plist"
            try:
                with info_plist.open("rb") as stream:
                    version = plistlib.load(stream).get("CFBundleShortVersionString")
                if version:
                    return str(version)
            except (OSError, ValueError, KeyError):
                pass
        return AppAPI._command_version([edge_path, "--version"])

    @staticmethod
    def _version_components(value: str) -> tuple[int, ...]:
        import re

        match = re.search(r"\b(\d+(?:\.\d+){1,3})\b", value or "")
        return tuple(int(part) for part in match.group(1).split(".")) if match else ()

    @staticmethod
    def _command_version(command: list[str]) -> str:
        try:
            result = subprocess.run(command, capture_output=True, text=True, timeout=4, check=False)
            return (result.stdout or result.stderr).strip().splitlines()[0][:160]
        except Exception as exc:
            return f"unavailable ({exc})"

    @staticmethod
    def _clipboard_copy(text: str) -> tuple[bool, str]:
        commands = []
        if sys.platform == "darwin":
            commands = [["pbcopy"]]
        elif sys.platform.startswith("win"):
            commands = [["clip"]]
        else:
            commands = [["xclip", "-selection", "clipboard"], ["xsel", "--clipboard", "--input"]]
        for command in commands:
            try:
                subprocess.run(command, input=text, text=True, check=True, timeout=5)
                return True, "Clipboard updated"
            except (FileNotFoundError, subprocess.SubprocessError):
                continue
        return False, "Native clipboard utility unavailable"

    def run_diagnostics(self, input_path: str = "") -> Dict[str, Any]:
        """Run safe local diagnostics without reading credentials or cookies."""
        checks: list[Dict[str, str]] = []

        def check(name: str, status: str, detail: str) -> None:
            checks.append({"name": name, "status": status, "detail": detail})

        edge_path = self._find_edge()
        check("Operating system", "PASS", f"{platform.system()} {platform.release()} ({platform.machine()})")
        check("Python runtime", "PASS", platform.python_version())
        check("Application version", "PASS", self._application_version())
        edge_version = self._edge_version(edge_path)
        edge_status = "PASS" if edge_path and not edge_version.startswith("unavailable") else "WARN"
        check("Microsoft Edge", edge_status, edge_version)

        # ---- Edge profile ----
        try:
            from src.engine.authenticator import _edge_user_data_dir, _edge_profile_directory
            user_data = _edge_user_data_dir()
            if user_data:
                profile = _edge_profile_directory(user_data)
                check("Edge profile", "PASS", f"{self._safe_user_path(str(user_data))}/{profile}")
            else:
                check("Edge profile", "WARN", "Edge user-data directory was not found")
        except Exception as exc:
            check("Edge profile", "WARN", f"Could not detect the Edge profile: {exc}")

        # ---- Coupa session ----
        try:
            cookies = load_cached_cookies()
            if not cookies:
                check("Coupa session", "WARN", "No cached Coupa session; sign-in is required")
            else:
                valid, reason = asyncio.run(validate_cookies_detailed(cookies))
                if valid:
                    check("Coupa session", "PASS", "Cached session is valid")
                elif reason == "unavailable":
                    check("Coupa session", "WARN", "Session validation unavailable — Coupa or network may be unreachable")
                else:
                    check("Coupa session", "WARN", "Cached session expired; re-authentication is required")
        except Exception as exc:
            check("Coupa session", "WARN", f"Could not validate session: {exc}")

        # ---- Python portable edition ----
        if self._is_python_portable():
            check("Distribution", "PASS", "Python portable edition (no installation required)")

        driver_path = self._edge_driver_path()
        driver_version = self._command_version([driver_path, "--version"]) if driver_path else "EdgeDriver was not resolved by Selenium Manager"
        edge_numbers = self._version_components(edge_version)
        driver_numbers = self._version_components(driver_version)
        if not driver_path:
            driver_status = "WARN"
            driver_detail = driver_version
        elif edge_numbers and driver_numbers and edge_numbers[:3] != driver_numbers[:3]:
            driver_status = "WARN"
            driver_detail = f"{driver_version} — browser/driver versions differ"
        else:
            driver_status = "PASS"
            driver_detail = f"{driver_version} ({Path(driver_path).name})"
        check("Microsoft EdgeDriver", driver_status, driver_detail)

        webview_version = self._package_version("pywebview", "webview")
        check("WebView runtime", "PASS" if webview_version != "not installed" else "FAIL", f"pywebview {webview_version}")

        runtime_packages = {
            "pandas": ("pandas", "pandas"),
            "openpyxl": ("openpyxl", "openpyxl"),
            "httpx": ("httpx", "httpx"),
            "selenium": ("selenium", "selenium"),
            "beautifulsoup4": ("beautifulsoup4", "bs4"),
            "extract-msg": ("extract-msg", "extract_msg"),
            "fpdf2": ("fpdf2", "fpdf"),
        }
        for label, (distribution, import_name) in runtime_packages.items():
            version = self._package_version(distribution, import_name)
            check(f"Dependency: {label}", "PASS" if version != "not installed" else "FAIL", version)

        try:
            usage = shutil.disk_usage(Path.home())
            free_gb = usage.free / (1024 ** 3)
            check("Disk space", "PASS" if free_gb >= 1 else "WARN", f"{free_gb:.1f} GB free")
        except OSError as exc:
            check("Disk space", "WARN", str(exc))

        for label, directory in {
            "Download directory": self.default_download_dir,
            "Application data": str(Path.home() / ".coupa_turbo"),
        }.items():
            try:
                Path(directory).mkdir(parents=True, exist_ok=True)
                probe = Path(directory) / ".diagnostic-write-test"
                probe.write_text("ok", encoding="utf-8")
                probe.unlink(missing_ok=True)
                check(label, "PASS", self._safe_user_path(directory))
            except OSError as exc:
                check(label, "FAIL", f"{self._safe_user_path(directory)} — {exc}")

        try:
            db_check = self.db.conn.execute("PRAGMA integrity_check").fetchone()[0]
            check("Session database", "PASS" if db_check == "ok" else "FAIL", str(db_check))
        except Exception as exc:
            check("Session database", "FAIL", str(exc))

        async def network_checks() -> None:
            import httpx
            for label, url in {
                "Coupa connectivity": "https://unilever.coupahost.com",
                "GitHub connectivity": "https://api.github.com",
            }.items():
                try:
                    async with httpx.AsyncClient(
                        timeout=5,
                        follow_redirects=True,
                        verify=system_ssl_context(),
                    ) as client:
                        response = await client.get(url)
                    check(label, "PASS" if response.status_code < 500 else "WARN", f"HTTP {response.status_code}")
                except Exception as exc:
                    check(label, "WARN", str(exc))

        try:
            asyncio.run(network_checks())
        except Exception as exc:
            check("Network diagnostics", "WARN", str(exc))

        if input_path:
            state = self.inspect_input_file(input_path)
            input_name = Path(input_path).name
            if not state.get("exists"):
                check("Selected input", "FAIL", f"{input_name}: file not found")
            elif state.get("open_detected"):
                check("Selected input", "WARN", f"{input_name}: Excel appears to be open")
            elif state.get("ready"):
                check("Selected input", "PASS", f"{input_name}: saved and readable")
            else:
                check("Selected input", "WARN", f"{input_name}: waiting for a stable save")

        passed = sum(item["status"] == "PASS" for item in checks)
        failed = sum(item["status"] == "FAIL" for item in checks)
        warnings = sum(item["status"] == "WARN" for item in checks)
        lines = [
            "COUPA TURBO DOWNLOADER - HOST DIAGNOSTIC REPORT",
            f"Generated: {datetime.now(timezone.utc).isoformat()}",
            "",
            f"Summary: {passed} pass | {warnings} warning(s) | {failed} failure(s)",
            "",
        ]
        for item in checks:
            lines.append(f"[{item['status']}] {item['name']}: {item['detail']}")
        lines.extend([
            "",
            "Privacy: credentials, cookies, PO contents, and full user paths are not included.",
            "Please send this report together with a screenshot of the error and the approximate run time.",
        ])
        report = "\n".join(lines)
        return {"success": True, "report": report, "checks": checks, "summary": {"passed": passed, "warnings": warnings, "failed": failed}}

    def copy_diagnostics_report(self, report: str) -> Dict[str, Any]:
        success, message = self._clipboard_copy(report or "")
        return {"success": success, "message": message}

    def save_diagnostics_report(self, report: str) -> Dict[str, Any]:
        try:
            destination = Path.home() / ".coupa_turbo" / f"diagnostics_{time.strftime('%Y%m%d-%H%M%S')}.txt"
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_text(report or "", encoding="utf-8")
            return {"success": True, "path": str(destination)}
        except Exception as exc:
            return {"success": False, "error": str(exc)}

    def set_auth_cookies(self, cookies: Dict[str, str]) -> None:
        """Store authenticated cookies for use by the crawler."""
        self._cookies = cookies

    def reset_authentication(self) -> Dict[str, Any]:
        """Clear cached Coupa state and any legacy app-owned Edge profile."""
        result = clear_cached_authentication(remove_app_profile=True)
        if result.get("success"):
            self._cookies = None
        return result

    def check_auth(self) -> Dict[str, Any]:
        """Check whether the cached Coupa session is still usable."""
        try:
            cookies = load_cached_cookies()
            if not cookies:
                self._cookies = None
                return {"authenticated": False, "state": "login_required", "message": "No cached session."}
            valid, reason = asyncio.run(validate_cookies_detailed(cookies))
            if valid:
                self._cookies = cookies
                return {"authenticated": True, "state": "cached", "message": "Cached session is valid."}
            if reason == "unavailable":
                # Do not label a valid-looking cache as expired because of a
                # timeout, proxy, or temporary Coupa/network failure.
                self._cookies = cookies
                return {
                    "authenticated": False,
                    "state": "unavailable",
                    "has_cached_session": True,
                    "message": "Could not verify cached session right now; the cached session will be tried during the run.",
                }
            self._cookies = None
            return {"authenticated": False, "state": "expired", "message": "Cached session expired."}
        except Exception as e:
            self._cookies = None
            return {"authenticated": False, "state": "unavailable", "message": str(e)}

    def reset_new_run(self, filepath: str = "") -> Dict[str, Any]:
        """Reset the wizard and remove only a template created by the app."""
        import re

        value = str(filepath or "").strip()
        if not value:
            return {"success": True, "deleted": False, "preserved": False}
        path = Path(value).expanduser().resolve()
        template_dir = (Path.home() / "Documents" / "CoupaTurboDownloader" / "templates").resolve()
        is_generated_template = (
            path.parent == template_dir
            and re.fullmatch(r"input_template_\d{8}-\d{6}\.xlsx", path.name) is not None
        )
        if not is_generated_template:
            return {"success": True, "deleted": False, "preserved": True}
        try:
            path.unlink(missing_ok=True)
            return {"success": True, "deleted": True, "path": str(path)}
        except OSError as exc:
            return {"success": False, "error": f"Could not delete the generated template: {exc}"}

    def generate_input_template(self) -> Dict[str, Any]:
        """Generate a template CSV file ready for the user to populate with PO data.

        Creates a file next to the default input CSV location with a timestamp
        so the user can fill it out without overwriting anything.
        """
        template_dir = Path.home() / "Documents" / "CoupaTurboDownloader" / "templates"
        template_dir.mkdir(parents=True, exist_ok=True)

        timestamp = time.strftime("%Y%m%d-%H%M%S")
        template_path = template_dir / f"input_template_{timestamp}.xlsx"

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        input_headers = [
            "PO_NUMBER", "SUPPLIER", "<|>",
            "Company", "Year", "Quarter", "Business Unit",
        ]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Input"
        sheet.append(input_headers)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(input_headers))}1"

        header_fill = PatternFill("solid", fgColor="173B56")
        separator_fill = PatternFill("solid", fgColor="DCECF5")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        sheet[1][2].fill = separator_fill
        sheet[1][2].font = Font(bold=True, color="173B56")
        widths = [20, 28, 8, 24, 12, 14, 24]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

        instructions = workbook.create_sheet("Instructions")
        instructions.column_dimensions["A"].width = 24
        instructions.column_dimensions["B"].width = 100
        instructions.append(["Coupa Turbo Downloader — input guide", ""])
        instructions.append(["Required fields", "PO_NUMBER and SUPPLIER must be filled for every row."])
        instructions.append(["The <|> separator", "Columns before <|> identify the PO. Columns after <|> define the destination folder hierarchy."])
        instructions.append(["Folder levels", "The first column after <|> is level 1, the next is level 2, and so on. Blank values become Unknown."])
        instructions.append(["Example", "Company=ACME, Year=2026, Quarter=Q3 → ACME/2026/Q3/PO_NUMBER"])
        instructions.append(["Workflow", "Fill the Input sheet, save and close Excel, choose the file in the app, validate, correct any errors, then start."])
        for cell in instructions[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for row in instructions.iter_rows(min_row=2):
            row[0].font = Font(bold=True, color="173B56")
            row[1].alignment = Alignment(wrap_text=True, vertical="top")
            row[0].alignment = Alignment(vertical="top")
        instructions.freeze_panes = "A2"
        workbook.save(template_path)

        template_path_str = str(template_path)
        try:
            self._open_path(template_path)
        except Exception:
            pass

        return {
            "success": True,
            "path": template_path_str,
            "format": "xlsx",
            "message": f"Excel template created at {template_path_str}",
        }

    def validate_input_file(self, filepath: str) -> Dict[str, Any]:
        """Validate a populated CSV file before importing.

        Checks:
          - File exists and is readable
          - Required columns present (PO_NUMBER or equivalent)
          - Supplier/company column present
          - No empty PO numbers
          - No duplicate PO numbers

        Returns a list of errors/warnings so the user can fix them in-place.
        """
        import re
        from pathlib import Path

        path = Path(filepath)
        if not path.exists():
            return {"valid": False, "errors": [f"File not found: {filepath}"], "warnings": [], "file_state": {"ready": False}}

        file_state = self.inspect_input_file(filepath)
        if file_state.get("open_detected"):
            return {
                "valid": False,
                "errors": ["The input file appears to be open in Excel. Save and close it before validating."],
                "warnings": [],
                "file_state": file_state,
            }
        if not file_state.get("readable", True):
            return {"valid": False, "errors": ["The input file is not readable."], "warnings": [], "file_state": file_state}

        try:
            if path.suffix.lower() in {".xlsx", ".xls"}:
                df = pd.read_excel(path, dtype=str)
            else:
                with open(path, encoding="utf-8-sig") as f:
                    sample = f.read(4096)
                sep = ";" if sample.count(";") > sample.count(",") else ","
                try:
                    df = pd.read_csv(path, sep=sep, dtype=str, encoding="utf-8-sig")
                except UnicodeDecodeError:
                    df = pd.read_csv(path, sep=sep, dtype=str, encoding="cp1252")
        except Exception as e:
            return {"valid": False, "errors": [f"Failed to parse input file: {e}"], "warnings": []}

        hierarchy_columns = []
        sep_column = next((column for column in df.columns if str(column).strip() == "<|>"), None)
        if sep_column is not None:
            hierarchy_columns = [str(column) for column in list(df.columns)[list(df.columns).index(sep_column) + 1:]]

        errors: list[str] = []
        warnings: list[str] = []
        fixes: list[Dict[str, Any]] = []
        valid_count = 0

        # Normalize column names for matching
        def norm(col):
            return re.sub(r"[^a-z0-9]+", "", str(col).lower().strip())

        norm_cols = {norm(c): c for c in df.columns}

        # Required: PO column
        po_keys = ["ponumber", "po", "pedido"]
        po_col = None
        for key in po_keys:
            if key in norm_cols:
                po_col = norm_cols[key]
                break
        if not po_col:
            errors.append(
                f"Missing PO Number column. Expected one of: PO_NUMBER, PO, Pedido. "
                f"Found: {list(df.columns)}"
            )

        # Required: company/supplier column
        company_keys = ["legalentity", "companycode", "empresa", "supplier"]
        company_col = None
        for key in company_keys:
            if key in norm_cols:
                company_col = norm_cols[key]
                break
        if not company_col:
            errors.append(
                f"Missing Supplier/Company column. Expected one of: SUPPLIER, LegalEntity, CompanyCode, Empresa. "
                f"Found: {list(df.columns)}"
            )

        if not errors:
            # Validate rows. Empty required values are errors because the user
            # should be able to correct the same file before starting a run.
            values = df[[po_col, company_col]].fillna("").astype(str).apply(lambda col: col.str.strip())
            values = values.replace({"nan": "", "None": ""})
            empty_mask = values[po_col].eq("") | values[company_col].eq("")
            normalized_frame = df.fillna("").astype(str).apply(lambda col: col.str.strip())
            fully_blank_mask = normalized_frame.eq("").all(axis=1)
            # Pandas no longer accepts a Series as a multidimensional indexer
            # for Index.__getitem__. Convert masks to plain NumPy booleans.
            blank_mask = fully_blank_mask.to_numpy(dtype=bool)
            partial_mask = (empty_mask & ~fully_blank_mask).to_numpy(dtype=bool)
            blank_rows = [int(index) + 2 for index in df.index[blank_mask]]
            partial_rows = [int(index) + 2 for index in df.index[partial_mask]]
            if blank_rows:
                errors.append(f"Blank row(s) found: {blank_rows[:10]}")
                fixes.append({
                    "action": "remove_blank_rows",
                    "count": len(blank_rows),
                    "description": f"Remove {len(blank_rows)} completely blank row(s)",
                })
            if partial_rows:
                errors.append(f"Missing PO Number or Supplier on row(s): {partial_rows[:10]}")

            clean = df.loc[~empty_mask].copy()
            duplicates = clean[clean.duplicated(subset=[po_col], keep=False)]
            if not duplicates.empty:
                dup_pos = list(duplicates[po_col].astype(str).unique()[:10])
                errors.append(f"Duplicate PO Number(s): {dup_pos}")
                fixes.append({
                    "action": "remove_duplicate_pos",
                    "count": int(duplicates[po_col].nunique()),
                    "description": "Keep the first row for each duplicated PO Number",
                })

            invalid_pos: list[str] = []
            for value in clean[po_col].astype(str):
                if not (value.upper().startswith(("PO", "PM")) or value.isdigit()):
                    invalid_pos.append(value)
            if invalid_pos:
                warnings.append(
                    f"{len(invalid_pos)} PO(s) have an unusual format (expected PO/PM prefix or digits): "
                    f"{invalid_pos[:5]}"
                )

            valid_count = len(clean.drop_duplicates(subset=[po_col]))
            if valid_count == 0:
                errors.append("No valid PO entries found after cleaning.")

        return {
            "valid": len(errors) == 0,
            "errors": errors,
            "warnings": warnings,
            "total_rows": len(df),
            "valid_po_count": valid_count if not errors else 0,
            "hierarchy_columns": hierarchy_columns,
            "fixes": fixes,
            "file_state": file_state,
        }

    def repair_input_file(self, filepath: str, actions: List[str]) -> Dict[str, Any]:
        """Apply only explicitly requested safe repairs, preserving a backup."""
        path = Path(filepath)
        if not path.exists():
            return {"success": False, "error": "File not found."}
        state = self.inspect_input_file(filepath)
        if state.get("open_detected"):
            return {"success": False, "error": "Save and close the input file before applying repairs."}

        actions = set(actions or [])
        if not actions.intersection({"remove_blank_rows", "remove_duplicate_pos"}):
            return {"success": False, "error": "No supported repair was selected."}

        if path.suffix.lower() == ".xls":
            return {"success": False, "error": "Automatic repair for legacy .xls is unavailable. Save it as .xlsx and try again."}

        timestamp = time.strftime("%Y%m%d-%H%M%S")
        backup = path.with_name(f"{path.stem}.backup_{timestamp}{path.suffix}")
        shutil.copy2(path, backup)
        removed_blank = 0
        removed_duplicates = 0

        def norm(column: Any) -> str:
            import re
            return re.sub(r"[^a-z0-9]+", "", str(column).lower().strip())

        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook
            workbook = load_workbook(path)
            sheet = workbook["Input"] if "Input" in workbook.sheetnames else workbook.active
            headers = [cell.value for cell in sheet[1]]
            normalized = {norm(value): index + 1 for index, value in enumerate(headers) if value is not None}
            po_column = next((normalized[key] for key in ("ponumber", "po", "pedido") if key in normalized), None)
            if not po_column:
                return {"success": False, "error": "PO Number column not found."}

            rows_to_remove: set[int] = set()
            seen_pos: set[str] = set()
            for row_number in range(2, sheet.max_row + 1):
                values = [sheet.cell(row_number, column).value for column in range(1, sheet.max_column + 1)]
                is_blank = all(value is None or str(value).strip() == "" for value in values)
                if is_blank and "remove_blank_rows" in actions:
                    rows_to_remove.add(row_number)
                    removed_blank += 1
                    continue
                po_value = str(sheet.cell(row_number, po_column).value or "").strip()
                if po_value and "remove_duplicate_pos" in actions:
                    if po_value in seen_pos:
                        rows_to_remove.add(row_number)
                        removed_duplicates += 1
                    else:
                        seen_pos.add(po_value)
            for row_number in sorted(rows_to_remove, reverse=True):
                sheet.delete_rows(row_number)
            workbook.save(path)
        else:
            with open(path, encoding="utf-8-sig") as handle:
                sample = handle.read(4096)
            separator = ";" if sample.count(";") > sample.count(",") else ","
            try:
                frame = pd.read_csv(path, sep=separator, dtype=str, encoding="utf-8-sig")
            except UnicodeDecodeError:
                frame = pd.read_csv(path, sep=separator, dtype=str, encoding="cp1252")
            normalized = {norm(value): value for value in frame.columns}
            po_column = next((normalized[key] for key in ("ponumber", "po", "pedido") if key in normalized), None)
            if not po_column:
                return {"success": False, "error": "PO Number column not found."}
            blank_mask = frame.fillna("").astype(str).apply(lambda row: row.str.strip().eq("").all(), axis=1)
            if "remove_blank_rows" in actions:
                removed_blank = int(blank_mask.sum())
                frame = frame.loc[~blank_mask]
            if "remove_duplicate_pos" in actions:
                duplicate_mask = frame.duplicated(subset=[po_column], keep="first")
                removed_duplicates = int(duplicate_mask.sum())
                frame = frame.loc[~duplicate_mask]
            frame.to_csv(path, sep=separator, index=False, encoding="utf-8-sig")

        return {
            "success": True,
            "backup_path": str(backup),
            "removed_blank_rows": removed_blank,
            "removed_duplicate_rows": removed_duplicates,
            "message": "Safe repairs applied. The original file was backed up.",
        }

    def import_file(self, filepath: str) -> Dict[str, Any]:
        import re
        try:
            if not os.path.exists(filepath):
                return {'success': False, 'error': f"File not found: {filepath}"}

            _, ext = os.path.splitext(filepath.lower())
            if ext == '.csv':
                with open(filepath, encoding='utf-8-sig') as f:
                    sample = f.read(4096)
                sep = ';' if sample.count(';') > sample.count(',') else ','
                try:
                    df = pd.read_csv(filepath, sep=sep, dtype=str, encoding='utf-8-sig')
                except UnicodeDecodeError:
                    df = pd.read_csv(filepath, sep=sep, dtype=str, encoding='cp1252')
            elif ext in ['.xls', '.xlsx']:
                df = pd.read_excel(filepath, dtype=str)
            else:
                return {'success': False, 'error': "Unsupported file format. Use XLSX or CSV."}

            def clean_folder_part(value: Any) -> str:
                text = str(value or '').strip()
                if text.lower() in {'', 'nan', 'none'}:
                    text = 'Unknown'
                text = text.replace('/', '_').replace('\\', '_')
                text = '_'.join(text.split())
                text = text.strip('._')
                return text or 'Unknown'

            def extract_hierarchy_columns(frame: pd.DataFrame) -> tuple[list[str], bool]:
                sep_col = None
                for c in frame.columns:
                    if str(c).strip() == '<|>':
                        sep_col = c
                        break
                if sep_col is None:
                    return [], False

                cols = list(frame.columns)
                hierarchy_cols = cols[cols.index(sep_col) + 1:]
                if not hierarchy_cols:
                    return [], False

                for col in hierarchy_cols:
                    series = frame[col].fillna('').astype(str).str.strip()
                    series = series[(series != '') & (series.str.lower() != 'nan')]
                    if not series.empty:
                        return hierarchy_cols, True
                return hierarchy_cols, False

            hierarchy_cols, has_hierarchy_data = extract_hierarchy_columns(df)

            def norm(col):
                return re.sub(r'[^a-z0-9]+', '', str(col).lower().strip())
            norm_cols = {norm(col): col for col in df.columns}

            po_keys = ['ponumber', 'po', 'pedido']
            company_keys = ['legalentity', 'companycode', 'empresa', 'supplier']

            po_col = None
            for key in po_keys:
                if key in norm_cols:
                    po_col = norm_cols[key]
                    break

            company_col = None
            for key in company_keys:
                if key in norm_cols:
                    company_col = norm_cols[key]
                    break

            if not po_col:
                return {'success': False, 'error': f"Could not find PO Number column. Found columns: {list(df.columns)}"}
            if not company_col:
                return {'success': False, 'error': f"Could not find Company Code/Legal Entity/Supplier column. Found columns: {list(df.columns)}"}

            session_id = self.db.create_session(os.path.basename(filepath))

            df = df.dropna(subset=[po_col, company_col])
            df = df.drop_duplicates(subset=[po_col, company_col])
            df = df.sort_values(by=company_col, kind='stable')

            total_pos = 0
            for _, row in df.iterrows():
                po_val = str(row[po_col]).strip()
                company_val = str(row[company_col]).strip()
                if po_val and po_val.lower() != 'nan' and company_val and company_val.lower() != 'nan':
                    if has_hierarchy_data and hierarchy_cols:
                        parts = [clean_folder_part(row.get(col, '')) for col in hierarchy_cols]
                        output_subdir = PurePosixPath(*parts).as_posix()
                    else:
                        output_subdir = company_val

                    self.db.add_po(PODownload(
                        session_id=session_id,
                        po_number=po_val,
                        company_code=company_val,
                        output_subdir=output_subdir,
                        status="PENDING"
                    ))
                    total_pos += 1

            return {
                'success': True,
                'session_id': session_id,
                'total_pos': total_pos
            }

        except Exception as e:
            return {'success': False, 'error': str(e)}

    def _session_summary(self, session_id: int) -> Dict[str, int]:
        cursor = self.db.conn.cursor()
        row = cursor.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN status = 'SUCCESS' THEN 1 ELSE 0 END) AS success,
                SUM(CASE WHEN status IN ('ERROR', 'SKIPPED_VERIFICATION_REQUIRED') THEN 1 ELSE 0 END) AS errors,
                SUM(CASE WHEN status = 'PENDING' THEN 1 ELSE 0 END) AS pending
            FROM po_downloads WHERE session_id = ?
            """,
            (session_id,),
        ).fetchone()
        return {
            "total": int(row["total"] or 0),
            "success": int(row["success"] or 0),
            "errors": int(row["errors"] or 0),
            "pending": int(row["pending"] or 0),
        }

    def get_session_history(self) -> List[Dict[str, Any]]:
        cursor = self.db.conn.cursor()
        cursor.execute("SELECT * FROM sessions ORDER BY id DESC")
        history: list[Dict[str, Any]] = []
        for row in cursor.fetchall():
            item = dict(row)
            stats = self._session_summary(int(item["id"]))
            item.update({
                "total_pos": stats["total"],
                "success_count": stats["success"],
                "error_count": stats["errors"],
                "pending_count": stats["pending"],
            })
            if item["status"] in {"PENDING", "RUNNING"} and stats["total"] > 0 and stats["pending"] == 0:
                if stats["errors"] == 0:
                    item["status"] = "SUCCESS"
                elif stats["success"] > 0:
                    item["status"] = "PARTIAL"
                else:
                    item["status"] = "FAILED"
            history.append(item)
        return history

    def get_session_details(self, session_id: int) -> Dict[str, Any]:
        cursor = self.db.conn.cursor()
        cursor.execute("SELECT * FROM po_downloads WHERE session_id = ?", (session_id,))
        pos = [dict(row) for row in cursor.fetchall()]
        session = self.db.get_session(session_id)
        return {
            'session': session,
            'pos': pos
        }

    def confirm_and_retry_company(self, session_id: int, company_code: str) -> Dict[str, Any]:
        try:
            cursor = self.db.conn.cursor()
            cursor.execute('''
                UPDATE po_downloads
                SET status = 'PENDING', updated_at = strftime('%Y-%m-%d %H:%M:%f', 'now')
                WHERE session_id = ? AND company_code = ? AND status = 'SKIPPED_VERIFICATION_REQUIRED'
            ''', (session_id, company_code))
            self.db.conn.commit()
            return {'success': True}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def export_session_report(self, session_id: int, dest_filepath: str) -> Dict[str, Any]:
        try:
            cursor = self.db.conn.cursor()
            cursor.execute('''
                SELECT po_number, company_code, status, download_folder, attachment_count, error_message, updated_at
                FROM po_downloads
                WHERE session_id = ?
            ''', (session_id,))
            rows = cursor.fetchall()
            df = pd.DataFrame([dict(r) for r in rows])

            df.to_excel(dest_filepath, index=False)
            return {'success': True, 'filepath': dest_filepath}
        except Exception as e:
            return {'success': False, 'error': str(e)}


def _check_runtime_stop(
    session_id: int,
    runtime: Dict[int, Dict[str, Any]],
    lock: threading.Lock,
) -> None:
    with lock:
        rt = runtime.get(session_id)
        if rt and rt.get("stop_requested"):
            rt["status"] = "STOPPED"
            raise SessionStoppedError(session_id)


async def _check_runtime_pause(
    session_id: int,
    runtime: Dict[int, Dict[str, Any]],
    lock: threading.Lock,
) -> None:
    while True:
        with lock:
            rt = runtime.get(session_id)
            if not rt:
                return
            if rt.get("stop_requested"):
                rt["status"] = "STOPPED"
                raise SessionStoppedError(session_id)
            if not rt.get("paused", False):
                return
        await asyncio.sleep(0.2)


class SessionStoppedError(Exception):
    def __init__(self, session_id: int):
        self.session_id = session_id
        super().__init__(f"Session {session_id} stopped by user")
